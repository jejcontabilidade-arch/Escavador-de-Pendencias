import openpyxl
import os
import re

def clean_name(filename):
    # Remove extensions and leading numbers/IDs
    name = os.path.splitext(filename)[0]
    name = re.sub(r'^\d+_(L_ESSENCE)?', '', name)
    name = re.sub(r'_\d+$', '', name)
    name = re.sub(r'SENHA\s*[a-zA-Z0-9@#$_-]+', '', name, flags=re.IGNORECASE)
    name = re.sub(r'senha\s*[a-zA-Z0-9@#$_-]+', '', name, flags=re.IGNORECASE)
    name = name.replace("_", " ")
    name = re.sub(r'\s+', ' ', name).strip()
    return name

def populate_sheet():
    excel_path = r"C:\Certificados_Escavador\Controle_Certificados.xlsx"
    folder = os.path.abspath(os.path.join("certificados", "CONDOMINIO"))
    
    if not os.path.exists(excel_path):
        print(f"Spreadsheet not found at: {excel_path}")
        return
        
    if not os.path.exists(folder):
        print(f"Folder not found at: {folder}")
        return
        
    wb = openpyxl.load_workbook(excel_path)
    ws = wb.active
    
    # Read existing CNPJs
    existing_cnpjs = []
    for r in range(2, ws.max_row + 1):
        cnpj_val = str(ws.cell(row=r, column=3).value or "").strip()
        existing_cnpjs.append("".join(filter(str.isdigit, cnpj_val)))
        
    files = [f for f in os.listdir(folder) if f.lower().endswith(".pfx") or f.lower().endswith(".p12")]
    
    added_count = 0
    cod_start = ws.max_row + 1
    
    for f in sorted(files):
        # Only add files that have 14 digits of CNPJ in filename
        cnpj_match = re.search(r'\d{14}', f)
        if cnpj_match:
            cnpj = cnpj_match.group(0)
            if cnpj not in existing_cnpjs:
                name = clean_name(f)
                ws.append([
                    cod_start,
                    name,
                    cnpj,
                    "CONDOMINIO",
                    "ATIVO",
                    f
                ])
                existing_cnpjs.append(cnpj)
                cod_start += 1
                added_count += 1
                
    wb.save(excel_path)
    print(f"Workbook saved. Added {added_count} condomínios to the local sheet.")

if __name__ == "__main__":
    populate_sheet()
