import openpyxl
import os

def add_clients():
    excel_path = r"C:\Certificados_Escavador\Controle_Certificados.xlsx"
    if not os.path.exists(excel_path):
        print(f"Spreadsheet not found at: {excel_path}")
        return
        
    wb = openpyxl.load_workbook(excel_path)
    ws = wb.active
    
    # Check if they are already added to avoid duplicates
    existing_cnpjs = []
    for r in range(2, ws.max_row + 1):
        cnpj_val = str(ws.cell(row=r, column=3).value or "").strip()
        existing_cnpjs.append("".join(filter(str.isdigit, cnpj_val)))
        
    # Client 1: CONVENCAO PASTEUR
    cnpj1 = "03636388000100"
    if cnpj1 not in existing_cnpjs:
        ws.append([
            3, 
            "CONDOMINIO DO CONJUNTO PASTEUR BLOCOS 1 E 4", 
            cnpj1, 
            "CONDOMINIO", 
            "ATIVO", 
            "173217607_CONDOMINIO_DO_CONJUNTO_PASTEUR_BLOCOS_1_E_4_03636388000100.pfx"
        ])
        print(f"Added CONDOMINIO DO CONJUNTO PASTEUR to sheet.")
        
    # Client 2: SQS 413 BLOCO A
    cnpj2 = "00270866000112"
    if cnpj2 not in existing_cnpjs:
        ws.append([
            4, 
            "CONDOMINIO DO BLOCO A DA SQS 413", 
            cnpj2, 
            "CONDOMINIO", 
            "ATIVO", 
            "CONDOMINIO DO BLOCO A DA SQS 413_00270866000112 SENHA 707070.pfx"
        ])
        print(f"Added CONDOMINIO DO BLOCO A DA SQS 413 to sheet.")
        
    wb.save(excel_path)
    print("Workbook saved successfully.")

if __name__ == "__main__":
    add_clients()
