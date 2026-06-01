import os
import sys
import csv
import time
import json
import datetime
import requests
import re
import shutil
from playwright.sync_api import sync_playwright, TimeoutError
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from cryptography.hazmat.primitives import serialization
from cryptography.hazmat.primitives.serialization import pkcs12

# Função de log
def log(msg, level="INFO"):
    timestamp = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    msg_str = str(msg)
    
    encoding = sys.stdout.encoding or 'utf-8'
    try:
        msg_encoded = msg_str.encode(encoding, errors='replace').decode(encoding)
        print(f"[{timestamp}] [{level}] {msg_encoded}")
    except Exception:
        try:
            print(f"[{timestamp}] [{level}] {msg_str.encode('ascii', errors='replace').decode('ascii')}")
        except Exception:
            pass
            
    log_dir = "logs"
    try:
        os.makedirs(log_dir, exist_ok=True)
        today = datetime.date.today().strftime("%Y-%m-%d")
        with open(os.path.join(log_dir, f"nota_fiscal_xml_{today}.log"), "a", encoding="utf-8") as f:
            f.write(f"[{timestamp}] [{level}] {msg_str}\n")
    except Exception:
        pass

def dialogo_certificado_aberto():
    import ctypes
    EnumWindows = ctypes.windll.user32.EnumWindows
    EnumWindowsProc = ctypes.WINFUNCTYPE(ctypes.c_bool, ctypes.c_int, ctypes.c_int)
    GetWindowText = ctypes.windll.user32.GetWindowTextW
    GetWindowTextLength = ctypes.windll.user32.GetWindowTextLengthW
    IsWindowVisible = ctypes.windll.user32.IsWindowVisible
    
    aberto = [False]
    
    def foreach_window(hwnd, lParam):
        if IsWindowVisible(hwnd):
            length = GetWindowTextLength(hwnd)
            buff = ctypes.create_unicode_buffer(length + 1)
            GetWindowText(hwnd, buff, length + 1)
            title = buff.value
            title_lower = title.lower()
            
            # Blacklist para evitar planilhas, códigos e editores abertos do usuário
            blacklist = ["- excel", "- word", "- notepad", "visual studio", "code", ".py", ".xlsx", ".xls", "cmd.exe", "powershell", "escavador"]
            if any(b in title_lower for b in blacklist):
                return True
                
            termos_cert = ["selecione um certificado", "selecione o certificado", "confirmar certificado", 
                           "select a certificate", "confirm certificate", "segurança do windows", 
                           "windows security", "credenciais de segurança", "security credentials", 
                           "pin do certificado", "insira o pin", "controle de acesso"]
            if any(t in title_lower for t in termos_cert) or (any(x in title_lower for x in ["certificado", "segurança"]) and not any(b in title_lower for b in blacklist)):
                aberto[0] = True
                return False
        return True

    EnumWindows(EnumWindowsProc(foreach_window), 0)
    return aberto[0]

def focar_janela_certificado(keywords_browser):
    import ctypes
    EnumWindows = ctypes.windll.user32.EnumWindows
    EnumWindowsProc = ctypes.WINFUNCTYPE(ctypes.c_bool, ctypes.c_int, ctypes.c_int)
    GetWindowText = ctypes.windll.user32.GetWindowTextW
    GetWindowTextLength = ctypes.windll.user32.GetWindowTextLengthW
    IsWindowVisible = ctypes.windll.user32.IsWindowVisible
    SetForegroundWindow = ctypes.windll.user32.SetForegroundWindow
    ShowWindow = ctypes.windll.user32.ShowWindow
    
    found_hwnds = []
    
    def foreach_window(hwnd, lParam):
        if IsWindowVisible(hwnd):
            length = GetWindowTextLength(hwnd)
            buff = ctypes.create_unicode_buffer(length + 1)
            GetWindowText(hwnd, buff, length + 1)
            title = buff.value
            title_lower = title.lower()
            
            # Blacklist para evitar planilhas, códigos e editores abertos do usuário
            blacklist = ["- excel", "- word", "- notepad", "visual studio", "code", ".py", ".xlsx", ".xls", "cmd.exe", "powershell", "escavador"]
            if any(b in title_lower for b in blacklist):
                return True
                
            termos_cert = ["selecione um certificado", "selecione o certificado", "confirmar certificado", 
                           "select a certificate", "confirm certificate", "segurança do windows", 
                           "windows security", "credenciais de segurança", "security credentials", 
                           "pin do certificado", "insira o pin", "controle de acesso"]
            
            # 1. Diálogos de certificado ou segurança
            if any(t in title_lower for t in termos_cert) or (any(x in title_lower for x in ["certificado", "segurança"]) and not any(b in title_lower for b in blacklist)):
                found_hwnds.append((hwnd, title, 2))
            # 2. Janela do navegador da nossa automação
            elif any(x in title_lower for x in keywords_browser):
                found_hwnds.append((hwnd, title, 1))
        return True

    EnumWindows(EnumWindowsProc(foreach_window), 0)
    
    if not found_hwnds:
        log("[AUTO-LOGIN] Nenhuma janela de certificado ou navegador encontrada para focar.", "WARNING")
        return False
        
    found_hwnds.sort(key=lambda x: x[2], reverse=True)
    target_hwnd, title, priority = found_hwnds[0]
    
    log(f"[AUTO-LOGIN] Janela alvo encontrada (Prioridade {priority}): '{title}'", "SYSTEM")
    try:
        current_active = ctypes.windll.user32.GetForegroundWindow()
        if current_active != target_hwnd:
            # Simula toque no ALT para liberar o privilégio de SetForegroundWindow no Windows
            ctypes.windll.user32.keybd_event(0x12, 0, 0, 0)
            ctypes.windll.user32.keybd_event(0x12, 0, 2, 0)
            ShowWindow(target_hwnd, 9) # SW_RESTORE
            SetForegroundWindow(target_hwnd)
            time.sleep(0.5)
        else:
            log("[AUTO-LOGIN] Janela alvo já está ativa. Ignorando refocus.", "SYSTEM")
        return True
    except Exception as e:
        log(f"[AUTO-LOGIN] Erro ao focar janela: {e}", "WARNING")
        return False

def press_enter(first_char=None):
    import ctypes
    focar_janela_certificado(["nota fiscal", "nfe.fazenda", "chrome", "edge"])
    if first_char:
        char_upper = first_char.upper()
        if len(char_upper) == 1 and (char_upper.isalnum() or char_upper == " "):
            vk_code = ord(char_upper)
            log(f"Enviando tecla '{char_upper}' (VK: {hex(vk_code)}) para focar no certificado...", "SYSTEM")
            ctypes.windll.user32.keybd_event(vk_code, 0, 0, 0) # Key Down
            time.sleep(0.05)
            ctypes.windll.user32.keybd_event(vk_code, 0, 2, 0) # Key Up
            time.sleep(0.5)
            
    log("Enviando comando ENTER para o Windows...", "SYSTEM")
    VK_RETURN = 0x0D
    ctypes.windll.user32.keybd_event(VK_RETURN, 0, 0, 0) # Key Down
    ctypes.windll.user32.keybd_event(VK_RETURN, 0, 2, 0) # Key Up
    log("ENTER enviado.", "SYSTEM")

def executar_confirmacao_certificado_em_loop(config, log_prefix):
    import time
    first_char = config.get("cert_first_char", "J")
    log(f"{log_prefix} Iniciando rotina de confirmacao de certificado SSL (Letra: '{first_char}')...", "SYSTEM")
    
    # Aguarda 3.5 segundos para a janela do Chrome iniciar o redirecionamento e abrir o diálogo
    time.sleep(3.5)
    
    # Realiza 3 tentativas espaçadas de focar e dar ENTER
    for attempt in range(1, 4):
        log(f"{log_prefix} Tentativa {attempt}/3 de confirmacao...", "SYSTEM")
        press_enter(first_char if attempt == 1 else None)
        time.sleep(1.5)
        
    log(f"{log_prefix} Rotina de confirmacao concluida.", "SUCCESS")


# Função para carregar as configurações locais
def load_config():
    config_path = "config.json"
    private_config_path = "config_private.json"
    config = {
        "headless": False,
        "timeout_ms": 30000,
        "relatorios_dir": "relatorios",
        "clientes_file": "clientes.csv",
        "portal_url": "https://www.nfe.fazenda.gov.br/portal/principal.aspx",
        "download_timeout_ms": 60000,
        "whatsapp_enabled": True,
        "whatsapp_number": "",
        "whatsapp_zapi_instance": "",
        "whatsapp_zapi_token": "",
        "whatsapp_zapi_client_token": "",
        "openai_api_key": ""
    }
    if os.path.exists(config_path):
        try:
            with open(config_path, "r", encoding="utf-8") as f:
                dados = json.load(f)
                config.update(dados)
        except Exception:
            pass
    if os.path.exists(private_config_path):
        try:
            with open(private_config_path, "r", encoding="utf-8") as f:
                dados_privados = json.load(f)
                config.update(dados_privados)
        except Exception:
            pass
    return config

# Função para enviar mensagens via WhatsApp
def enviar_whatsapp(mensagem, config, destinatario=None):
    if not config.get("whatsapp_enabled"):
        log("Notificação via WhatsApp desabilitada nas configurações.", "INFO")
        return False
        
    number = destinatario or config.get("whatsapp_number")
    if not number:
        log("Erro: Número de telefone do WhatsApp não configurado.", "ERROR")
        return False
        
    def worker():
        url = "http://127.0.0.1:3000/api/send-message"
        payload = {
            "to": number,
            "message": mensagem
        }
        log(f"Enviando notificação WhatsApp para {number} via gateway local em segundo plano...", "INFO")
        try:
            response = requests.post(url, json=payload, timeout=15)
            if response.status_code in [200, 201]:
                log("Notificação via WhatsApp enviada com sucesso!", "SUCCESS")
            else:
                log(f"Falha ao enviar WhatsApp. Status: {response.status_code}, Resposta: {response.text}", "ERROR")
        except Exception as e:
            log(f"Erro ao enviar requisição HTTP do WhatsApp: {e}", "ERROR")

    import threading
    threading.Thread(target=worker, daemon=True).start()
    return True

def enviar_documento_whatsapp_local(caminho_arquivo, nome_arquivo, config, destinatario):
    if not config.get("whatsapp_enabled") or not destinatario:
        return False
        
    def worker():
        try:
            import base64
            with open(caminho_arquivo, "rb") as f:
                base64_content = base64.b64encode(f.read()).decode("utf-8")
                
            ext = os.path.splitext(nome_arquivo)[1].lower()
            mime = "application/octet-stream"
            if ext == ".pdf":
                mime = "application/pdf"
            elif ext == ".xlsx":
                mime = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            elif ext == ".xml":
                mime = "application/xml"
                
            document_payload = f"data:{mime};base64,{base64_content}"
            
            url = "http://127.0.0.1:3000/api/send-document"
            payload = {
                "to": destinatario,
                "document": document_payload,
                "fileName": nome_arquivo
            }
            
            log(f"[WHATSAPP] Enviando documento '{nome_arquivo}' via gateway local em segundo plano...", "INFO")
            response = requests.post(url, json=payload, timeout=30)
            if response.status_code in [200, 201]:
                log(f"[WHATSAPP] Documento '{nome_arquivo}' enviado com sucesso!", "SUCCESS")
            else:
                log(f"[WHATSAPP] Falha ao enviar documento. Status: {response.status_code}, Resposta: {response.text}", "ERROR")
        except Exception as e:
            log(f"[WHATSAPP] Erro ao enviar documento via gateway: {e}", "ERROR")

    import threading
    threading.Thread(target=worker, daemon=True).start()
    return True

# Limpar o nome da empresa para diretórios seguros
def clean_filename(name):
    return "".join(c if c.isalnum() or c in " _-ÇçÁáÉéÍíÓóÚúÃãÕõÂâÊêÔôÀàÜü" else "_" for c in name).strip()

# Formatar CNPJ
def format_cnpj(cnpj):
    c = "".join(filter(str.isdigit, str(cnpj)))
    if 11 < len(c) <= 14:
        c = c.zfill(14)
        return f"{c[:2]}.{c[2:5]}.{c[5:8]}/{c[8:12]}-{c[12:]}"
    elif 0 < len(c) <= 11:
        c = c.zfill(11)
        return f"{c[:3]}.{c[3:6]}.{c[6:9]}-{c[9:]}"
    return cnpj

# Salvar status do processamento do cliente
def save_client_status_nota_fiscal_xml(cnpj, nome, status, details="", mes_ano=None):
    today = datetime.date.today().strftime("%Y-%m-%d")
    nome_limpo = clean_filename(nome)
    cnpj_limpo = "".join(filter(str.isdigit, cnpj))
    
    if not mes_ano:
        mes_ano = datetime.datetime.now().strftime("%m_%Y")
        
    folder_name = f"XML_{mes_ano}_{cnpj_limpo}_{nome_limpo}"
    client_dir = os.path.join("documentos de nota fiscal xml", folder_name)
    os.makedirs(client_dir, exist_ok=True)
    
    status_path = os.path.join(client_dir, "status_nota_fiscal_xml.json")
    contador_falhas = 0
    
    if os.path.exists(status_path):
        try:
            with open(status_path, "r", encoding="utf-8") as f:
                data = json.load(f)
                if data.get("data_consulta") == today:
                    contador_falhas = data.get("contador_falhas_hoje", 0)
        except Exception:
            pass
            
    if status == "Erro":
        contador_falhas += 1
    elif status == "Sucesso":
        contador_falhas = 0
        
    status_data = {
        "cnpj": cnpj,
        "nome": nome,
        "data_consulta": today,
        "hora_consulta": datetime.datetime.now().strftime("%H:%M:%S"),
        "status": status,
        "detalhes": details,
        "contador_falhas_hoje": contador_falhas
    }
    
    with open(status_path, "w", encoding="utf-8") as f:
        json.dump(status_data, f, indent=4, ensure_ascii=False)
    return client_dir

# Salvar estado global em tempo real
def salvar_estado_global(rodando, empresa_atual, total, processados, sucessos, falhas):
    os.makedirs("temp", exist_ok=True)
    today = datetime.date.today().strftime("%Y-%m-%d")
    state_data = {
        "rodando": rodando,
        "empresa_atual": empresa_atual,
        "total_clientes": total,
        "processados": processados,
        "sucessos": sucessos,
        "falhas": falhas,
        "data_consulta": today
    }
    try:
        with open("temp/state_nota_fiscal_xml.json", "w", encoding="utf-8") as f:
            json.dump(state_data, f, indent=4, ensure_ascii=False)
    except Exception as e:
        log(f"Erro ao salvar state_nota_fiscal_xml.json: {e}", "WARNING")

# Aguardar resolução do CAPTCHA na tela
def esperar_captcha(page, client_name, cnpj, step_name, config, destinatario=None):
    time.sleep(2.0)
    limite_tempo = 300  # 5 minutos no total
    inicio = time.time()
    
    while time.time() - inicio < limite_tempo:
        # 1. Verificar se há CAPTCHA na tela
        has_hcaptcha = False
        has_recaptcha = False
        try:
            has_hcaptcha = page.locator('iframe[src*="hcaptcha"], iframe[title*="hCaptcha"]').count() > 0
        except Exception: pass
        try:
            has_recaptcha = page.locator('iframe[src*="recaptcha"], iframe[title*="recaptcha"]').count() > 0
        except Exception: pass
            
        if not has_hcaptcha and not has_recaptcha:
            log("Nenhum CAPTCHA detectado ou já resolvido.", "SUCCESS")
            return True
            
        # 2. Identificar a frame de desafio (challenge) ativa
        challenge_frame = None
        is_hcaptcha = False
        is_recaptcha = False
        
        # Procura hCaptcha challenge
        for f in page.frames:
            if "frame=challenge" in f.url:
                try:
                    iframe_el = page.locator('iframe[src*="frame=challenge"]').first
                    if iframe_el.count() > 0 and iframe_el.is_visible():
                        challenge_frame = f
                        is_hcaptcha = True
                        break
                except Exception:
                    pass
        
        # Procura reCAPTCHA challenge
        if not challenge_frame:
            for f in page.frames:
                if "bframe" in f.url or "recaptcha/api2/bframe" in f.url:
                    try:
                        iframe_el = page.locator('iframe[src*="bframe"], iframe[src*="recaptcha/api2/bframe"]').first
                        if iframe_el.count() > 0 and iframe_el.is_visible():
                            challenge_frame = f
                            is_recaptcha = True
                            break
                    except Exception:
                        pass
                    
        # Se a frame de desafio não está visível ainda (apenas o checkbox),
        # tenta clicar no checkbox para abrir o desafio
        if not challenge_frame:
            try:
                iframe_check = page.frame_locator('iframe[src*="frame=checkbox"], iframe[src*="api2/anchor"], iframe[title*="caixa de seleção"]')
                checkbox = iframe_check.locator('#checkbox, #anchor, .check, div[role="checkbox"]').first
                if checkbox.is_visible():
                    checkbox.click(force=True)
                    time.sleep(2.0)
                    continue
            except Exception:
                pass
                
        # Se a frame de desafio está aberta, tiramos print e pedimos ajuda ao agente
        if challenge_frame:
            os.makedirs("temp", exist_ok=True)
            screenshot_path = os.path.join("temp", "captcha.png")
            page.screenshot(path=screenshot_path)
            print(f"[AGENTE_ACORDE_CAPTCHA] {client_name} | {cnpj}", flush=True)
            log(f"[CAPTCHA] Desafio detectado. Print salvo em '{screenshot_path}'. Aguardando resolução do Agente...", "WARNING")
            
            # Escreve o arquivo de requisição para o Agente
            req_data = {
                "status": "waiting",
                "client_name": client_name,
                "cnpj": cnpj,
                "type": "hcaptcha" if is_hcaptcha else "recaptcha",
                "screenshot_path": os.path.abspath(screenshot_path)
            }
            req_path = os.path.join("temp", "captcha_request.json")
            with open(req_path, "w", encoding="utf-8") as f_req:
                json.dump(req_data, f_req, indent=4, ensure_ascii=False)
                
            # Limpa resposta anterior se houver
            resp_path = os.path.join("temp", "captcha_response.json")
            if os.path.exists(resp_path):
                try: os.remove(resp_path)
                except: pass
                
            # Aguarda a resposta do Agente (limite de 2 minutos para cada tentativa)
            log("[CAPTCHA] Aguardando arquivo 'temp/captcha_response.json' com os cliques...", "INFO")
            enviar_whatsapp(f"🤖 *Alerta de CAPTCHA! (Mapeamento)*\n\nA automação tirou um print do CAPTCHA em 'temp/captcha.png'. Aguardando resolução do Agente.", config, destinatario)
            
            inicio_espera = time.time()
            resolvido = False
            while time.time() - inicio_espera < 120:
                if os.path.exists(resp_path):
                    try:
                        with open(resp_path, "r", encoding="utf-8") as f_resp:
                            resp_data = json.load(f_resp)
                        click_indices = resp_data.get("click_indices", [])
                        
                        log(f"[CAPTCHA] Resposta recebida do Agente. Clicando nos índices: {click_indices}", "INFO")
                        
                        if is_hcaptcha:
                            cells = challenge_frame.locator('.task-image, .image-wrapper, .image')
                            verify_button = challenge_frame.locator('.button-submit, button:has-text("Verificar"), button:has-text("Verify"), button:has-text("Avançar"), button:has-text("Next")').first
                        else:
                            cells = challenge_frame.locator('.rc-imageselect-tile')
                            verify_button = challenge_frame.locator('#recaptcha-verify-button').first
                            
                        # Clicar nos índices fornecidos com fallback JS
                        for idx in click_indices:
                            if 0 <= idx < cells.count():
                                cell = cells.nth(idx)
                                try:
                                    cell.scroll_into_view_if_needed()
                                    cell.click(force=True)
                                except Exception:
                                    try:
                                        cell.evaluate("el => el.click()")
                                    except Exception:
                                        pass
                                time.sleep(0.5)
                                
                        # Clicar no botão de verificar com scroll e fallback robusto
                        if verify_button.is_visible():
                            try:
                                verify_button.scroll_into_view_if_needed()
                                verify_button.click(force=True)
                            except Exception:
                                try:
                                    verify_button.evaluate("el => el.click()")
                                except Exception as e_ev:
                                    log(f"Erro ao forçar clique via JS no botão de verificar: {e_ev}", "WARNING")
                            
                        time.sleep(3.0)
                        resolvido = True
                        break
                    except Exception as e_resp:
                        log(f"Erro ao processar resposta do captcha: {e_resp}", "WARNING")
                        
                time.sleep(1)
                
            # Limpa arquivos de comunicação da rodada
            try:
                if os.path.exists(req_path): os.remove(req_path)
                if os.path.exists(resp_path): os.remove(resp_path)
            except: pass
            
            if not resolvido:
                log("Timeout aguardando resposta do Agente para esta tentativa de CAPTCHA.", "WARNING")
                
        else:
            time.sleep(2.0)
            
    return False



# Gerar planilha Excel bonita
def gerar_excel_resumo(resultados):
    excel_path = "nota_fiscal_xml.xlsx"
    desktop_path = os.path.join(os.path.expanduser("~"), "Desktop", "nota_fiscal_xml.xlsx")
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Nota Fiscal XML"
    
    ws.views.sheetView[0].showGridLines = True
    
    # 1. Montar o Cabeçalho Corporativo Estético de Topo (linhas 1 a 3)
    ws.merge_cells("A1:O1")
    ws.merge_cells("A2:O2")
    ws.merge_cells("A3:O3")
    
    ws["A1"] = "J&J CONTABILIDADE — RELATÓRIO DE ENTRADA E CONCILIAÇÃO DE NOTAS FISCAIS (NF-e)"
    ws["A2"] = "Varredura de Manifestações de Destinatário e XMLs via Portal Nacional da NF-e"
    
    agora = datetime.datetime.now().strftime("%d/%m/%Y às %H:%M")
    ws["A3"] = f"Relatório emitido em {agora} | Escavador de Pendências Automatizado v1.0"
    
    # Estilização dos banners de topo
    fill_banner = PatternFill(start_color="1B4F72", end_color="1B4F72", fill_type="solid") # Azul escuro profundo
    fill_sub_meta = PatternFill(start_color="F8F9F9", end_color="F8F9F9", fill_type="solid") # Cinza ultra claro
    
    font_banner = Font(name="Segoe UI", size=13, bold=True, color="FFFFFF")
    font_sub = Font(name="Segoe UI", size=10, bold=False, color="2C3E50")
    font_meta = Font(name="Segoe UI", size=9, italic=True, color="7F8C8D")
    
    align_center_v = Alignment(horizontal="center", vertical="center")
    
    # Aplicar cores de fundo a todas as colunas do intervalo mesclado para evitar que fiquem brancas no Excel
    for col_idx in range(1, 16):
        ws.cell(row=1, column=col_idx).fill = fill_banner
        ws.cell(row=2, column=col_idx).fill = fill_sub_meta
        ws.cell(row=3, column=col_idx).fill = fill_sub_meta
        
    ws["A1"].font = font_banner
    ws["A1"].alignment = align_center_v
    
    ws["A2"].font = font_sub
    ws["A2"].alignment = align_center_v
    
    ws["A3"].font = font_meta
    ws["A3"].alignment = align_center_v
    
    # Ajustar alturas das linhas de topo
    ws.row_dimensions[1].height = 35
    ws.row_dimensions[2].height = 20
    ws.row_dimensions[3].height = 18
    ws.row_dimensions[4].height = 12  # Linha 4 vazia para separação visual
    
    # 2. Cabeçalho da Tabela (linha 5)
    headers = [
        "Número NF", "Série", "Data Emissão", 
        "CNPJ Emitente", "Nome Emitente", 
        "CNPJ Destinatário", "Nome Destinatário", 
        "Valor Produtos", "Valor Nota (R$)", 
        "PIS (R$)", "COFINS (R$)", "ICMS (R$)", 
        "Chave de Acesso", "Status de Download", "Data/Hora Consulta"
    ]
    
    font_header = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
    fill_header = PatternFill(start_color="2471A3", end_color="2471A3", fill_type="solid") # Azul médio profissional
    align_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    align_left = Alignment(horizontal="left", vertical="center")
    align_right = Alignment(horizontal="right", vertical="center")
    
    border_side = Side(border_style="thin", color="D3D3D3")
    border_cell = Border(left=border_side, right=border_side, top=border_side, bottom=border_side)
    
    for col_idx, h in enumerate(headers, start=1):
        cell = ws.cell(row=5, column=col_idx, value=h)
        cell.font = font_header
        cell.fill = fill_header
        cell.alignment = align_center
        cell.border = border_cell
        
    ws.row_dimensions[5].height = 26
    
    # 3. Linhas de Dados (linha 6 em diante)
    font_data = Font(name="Segoe UI", size=10)
    
    for idx, r in enumerate(resultados, start=6):
        # Tratar robustamente se por algum motivo vier formato de tupla/lista antiga
        if isinstance(r, (list, tuple)):
            cnpj = r[0] if len(r) > 0 else "N/A"
            nome = r[1] if len(r) > 1 else "N/A"
            chave = r[2] if len(r) > 2 else "N/A"
            status = r[3] if len(r) > 3 else "Pendente"
            dt = r[4] if len(r) > 4 else "-"
            
            r = {
                "numero_nf": "N/A",
                "serie": "N/A",
                "data_emissao": "N/A",
                "cnpj_emitente": "N/A",
                "nome_emitente": "N/A",
                "cnpj_destinatario": format_cnpj(cnpj),
                "nome_destinatario": nome,
                "valor_produtos": 0.0,
                "valor_nota": 0.0,
                "pis": 0.0,
                "cofins": 0.0,
                "icms": 0.0,
                "chave_acesso": chave,
                "status_download": status,
                "data_hora_consulta": dt
            }
            
        row_data = [
            r.get("numero_nf", "N/A"),
            r.get("serie", "N/A"),
            r.get("data_emissao", "N/A"),
            r.get("cnpj_emitente", "N/A"),
            r.get("nome_emitente", "N/A"),
            r.get("cnpj_destinatario", "N/A"),
            r.get("nome_destinatario", "N/A"),
            r.get("valor_produtos", 0.0),
            r.get("valor_nota", 0.0),
            r.get("pis", 0.0),
            r.get("cofins", 0.0),
            r.get("icms", 0.0),
            r.get("chave_acesso", "N/A"),
            r.get("status_download", "Pendente"),
            r.get("data_hora_consulta", "-")
        ]
        
        for col_idx, val in enumerate(row_data, start=1):
            cell = ws.cell(row=idx, column=col_idx, value=val)
            cell.font = font_data
            cell.border = border_cell
            
            # Alinhamentos e formatações específicas
            if col_idx in [1, 2, 3, 4, 6, 13, 14, 15]:
                cell.alignment = align_center
            elif col_idx in [5, 7]:
                cell.alignment = align_left
            elif col_idx in [8, 9, 10, 11, 12]:
                cell.alignment = align_right
                cell.number_format = 'R$ #,##0.00'
                
        status = r.get("status_download", "Pendente")
        fill_status = PatternFill(fill_type=None)
        if status == "Baixado":
            fill_status = PatternFill(start_color="D4EFDF", end_color="D4EFDF", fill_type="solid") # Verde claro
        elif status in ["Não Possui Manifestação", "Sem Registros"]:
            fill_status = PatternFill(start_color="FCF3CF", end_color="FCF3CF", fill_type="solid") # Amarelo claro
        elif "Erro" in str(status):
            fill_status = PatternFill(start_color="FADBD8", end_color="FADBD8", fill_type="solid") # Vermelho claro
            
        ws.cell(row=idx, column=14).fill = fill_status
                
    # Largura das colunas
    ws.column_dimensions['A'].width = 15  # Número NF
    ws.column_dimensions['B'].width = 10  # Série
    ws.column_dimensions['C'].width = 20  # Data Emissão
    ws.column_dimensions['D'].width = 20  # CNPJ Emitente
    ws.column_dimensions['E'].width = 45  # Nome Emitente
    ws.column_dimensions['F'].width = 20  # CNPJ Destinatário
    ws.column_dimensions['G'].width = 45  # Nome Destinatário
    ws.column_dimensions['H'].width = 18  # Valor Produtos
    ws.column_dimensions['I'].width = 18  # Valor Nota
    ws.column_dimensions['J'].width = 15  # PIS
    ws.column_dimensions['K'].width = 15  # COFINS
    ws.column_dimensions['L'].width = 15  # ICMS
    ws.column_dimensions['M'].width = 48  # Chave de Acesso
    ws.column_dimensions['N'].width = 25  # Status de Download
    ws.column_dimensions['O'].width = 22  # Data/Hora Consulta
    
    # Salvar
    try:
        wb.save(excel_path)
        log(f"Planilha local {excel_path} salva com sucesso.", "SUCCESS")
        try:
            shutil.copy(excel_path, desktop_path)
            log(f"Cópia da planilha salva no Desktop em: {desktop_path}", "SUCCESS")
        except Exception as e_desktop:
            log(f"Aviso ao salvar cópia no Desktop: {e_desktop}", "WARNING")
    except Exception as e:
        log(f"Erro ao salvar planilha excel: {e}", "ERROR")

def obter_dados_certificado():
    import glob
    import re
    
    # Valores default para fallback
    cnpj = "05443435000124"
    nome = "J&J SERVICOS PROFISSIONAIS LTDA"
    first_char = "J"
    
    pfx_files = glob.glob("*.pfx")
    if pfx_files:
        filename = os.path.basename(pfx_files[0])
        log(f"Arquivo de certificado encontrado: '{filename}'", "INFO")
        
        # Detecta o CNPJ de 14 dígitos
        match_cnpj = re.search(r'\d{14}', filename)
        if match_cnpj:
            cnpj = match_cnpj.group(0)
            
            if cnpj == "05443435000124":
                # Força o nome do certificado conhecido para unificar pastas (J_J SERVICOS PROFISSIONAIS LTDA)
                nome = "J&J SERVICOS PROFISSIONAIS LTDA"
                first_char = "J"
            else:
                # Detecta o nome antes do CNPJ
                part_before = filename[:match_cnpj.start()]
                part_before = re.sub(r'^\d+_', '', part_before)
                part_before = part_before.strip('_')
                if part_before:
                    nome = part_before.replace('_', ' ')
                    first_char = nome[0].upper()
                
    log(f"Dados do Certificado Ativo: Nome='{nome}', CNPJ='{cnpj}', Caractere Inicial='{first_char}'", "SUCCESS")
    return cnpj, nome, first_char

def merge_resultados(anteriores, hoje):
    chaves_hoje = set(r[2] for r in hoje if r[2] and r[2] != "N/A")
    resultados_finais = []
    for r in anteriores:
        chave_ant = r[2]
        if chave_ant and chave_ant != "N/A" and chave_ant in chaves_hoje:
            continue
        resultados_finais.append(r)
    resultados_finais.extend(hoje)
    return resultados_finais

def converter_xml_nfe_para_json(caminho_xml, caminho_json):
    import xml.etree.ElementTree as ET
    import json
    import re
    
    def xml_para_dicionario(elemento):
        tag = elemento.tag.split('}')[-1] if '}' in elemento.tag else elemento.tag
        dicionario = {}
        
        if elemento.attrib:
            dicionario["@atributos"] = {k.split('}')[-1]: v for k, v in elemento.attrib.items()}
            
        filhos = list(elemento)
        if filhos:
            grupos_de_filhos = {}
            for filho in filhos:
                tag_filho = filho.tag.split('}')[-1] if '}' in filho.tag else filho.tag
                if tag_filho not in grupos_de_filhos:
                    grupos_de_filhos[tag_filho] = []
                grupos_de_filhos[tag_filho].append(filho)
                
            for tag_filho, grupo in grupos_de_filhos.items():
                tags_repetiveis = ["det", "dup", "pag", "vol", "reboque", "lacres", "obsCont", "obsFisco"]
                if len(grupo) > 1 or tag_filho in tags_repetiveis:
                    dicionario[tag_filho] = [xml_para_dicionario(c) for c in grupo]
                else:
                    dicionario[tag_filho] = xml_para_dicionario(grupo[0])
        else:
            texto = elemento.text.strip() if elemento.text else ""
            if texto:
                if re.match(r'^-?\d+\.\d+$', texto):
                    try:
                        return float(texto)
                    except ValueError:
                        pass
                elif re.match(r'^-?\d+$', texto):
                    try:
                        return int(texto)
                    except ValueError:
                        pass
                return texto
            else:
                return None if not elemento.attrib else dicionario
                
        return dicionario

    try:
        arvore = ET.parse(caminho_xml)
        raiz = arvore.getroot()
        dados_nota = xml_para_dicionario(raiz)
        
        with open(caminho_json, "w", encoding="utf-8") as f:
            json.dump(dados_nota, f, indent=4, ensure_ascii=False)
        log(f"Dados estruturados em JSON salvos com sucesso: '{os.path.basename(caminho_json)}'", "SUCCESS")
        return True
    except Exception as e:
        log(f"Erro ao converter XML {os.path.basename(caminho_xml)} para JSON: {e}", "ERROR")
        return False

def extrair_dados_fiscais_nfe(caminho_json):
    import json
    dados_fiscais = {
        "numero_nf": "N/A",
        "serie": "N/A",
        "data_emissao": "N/A",
        "cnpj_emitente": "N/A",
        "nome_emitente": "N/A",
        "cnpj_destinatario": "N/A",
        "nome_destinatario": "N/A",
        "valor_produtos": 0.0,
        "valor_nota": 0.0,
        "pis": 0.0,
        "cofins": 0.0,
        "icms": 0.0
    }
    
    if not os.path.exists(caminho_json):
        return dados_fiscais
        
    try:
        with open(caminho_json, "r", encoding="utf-8") as f:
            dados = json.load(f)
            
        inf = dados.get("NFe", {}).get("infNFe", {})
        if not inf:
            inf = dados.get("infNFe", {})
            
        if inf:
            ide = inf.get("ide", {})
            emit = inf.get("emit", {})
            dest = inf.get("dest", {})
            icms_tot = inf.get("total", {}).get("ICMSTot", {})
            
            dados_fiscais["numero_nf"] = ide.get("nNF", "N/A")
            dados_fiscais["serie"] = ide.get("serie", "N/A")
            
            dh_emi = ide.get("dhEmi") or ide.get("dEmi") or "N/A"
            if dh_emi and dh_emi != "N/A" and "T" in str(dh_emi):
                try:
                    partes = str(dh_emi).split("T")
                    data = partes[0]
                    hora = partes[1].split("-")[0].split("+")[0]
                    dh_emi = f"{data} {hora}"
                except Exception:
                    pass
            dados_fiscais["data_emissao"] = dh_emi
            
            cnpj_emit = emit.get("CNPJ") or emit.get("CPF") or "N/A"
            dados_fiscais["cnpj_emitente"] = format_cnpj(str(cnpj_emit)) if cnpj_emit != "N/A" else "N/A"
            dados_fiscais["nome_emitente"] = emit.get("xNome", "N/A")
            
            cnpj_dest = dest.get("CNPJ") or dest.get("CPF") or "N/A"
            dados_fiscais["cnpj_destinatario"] = format_cnpj(str(cnpj_dest)) if cnpj_dest != "N/A" else "N/A"
            dados_fiscais["nome_destinatario"] = dest.get("xNome", "N/A")
            
            def converter_float(val):
                if val is None or val == "":
                    return 0.0
                try:
                    return float(val)
                except ValueError:
                    return 0.0
                    
            dados_fiscais["valor_produtos"] = converter_float(icms_tot.get("vProd"))
            dados_fiscais["valor_nota"] = converter_float(icms_tot.get("vNF"))
            dados_fiscais["pis"] = converter_float(icms_tot.get("vPIS"))
            dados_fiscais["cofins"] = converter_float(icms_tot.get("vCOFINS"))
            dados_fiscais["icms"] = converter_float(icms_tot.get("vICMS"))
            
    except Exception as e:
        log(f"Erro ao extrair dados fiscais do JSON {os.path.basename(caminho_json)}: {e}", "WARNING")
        
    return dados_fiscais

def limpar_cache_perfil_chrome(user_data_dir, cnpj=None):
    import subprocess
    import time
    try:
        filter_str = f"*chrome_profile_{cnpj}*" if cnpj else "*chrome_profile_nota_fiscal_xml*"
        cmd = f'powershell -NoProfile -Command "Get-CimInstance Win32_Process -Filter \\"Name = \'chrome.exe\' or Name = \'chromedriver.exe\'\\" | Where-Object {{ $_.CommandLine -like \'{filter_str}\' }} | ForEach-Object {{ Stop-Process -Id $_.ProcessId -Force }}"'
        subprocess.run(cmd, shell=True, capture_output=True, timeout=5)
    except Exception:
        pass
        
    if os.path.exists(user_data_dir):
        for _ in range(3):
            try:
                shutil.rmtree(user_data_dir)
                log(f"[CACHE] Limpeza completa realizada para o perfil: '{os.path.basename(user_data_dir)}'", "SUCCESS")
                break
            except Exception:
                time.sleep(0.5)

# Helpers adicionais para classificação mensal e modernização do certificado
def extrair_mes_ano_emissao(xml_path):
    import xml.etree.ElementTree as ET
    try:
        tree = ET.parse(xml_path)
        root = tree.getroot()
        for elem in root.iter():
            if '}' in elem.tag:
                elem.tag = elem.tag.split('}', 1)[1]
                
        ide = root.find(".//ide")
        if ide is not None:
            dh_emi = ide.find("dhEmi")
            if dh_emi is not None and dh_emi.text:
                date_str = dh_emi.text.strip()
                partes = date_str.split("T")[0].split("-")
                if len(partes) >= 2:
                    return f"{partes[1]}_{partes[0]}"
            d_emi = ide.find("dEmi")
            if d_emi is not None and d_emi.text:
                date_str = d_emi.text.strip()
                partes = date_str.split("-")
                if len(partes) >= 2:
                    return f"{partes[1]}_{partes[0]}"
    except Exception as e:
        log(f"Erro ao extrair data de emissão do XML: {e}", "WARNING")
    now = datetime.datetime.now()
    return now.strftime("%m_%Y")

def buscar_xml_existente_por_chave(chave):
    base_dir = "documentos de nota fiscal xml"
    if not os.path.exists(base_dir):
        return None
    for root, dirs, files in os.walk(base_dir):
        for f in files:
            if chave in f and f.endswith(".xml"):
                return os.path.join(root, f)
    return None

def cliente_possui_xml_existente(cnpj_clean):
    base_dir = "documentos de nota fiscal xml"
    if not os.path.exists(base_dir):
        return False
    try:
        for d in os.listdir(base_dir):
            if cnpj_clean in d:
                d_path = os.path.join(base_dir, d)
                if os.path.isdir(d_path):
                    xmls = [f for f in os.listdir(d_path) if f.endswith(".xml")]
                    if xmls:
                        return True
    except Exception:
        pass
    return False

def buscar_pfx_recursivo(caminho_dir, cnpj_limpo):
    if not os.path.exists(caminho_dir):
        return None, None
    for root, dirs, files in os.walk(caminho_dir):
        for f in files:
            if cnpj_limpo in f and (f.lower().endswith(".pfx") or f.lower().endswith(".p12")):
                return os.path.join(root, f), f
    return None, None

def modernizar_pfx_se_necessario(caminho_pfx, senha):
    os.makedirs(os.path.join("temp", "modern_certs"), exist_ok=True)
    pfx_basename = os.path.basename(caminho_pfx)
    modern_pfx_path = os.path.join("temp", "modern_certs", f"modern_{pfx_basename}")
    
    try:
        with open(caminho_pfx, "rb") as f:
            pfx_data = f.read()
        pass_bytes = senha.encode('utf-8') if senha else None
        
        private_key, certificate, additional_certificates = pkcs12.load_key_and_certificates(
            pfx_data, pass_bytes
        )
        
        modern_pfx_data = pkcs12.serialize_key_and_certificates(
            name=b"cert",
            key=private_key,
            cert=certificate,
            cas=additional_certificates,
            encryption_algorithm=serialization.BestAvailableEncryption(pass_bytes) if pass_bytes else serialization.NoEncryption()
        )
        
        with open(modern_pfx_path, "wb") as f:
            f.write(modern_pfx_data)
        log(f"[MODERNIZE] Certificado PFX re-criptografado com sucesso para formato moderno em: '{modern_pfx_path}'", "SUCCESS")
        return os.path.abspath(modern_pfx_path)
    except Exception as e:
        log(f"[MODERNIZE] Erro ao modernizar certificado '{caminho_pfx}': {e}. Usando o original.", "WARNING")
        return os.path.abspath(caminho_pfx)

def sincronizar_e_instalar_certificados_condominios():
    log("[MULTICLIENTE] Iniciando sincronização e instalação de certificados para Condomínios...", "SYSTEM")
    config = load_config()
    
    path_servidor = config.get("caminho_certificados_servidor", "")
    path_local = config.get("caminho_certificados_local", "")
    path_interno = os.path.abspath("certificados")
    senha_padrao = config.get("senha_padrao_pfx", "123456")
    
    caminho_ativo = None
    if path_servidor and os.path.exists(path_servidor):
        caminho_ativo = path_servidor
        log(f"[MULTICLIENTE] Usando pasta de certificados do servidor: '{caminho_ativo}'", "INFO")
    elif path_local and os.path.exists(path_local):
        caminho_ativo = path_local
        log(f"[MULTICLIENTE] Servidor inacessível. Usando pasta local de contingência: '{caminho_ativo}'", "INFO")
    elif os.path.exists(path_interno):
        caminho_ativo = path_interno
        log(f"[MULTICLIENTE] Servidor e contingência indisponíveis. Usando pasta interna do robô: '{caminho_ativo}'", "INFO")
    else:
        log(f"[MULTICLIENTE] Erro: Nenhum diretório de certificados válido ou acessível ({path_servidor}, {path_local} ou {path_interno}).", "ERROR")
        return []
        
    excel_path = os.path.join(caminho_ativo, "Controle_Certificados.xlsx")
    if not os.path.exists(excel_path):
        log(f"[MULTICLIENTE] Erro: Planilha de controle '{excel_path}' não localizada.", "ERROR")
        return []
        
    # 1. Obter certificados instalados na máquina do usuário via PowerShell
    certificados_instalados = []
    try:
        import subprocess
        cmd = "powershell -Command \"Get-ChildItem Cert:\\CurrentUser\\My | Select-Object -Property Subject, Thumbprint | ConvertTo-Json\""
        res = subprocess.run(cmd, capture_output=True, text=True, shell=True)
        if res.returncode == 0 and res.stdout.strip():
            try:
                data = json.loads(res.stdout.strip())
                if isinstance(data, dict):
                    certificados_instalados.append(data)
                elif isinstance(data, list):
                    certificados_instalados.extend(data)
            except Exception as e_json:
                log(f"[MULTICLIENTE] Aviso ao parsear JSON dos certificados instalados: {e_json}", "WARNING")
    except Exception as e_cmd:
        log(f"[MULTICLIENTE] Erro ao obter certificados via PowerShell: {e_cmd}", "WARNING")
        
    subjects_instalados = [c.get("Subject", "") for c in certificados_instalados if c.get("Subject")]
    
    # 2. Ler clientes.csv
    clientes_file = config.get("clientes_file", "clientes.csv")
    clientes_existentes = {}
    if os.path.exists(clientes_file):
        try:
            with open(clientes_file, "r", encoding="utf-8-sig") as f:
                reader = csv.DictReader(f)
                for row in reader:
                    c_cnpj = "".join(filter(str.isdigit, row.get("cnpj", "")))
                    if c_cnpj:
                        clientes_existentes[c_cnpj] = {
                            "nome": row.get("nome", "").strip(),
                            "ativo": row.get("ativo", "True").strip()
                        }
        except Exception as e_csv:
            log(f"[MULTICLIENTE] Erro ao ler clientes.csv: {e_csv}", "ERROR")
            
    # 3. Ler planilha Excel
    condominios_sincronizados = []
    try:
        wb = openpyxl.load_workbook(excel_path, data_only=True)
        ws = wb.active
        
        headers = [str(ws.cell(row=1, column=c).value).strip().upper() for c in range(1, ws.max_column + 1)]
        
        col_cnpj = -1
        col_cliente = -1
        col_tipo = -1
        col_arquivo = -1
        
        for idx, h in enumerate(headers, start=1):
            if "CNPJ" in h or "DADOS" in h:
                col_cnpj = idx
            elif "CLIENTE" in h or "NOME" in h or "IDENTIFIC" in h:
                col_cliente = idx
            elif "TIPO" in h:
                col_tipo = idx
            elif "ARQUIVO" in h or "CERTIFICADO" in h:
                col_arquivo = idx
                
        if col_cnpj == -1 or col_cliente == -1 or col_tipo == -1:
            log("[MULTICLIENTE] Cabeçalhos não identificados por completo. Usando mapeamento padrão.", "WARNING")
            col_tipo = 1
            col_cliente = 2
            col_cnpj = 3
            col_arquivo = 6
            
        for row_idx in range(2, ws.max_row + 1):
            tipo_val = str(ws.cell(row=row_idx, column=col_tipo).value or "").strip().upper()
            if "CONDOM" in tipo_val:
                cnpj_val = str(ws.cell(row=row_idx, column=col_cnpj).value or "").strip()
                nome_val = str(ws.cell(row=row_idx, column=col_cliente).value or "").strip()
                arquivo_val = str(ws.cell(row=row_idx, column=col_arquivo).value or "").strip() if col_arquivo != -1 else ""
                
                cnpj_limpo = "".join(filter(str.isdigit, cnpj_val))
                if not cnpj_limpo:
                    continue
                cnpj_limpo = cnpj_limpo.zfill(14)
                
                log(f"[MULTICLIENTE] Condomínio detectado na planilha: '{nome_val}' - CNPJ: {cnpj_limpo}", "INFO")
                
                if cnpj_limpo not in clientes_existentes:
                    log(f"[MULTICLIENTE] Adicionando novo condomínio ao clientes.csv: {nome_val} (CNPJ: {cnpj_limpo})", "SUCCESS")
                    try:
                        with open(clientes_file, "a", newline="", encoding="utf-8") as f_a:
                            writer = csv.writer(f_a)
                            writer.writerow([cnpj_limpo, nome_val, "True"])
                        clientes_existentes[cnpj_limpo] = {"nome": nome_val, "ativo": "True"}
                    except Exception as e_wcsv:
                        log(f"[MULTICLIENTE] Erro ao gravar novo cliente no CSV: {e_wcsv}", "ERROR")
                
                certificado_instalado = False
                for subject in subjects_instalados:
                    if cnpj_limpo in subject:
                        certificado_instalado = True
                        break
                        
                caminho_pfx_encontrado = None
                pfx_filename = None
                
                pastas_busca = [caminho_ativo]
                if path_local and path_local not in pastas_busca:
                    pastas_busca.append(path_local)
                if path_interno and path_interno not in pastas_busca:
                    pastas_busca.append(path_interno)
                
                # 1. Tentar encontrar pelo nome especificado na planilha
                if arquivo_val and arquivo_val != "None":
                    for pasta in pastas_busca:
                        if not os.path.exists(pasta):
                            continue
                        p_file = os.path.join(pasta, arquivo_val)
                        if os.path.exists(p_file):
                            caminho_pfx_encontrado = p_file
                            pfx_filename = os.path.basename(p_file)
                            break
                        else:
                            for root, dirs, files in os.walk(pasta):
                                if arquivo_val in files:
                                    caminho_pfx_encontrado = os.path.join(root, arquivo_val)
                                    pfx_filename = arquivo_val
                                    break
                            if caminho_pfx_encontrado:
                                break
                                
                # 2. Se não achou, buscar pelo CNPJ de forma recursiva nas pastas
                if not caminho_pfx_encontrado:
                    for pasta in pastas_busca:
                        caminho_pfx_encontrado, pfx_filename = buscar_pfx_recursivo(pasta, cnpj_limpo)
                        if caminho_pfx_encontrado:
                            break
                    
                if not certificado_instalado:
                    log(f"[MULTICLIENTE] Certificado para CNPJ {cnpj_limpo} não está instalado. Verificando PFX encontrado...", "INFO")
                    if caminho_pfx_encontrado:
                        senha_encontrada = senha_padrao
                        match_senha = re.search(r'senha\s*([a-zA-Z0-9@#$_-]+)', pfx_filename, re.IGNORECASE)
                        if match_senha:
                            senha_encontrada = match_senha.group(1)
                            log(f"[MULTICLIENTE] Senha extraída do nome do arquivo PFX: '{senha_encontrada}'", "SUCCESS")
                            
                        try:
                            senha_esc = senha_encontrada.replace("'", "''")
                            import subprocess
                            cmd_import = f"powershell -Command \"$p = ConvertTo-SecureString '{senha_esc}' -AsPlainText -Force; Import-PfxCertificate -FilePath '{caminho_pfx_encontrado}' -CertStoreLocation Cert:\\CurrentUser\\My -Password $p\""
                            res_import = subprocess.run(cmd_import, capture_output=True, text=True, shell=True)
                            if res_import.returncode == 0:
                                log(f"[MULTICLIENTE] Certificado instalado com sucesso para CNPJ {cnpj_limpo}: '{pfx_filename}'", "SUCCESS")
                                certificado_instalado = True
                            else:
                                log(f"[MULTICLIENTE] Erro ao importar PFX via PowerShell: {res_import.stderr}", "ERROR")
                        except Exception as e_imp:
                            log(f"[MULTICLIENTE] Exceção ao executar comando de importação: {e_imp}", "ERROR")
                    else:
                        log(f"[MULTICLIENTE] Erro: Arquivo PFX para o condomínio CNPJ {cnpj_limpo} não foi localizado.", "ERROR")
                else:
                    log(f"[MULTICLIENTE] Certificado para CNPJ {cnpj_limpo} já está instalado no Windows.", "SUCCESS")
                    
                if caminho_pfx_encontrado:
                    senha_encontrada = senha_padrao
                    match_senha = re.search(r'senha\s*([a-zA-Z0-9@#$_-]+)', pfx_filename, re.IGNORECASE)
                    if match_senha:
                        senha_encontrada = match_senha.group(1)
                        
                    condominios_sincronizados.append({
                        "cnpj": cnpj_limpo,
                        "nome": nome_val,
                        "pfx_path": os.path.abspath(caminho_pfx_encontrado),
                        "senha": senha_encontrada
                    })
                else:
                    log(f"[MULTICLIENTE] Condomínio '{nome_val}' (CNPJ: {cnpj_limpo}) será ignorado por falta de arquivo PFX.", "WARNING")
    except Exception as e_xl:
        log(f"[MULTICLIENTE] Erro ao ler planilha Excel de condomínios: {e_xl}", "ERROR")
        
    log(f"[MULTICLIENTE] Sincronização concluída. Total de condomínios prontos: {len(condominios_sincronizados)}", "SUCCESS")
    return condominios_sincronizados

def executar_varredura_nfe_para_cliente(page, config, cnpj_cliente, nome_cliente, erros_hoje, destinatario=None):
    today = datetime.date.today().strftime("%Y-%m-%d")
    cnpj_clean = "".join(filter(str.isdigit, cnpj_cliente))
    nome_limpo = clean_filename(nome_cliente)
    
    possui_xml_existente = cliente_possui_xml_existente(cnpj_clean)
    log(f"--- Iniciando Varredura para: {nome_cliente} (CNPJ: {cnpj_cliente}) ---", "INFO")
    
    current_mes_ano = datetime.datetime.now().strftime("%m_%Y")
    save_client_status_nota_fiscal_xml(cnpj_cliente, nome_cliente, "Pendente", "Iniciado", current_mes_ano)
    
    sucessos = 0
    falhas = 0
    total_chaves = 0
    chaves_baixadas = 0
    
    try:
        log("Acessando portal da NF-e...", "INFO")
        page.goto("https://www.nfe.fazenda.gov.br/portal/principal.aspx")
        page.wait_for_load_state("networkidle")
        
        log("Navegando para 'Manifestação Destinatário'...", "INFO")
        link_selector = 'a[href*="manifestacao"]:visible, a:has-text("Manifestação Destinatário"):visible, a:has-text("Manifestacao Destinatario"):visible'
        page.locator(link_selector).first.click()
        
        try:
            page.wait_for_load_state("networkidle", timeout=10000)
        except Exception:
            pass
            
        # Verifica se o portal reportou erro de certificado digital expirado ou inválido
        conteudo_visivel_init = ""
        try:
            conteudo_visivel_init = page.locator("body").inner_text()
        except Exception:
            pass
            
        if "certificado digital" in conteudo_visivel_init.lower() and ("expirado" in conteudo_visivel_init.lower() or "inválido" in conteudo_visivel_init.lower() or "invalido" in conteudo_visivel_init.lower()):
            raise Exception("Certificado digital expirado ou inválido no portal da NF-e")
            
        try:
            # Procura de forma robusta pelo radio button ou label correspondente usando seletores válidos
            opcao_nao_tenho = page.locator(
                'input#ctl00_ContentPlaceHolder1_rbtSemChave, '
                'input[value="rbtSemChave"], '
                'label[for*="rbtSemChave"], '
                'label:has-text("Não tenho a Chave de Acesso"), '
                'label:has-text("Nao tenho a Chave"), '
                'label:has-text("Não tenho a Chave")'
            ).first
            
            opcao_nao_tenho.wait_for(state="visible", timeout=15000)
            log("Selecionando opção 'Não tenho a Chave de Acesso'...", "INFO")
            opcao_nao_tenho.click()
            time.sleep(1.5)
        except Exception as e_opt:
            log(f"Aviso ao selecionar opção 'Não tenho a Chave': {e_opt}", "WARNING")
            
        log("Verificando campos de CNPJ...", "INFO")
        try:
            page.wait_for_selector('input[id*="iptCNPJBase"]', timeout=15000)
        except Exception as e_wait_cnpj:
            log(f"Aviso ao aguardar campo CNPJ: {e_wait_cnpj}", "WARNING")
            
        cnpj_base_loc = page.locator('input#ctl00_ContentPlaceHolder1_iptCNPJBase, input[name*="iptCNPJBase"]')
        cnpj_fim_loc = page.locator('input#ctl00_ContentPlaceHolder1_iptCNPJ, input[name="ctl00$ContentPlaceHolder1$iptCNPJ"]')
        
        is_readonly = False
        cnpj_preenchido = ""
        
        if cnpj_base_loc.is_visible(timeout=5000):
            is_readonly = (
                cnpj_base_loc.is_disabled() or 
                cnpj_base_loc.get_attribute("readonly") == "readonly" or 
                cnpj_base_loc.get_attribute("disabled") == "disabled" or
                "aspNetDisabled" in (cnpj_base_loc.get_attribute("class") or "")
            )
            val_base = cnpj_base_loc.input_value() or ""
            val_fim = cnpj_fim_loc.input_value() or ""
            cnpj_preenchido = "".join(filter(str.isdigit, val_base + val_fim))
            
        if is_readonly:
            log(f"Campos de CNPJ estão bloqueados/somente-leitura. CNPJ preenchido no portal: '{cnpj_preenchido}'", "INFO")
            if cnpj_preenchido == cnpj_clean:
                log("O CNPJ preenchido corresponde ao certificado ativo. Prosseguindo...", "SUCCESS")
            else:
                msg_erro = f"Certificado ativo no portal ({cnpj_preenchido}) é diferente do configurado ({cnpj_clean}) e o campo está bloqueado."
                log(f"[AVISO] {msg_erro}", "WARNING")
                raise Exception(msg_erro)
        else:
            log("Campos de CNPJ editáveis. Preenchendo CNPJ do certificado...", "INFO")
            cnpj_inputs = page.locator('input[name*="CNPJ"], input[name*="Cnpj"], input[id*="CNPJ"]').all()
            if len(cnpj_inputs) == 1:
                cnpj_inputs[0].fill(cnpj_clean)
            elif len(cnpj_inputs) >= 2:
                cnpj_inputs[0].fill(cnpj_clean[:8])
                cnpj_inputs[1].fill(cnpj_clean[8:])
            else:
                raise Exception("Campos de CNPJ não localizados na tela.")
                
        captcha_ok = esperar_captcha(page, nome_cliente, cnpj_cliente, "Manifestação", config, destinatario)
        if not captcha_ok:
            raise Exception("Timeout ou falha na resolução do CAPTCHA na Manifestação")
            
        log("Clicando em Pesquisar...", "INFO")
        btn_pesquisar_selector = 'input[value="Pesquisar"], input[name*="Pesquisar"], input[name*="btnPesquisar"], button:has-text("Pesquisar"), input[name*="btnConsultar"]'
        page.locator(btn_pesquisar_selector).first.click()
        
        # Espera o carregamento inicial da página após o clique
        try:
            page.wait_for_load_state("domcontentloaded", timeout=10000)
        except Exception:
            pass
            
        # Loop para aguardar dinamicamente que o resultado da pesquisa seja renderizado na tela
        # (seja uma chave de acesso, a mensagem de 'sem registros' ou alguma mensagem de erro do portal)
        limite_busca = 15
        inicio_busca = time.time()
        chaves = []
        conteudo_visivel = ""
        conteudo_html = ""
        
        log("Aguardando os resultados da pesquisa serem carregados na tela...", "INFO")
        while time.time() - inicio_busca < limite_busca:
            conteudo_html = page.content()
            conteudo_visivel = page.locator("body").inner_text()
            
            # 1. Procurar chaves de 44 dígitos
            chaves = list(dict.fromkeys(re.findall(r'(?<!\d)\d{44}(?!\d)', conteudo_html)))
            
            # 2. Verificar se há mensagens conhecidas
            if chaves:
                break
            if "Não existe registro para os dados informados" in conteudo_visivel:
                break
            if "Não existem registros para os dados informados" in conteudo_visivel:
                break
            if "Nenhum registro encontrado" in conteudo_visivel:
                break
            if "Nenhum documento localizado para o destinatario" in conteudo_visivel:
                break
            if "Nenhum documento localizado para o destinatário" in conteudo_visivel:
                break
            if "Erro" in conteudo_visivel and "Pesquisar" not in conteudo_visivel:
                break
                
            time.sleep(1.0)
            
        if any(msg in conteudo_visivel for msg in [
            "Não existe registro para os dados informados",
            "Não existem registros para os dados informados",
            "Nenhum registro encontrado",
            "Nenhum documento localizado para o destinatario",
            "Nenhum documento localizado para o destinatário"
        ]):
            log(f"Nenhum registro de manifestação encontrado para {nome_cliente}.", "INFO")
            save_client_status_nota_fiscal_xml(cnpj_cliente, nome_cliente, "Sucesso", "Sem Registros", current_mes_ano)
            return (0, 0, 0)
            
        if not chaves:
            log("Aviso: Chave de acesso não localizada na página.", "WARNING")
            msg_erro = "Chave de acesso não localizada na página."
            if "Erro" in conteudo_visivel:
                msg_erro = "Erro reportado pelo portal na consulta de manifestação."
            raise Exception(msg_erro)
            
        total_chaves = len(chaves)
        log(f"Localizada(s) {total_chaves} chave(s) de manifestação na tela para {nome_cliente}. Iniciando downloads...", "SUCCESS")
        
        for idx_chave, chave in enumerate(chaves):
            log(f"--- {nome_cliente}: Processando Chave {idx_chave+1}/{total_chaves}: {chave} ---", "INFO")
            
            caminho_xml_existente = buscar_xml_existente_por_chave(chave)
            if caminho_xml_existente:
                log(f"Chave {chave} já foi baixada anteriormente em: '{caminho_xml_existente}'. Pulando consulta...", "SUCCESS")
                dir_existente = os.path.dirname(caminho_xml_existente)
                caminho_json_esperado = os.path.join(dir_existente, f"{chave}.json")
                if not os.path.exists(caminho_json_esperado):
                    log(f"Metadados JSON ausentes para a chave {chave}. Gerando localmente a partir do XML...", "INFO")
                    converter_xml_nfe_para_json(caminho_xml_existente, caminho_json_esperado)
                chaves_baixadas += 1
                sucessos += 1
                continue
                
            page.goto("https://www.nfe.fazenda.gov.br/portal/principal.aspx")
            page.wait_for_load_state("domcontentloaded")
            
            page.locator('a[href*="tipoConsulta=resumo"]:visible, a:has-text("Consultar NF-e"):visible').first.click()
            page.wait_for_load_state("domcontentloaded")
            
            input_chave_selector = 'input[name="ctl00$ContentPlaceHolder1$txtChaveAcessoResumo"]'
            page.wait_for_selector(input_chave_selector)
            page.locator(input_chave_selector).fill(chave)
            
            captcha_nfe_ok = esperar_captcha(page, nome_cliente, cnpj_cliente, f"Consulta Chave {idx_chave+1}", config, destinatario)
            if not captcha_nfe_ok:
                log(f"Falha de captcha na chave {chave}. Pulando...", "ERROR")
                erros_hoje[chave] = ("Erro no Captcha", datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S"))
                falhas += 1
                continue
                
            log("Clicando em Continuar para carregar a nota...", "INFO")
            page.locator('input[name="ctl00$ContentPlaceHolder1$btnConsultarHCaptcha"]').first.click()
            page.wait_for_load_state("domcontentloaded")
            
            try:
                seletor_download = "input#ctl00_ContentPlaceHolder1_btnDownload, input#btnDownload, a#btnDownload, #btnDownload, input[value*='Download'], button:has-text('Download')"
                page.wait_for_selector(seletor_download, timeout=15000)
                
                log("Botão de Download localizado. Clicando...", "INFO")
                with page.expect_download(timeout=45000) as download_info:
                    page.locator(seletor_download).first.click()
                    
                download = download_info.value
                nome_arquivo = download.suggested_filename
                
                temp_xml_path = os.path.join("temp", nome_arquivo)
                os.makedirs("temp", exist_ok=True)
                download.save_as(temp_xml_path)
                
                mes_ano = extrair_mes_ano_emissao(temp_xml_path)
                
                pasta_destino = os.path.join("documentos de nota fiscal xml", f"XML_{mes_ano}_{cnpj_clean}_{nome_limpo}")
                os.makedirs(pasta_destino, exist_ok=True)
                
                caminho_salvar = os.path.join(pasta_destino, nome_arquivo)
                if os.path.exists(caminho_salvar):
                    os.remove(caminho_salvar)
                shutil.move(temp_xml_path, caminho_salvar)
                
                log(f"Nota salva com sucesso em: {caminho_salvar}", "SUCCESS")
                
                caminho_json = os.path.join(pasta_destino, f"{chave}.json")
                converter_xml_nfe_para_json(caminho_salvar, caminho_json)
                
                if destinatario:
                    enviar_documento_whatsapp_local(caminho_salvar, nome_arquivo, config, destinatario)
                
                chaves_baixadas += 1
                sucessos += 1
                
            except Exception as e_dl:
                log(f"Falha ao realizar download da chave {chave} para {nome_cliente}: {e_dl}", "ERROR")
                erros_hoje[chave] = ("Erro no Download", datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S"))
                falhas += 1
                
        if chaves_baixadas > 0:
            save_client_status_nota_fiscal_xml(cnpj_cliente, nome_cliente, "Sucesso", f"Baixados {chaves_baixadas} chaves", current_mes_ano)
        else:
            forcar_todos = "--forcar-todos" in sys.argv
            if not forcar_todos and possui_xml_existente:
                log(f"[RECOVERY] Nenhuma chave nova baixada para {nome_cliente}, mas mantendo os XMLs existentes.", "SUCCESS")
                save_client_status_nota_fiscal_xml(cnpj_cliente, nome_cliente, "Sucesso", "Preservado (Erro no download das novas chaves)", current_mes_ano)
            else:
                save_client_status_nota_fiscal_xml(cnpj_cliente, nome_cliente, "Erro", "Chaves encontradas, mas nenhuma baixada", current_mes_ano)
                
    except Exception as err_global:
        try:
            os.makedirs("temp", exist_ok=True)
            screenshot_path = os.path.join("temp", f"erro_{cnpj_clean}.png")
            page.screenshot(path=screenshot_path)
            log(f"Screenshot do erro salvo em: '{screenshot_path}'", "INFO")
        except Exception as e_snap:
            log(f"Não foi possível salvar screenshot do erro: {e_snap}", "WARNING")
            
        forcar_todos = "--forcar-todos" in sys.argv
        if not forcar_todos and possui_xml_existente:
            log(f"[RECOVERY] Falha no fluxo global de consulta para {nome_cliente}, mas mantendo os XMLs existentes.", "SUCCESS")
            save_client_status_nota_fiscal_xml(cnpj_cliente, nome_cliente, "Sucesso", "Preservado (Erro na nova consulta)", current_mes_ano)
        else:
            log(f"Erro crítico no processamento de {nome_cliente}: {err_global}", "ERROR")
            save_client_status_nota_fiscal_xml(cnpj_cliente, nome_cliente, "Erro", str(err_global), current_mes_ano)
            
    return (sucessos, falhas, total_chaves)

def compilar_resultados_fiscais(erros_hoje=None, cnpj_cert_ativo=None):
    if erros_hoje is None:
        erros_hoje = {}
        
    config = load_config()
    clientes_file = config.get("clientes_file", "clientes.csv")
    
    clientes_ativos = []
    if os.path.exists(clientes_file):
        try:
            with open(clientes_file, "r", encoding="utf-8-sig") as f:
                reader = csv.DictReader(f)
                for row in reader:
                    ativo_val = row.get("ativo", "True").strip().lower() in ["true", "1", "yes", "ativo", "sim", "s"]
                    if ativo_val:
                        clientes_ativos.append(row)
        except Exception as e:
            log(f"Erro ao ler clientes.csv para compilação: {e}", "ERROR")
            
    resultados = []
    today = datetime.date.today().strftime("%Y-%m-%d")
    chaves_incluidas = set()
    cnpj_cert_limpo = "".join(filter(str.isdigit, cnpj_cert_ativo)) if cnpj_cert_ativo else ""
    
    base_dir = "documentos de nota fiscal xml"
    
    for c in clientes_ativos:
        cnpj = c.get("cnpj", "")
        nome = c.get("nome", "")
        cnpj_limpo = "".join(filter(str.isdigit, cnpj))
        if not cnpj_limpo:
            continue
            
        pastas_cliente = []
        if os.path.exists(base_dir):
            try:
                for d in os.listdir(base_dir):
                    d_path = os.path.join(base_dir, d)
                    if os.path.isdir(d_path):
                        if cnpj_limpo in d:
                            pastas_cliente.append(d_path)
            except Exception as e_list:
                log(f"Erro ao listar diretórios para CNPJ {cnpj_limpo}: {e_list}", "WARNING")
                
        tem_notas = False
        latest_status_data = None
        
        for client_dir in pastas_cliente:
            status_path = os.path.join(client_dir, "status_nota_fiscal_xml.json")
            if os.path.exists(status_path):
                try:
                    with open(status_path, "r", encoding="utf-8") as f_st:
                        st_data = json.load(f_st)
                        if not latest_status_data:
                            latest_status_data = st_data
                        else:
                            dt_curr = f"{st_data.get('data_consulta', '')} {st_data.get('hora_consulta', '')}"
                            dt_lat = f"{latest_status_data.get('data_consulta', '')} {latest_status_data.get('hora_consulta', '')}"
                            if dt_curr > dt_lat:
                                latest_status_data = st_data
                except Exception as e_st:
                    log(f"Erro ao ler status JSON em {client_dir}: {e_st}", "WARNING")
                    
            try:
                arquivos = os.listdir(client_dir)
                arquivos_xml = [f for f in arquivos if f.endswith(".xml")]
                for xml_f in arquivos_xml:
                    match_key = re.search(r'(?<!\d)\d{44}(?!\d)', xml_f)
                    if match_key:
                        chave = match_key.group(0)
                        caminho_xml = os.path.join(client_dir, xml_f)
                        caminho_json = os.path.join(client_dir, f"{chave}.json")
                        if not os.path.exists(caminho_json):
                            converter_xml_nfe_para_json(caminho_xml, caminho_json)
                            
                arquivos = os.listdir(client_dir)
                arquivos_nfe = [f for f in arquivos if f.endswith(".json") and f != "status_nota_fiscal_xml.json"]
                
                for f in arquivos_nfe:
                    match_key = re.search(r'(?<!\d)\d{44}(?!\d)', f)
                    if match_key:
                        chave = match_key.group(0)
                        if chave in chaves_incluidas:
                            continue
                            
                        caminho_json = os.path.join(client_dir, f)
                        dados_fiscais = extrair_dados_fiscais_nfe(caminho_json)
                        
                        dt_consulta = f"{today} 00:00:00"
                        if latest_status_data:
                            dt_consulta = f"{latest_status_data.get('data_consulta', today)} {latest_status_data.get('hora_consulta', '00:00:00')}"
                            
                        registro = {
                            "numero_nf": dados_fiscais.get("numero_nf", "N/A"),
                            "serie": dados_fiscais.get("serie", "N/A"),
                            "data_emissao": dados_fiscais.get("data_emissao", "N/A"),
                            "cnpj_emitente": dados_fiscais.get("cnpj_emitente", "N/A"),
                            "nome_emitente": dados_fiscais.get("nome_emitente", "N/A"),
                            "cnpj_destinatario": dados_fiscais.get("cnpj_destinatario", format_cnpj(cnpj)),
                            "nome_destinatario": dados_fiscais.get("nome_destinatario", nome),
                            "valor_produtos": dados_fiscais.get("valor_produtos", 0.0),
                            "valor_nota": dados_fiscais.get("valor_nota", 0.0),
                            "pis": dados_fiscais.get("pis", 0.0),
                            "cofins": dados_fiscais.get("cofins", 0.0),
                            "icms": dados_fiscais.get("icms", 0.0),
                            "chave_acesso": chave,
                            "status_download": "Baixado",
                            "data_hora_consulta": dt_consulta
                        }
                        resultados.append(registro)
                        chaves_incluidas.add(chave)
                        tem_notas = True
            except Exception as e_dir:
                log(f"Erro ao ler diretório de notas {client_dir}: {e_dir}", "WARNING")
                
        if cnpj_cert_limpo and cnpj_limpo == cnpj_cert_limpo and erros_hoje:
            for chave, (err_status, err_dt) in erros_hoje.items():
                if chave not in chaves_incluidas:
                    registro = {
                        "numero_nf": "N/A",
                        "serie": "N/A",
                        "data_emissao": "N/A",
                        "cnpj_emitente": "N/A",
                        "nome_emitente": "N/A",
                        "cnpj_destinatario": format_cnpj(cnpj),
                        "nome_destinatario": nome,
                        "valor_produtos": 0.0,
                        "valor_nota": 0.0,
                        "pis": 0.0,
                        "cofins": 0.0,
                        "icms": 0.0,
                        "chave_acesso": chave,
                        "status_download": err_status,
                        "data_hora_consulta": err_dt
                    }
                    resultados.append(registro)
                    chaves_incluidas.add(chave)
                    tem_notas = True
                    
        if not tem_notas:
            status = "Pendente"
            detalhes = "Não Consultado"
            dt_completa = f"{today} 00:00:00"
            
            if latest_status_data:
                status = latest_status_data.get("status", "Pendente")
                detalhes = latest_status_data.get("detalhes", "Não Consultado")
                dt_completa = f"{latest_status_data.get('data_consulta', today)} {latest_status_data.get('hora_consulta', '00:00:00')}"
                
            status_exibicao = status
            if status == "Sucesso":
                if "chaves" in detalhes or "Baixados" in detalhes:
                    status_exibicao = "Baixado"
                else:
                    status_exibicao = "Sem Registros"
            elif status == "Erro":
                status_exibicao = f"Erro: {detalhes}"
                
            registro = {
                "numero_nf": "N/A",
                "serie": "N/A",
                "data_emissao": "N/A",
                "cnpj_emitente": "N/A",
                "nome_emitente": "N/A",
                "cnpj_destinatario": format_cnpj(cnpj),
                "nome_destinatario": nome,
                "valor_produtos": 0.0,
                "valor_nota": 0.0,
                "pis": 0.0,
                "cofins": 0.0,
                "icms": 0.0,
                "chave_acesso": "N/A",
                "status_download": status_exibicao,
                "data_hora_consulta": dt_completa
            }
            resultados.append(registro)
            
    return resultados

def main():
    config = load_config()
    
    destinatario = None
    if "--destinatario" in sys.argv:
        try:
            idx = sys.argv.index("--destinatario")
            destinatario = sys.argv[idx + 1]
        except Exception:
            pass
            
    cliente_filtro = None
    if "--cliente" in sys.argv:
        try:
            idx = sys.argv.index("--cliente")
            cliente_filtro = sys.argv[idx + 1]
        except Exception:
            pass
            
    # Execução Multicliente Condomínios
    if "--condominios" in sys.argv:
        condominios = sincronizar_e_instalar_certificados_condominios()
        if not condominios:
            log("[MULTICLIENTE] Nenhum condomínio com certificado válido encontrado para processar.", "WARNING")
            salvar_estado_global(False, "Nenhum condomínio encontrado", 1, 0, 0, 0)
            return
            
        if cliente_filtro:
            filtro_limpo = "".join(filter(str.isdigit, cliente_filtro))
            cliente_filtro_lower = cliente_filtro.strip().lower()
            if filtro_limpo:
                condominios = [c for c in condominios if filtro_limpo in "".join(filter(str.isdigit, c["cnpj"]))]
            else:
                condominios = [c for c in condominios if cliente_filtro_lower in c["nome"].lower()]
                
            if not condominios:
                msg_aviso = f"❌ Nenhum condomínio correspondente ao filtro '{cliente_filtro}' foi localizado para consulta de Notas Fiscais."
                log(f"[MULTICLIENTE] {msg_aviso}", "WARNING")
                if destinatario:
                    enviar_whatsapp(msg_aviso, config, destinatario)
                salvar_estado_global(False, f"Filtro '{cliente_filtro}' não encontrado", 1, 0, 0, 0)
                return
            else:
                log(f"[MULTICLIENTE] Filtro '{cliente_filtro}' aplicado. Processando apenas {len(condominios)} condomínio(s): {[c['nome'] for c in condominios]}", "SUCCESS")
            
        total_clientes = len(condominios)
        log(f"[MULTICLIENTE] Iniciando varredura para {total_clientes} condomínio(s)...", "SYSTEM")
        salvar_estado_global(True, "Iniciando Condomínios...", total_clientes, 0, 0, 0)
        
        sucessos_totais = 0
        falhas_totais = 0
        
        for idx_cond, cond in enumerate(condominios, start=1):
            log(f"\n=======================================================", "SYSTEM")
            log(f"Processando Cliente {idx_cond}/{total_clientes}: {cond['nome']} (CNPJ: {cond['cnpj']})", "SYSTEM")
            log(f"=======================================================", "SYSTEM")
            salvar_estado_global(True, f"Cliente {idx_cond} de {total_clientes}: {cond['nome']}", total_clientes, idx_cond-1, sucessos_totais, falhas_totais)
            
            modern_pfx = modernizar_pfx_se_necessario(cond["pfx_path"], cond["senha"])
            
            client_certs = [
                {"origin": "https://www.nfe.fazenda.gov.br", "pfxPath": modern_pfx, "passphrase": cond["senha"]},
                {"origin": "https://nfe.fazenda.gov.br", "pfxPath": modern_pfx, "passphrase": cond["senha"]},
                {"origin": "https://cav.receita.fazenda.gov.br", "pfxPath": modern_pfx, "passphrase": cond["senha"]},
                {"origin": "https://sso.acesso.gov.br", "pfxPath": modern_pfx, "passphrase": cond["senha"]}
            ]
            
            with sync_playwright() as p:
                user_data_dir = os.path.join(os.getcwd(), "temp", f"chrome_profile_{cond['cnpj']}")
                limpar_cache_perfil_chrome(user_data_dir, cond['cnpj'])
                os.makedirs(user_data_dir, exist_ok=True)
                
                lock_file = os.path.join(user_data_dir, "SingletonLock")
                if os.path.exists(lock_file):
                    try: os.remove(lock_file)
                    except: pass
                    
                log(f"Abrindo Google Chrome nativo para {cond['nome']}...", "SYSTEM")
                try:
                    context = p.chromium.launch_persistent_context(
                        user_data_dir=user_data_dir,
                        headless=False,
                        channel="chrome",
                        client_certificates=client_certs,
                        args=[
                            "--disable-blink-features=AutomationControlled",
                            "--disable-infobars",
                            "--disable-session-crashed-bubble",
                            "--disable-features=BubbleSessionCrashedBubble"
                        ],
                        no_viewport=True
                    )
                    page = context.new_page()
                    page.set_default_timeout(30000)
                    page.on("dialog", lambda dialog: (log(f"Diálogo aceito automaticamente: '{dialog.message}'", "SYSTEM"), dialog.accept()))
                    
                    erros_hoje_cond = {}
                    s_c, f_c, t_c = executar_varredura_nfe_para_cliente(
                        page, config, cond["cnpj"], cond["nome"], erros_hoje_cond, destinatario
                    )
                    
                    sucessos_totais += s_c
                    falhas_totais += f_c
                    
                    context.close()
                    
                except Exception as e_launch:
                    log(f"Erro crítico no navegador para {cond['nome']}: {e_launch}", "ERROR")
                    save_client_status_nota_fiscal_xml(cond["cnpj"], cond["nome"], "Erro", f"Erro no navegador: {e_launch}")
                    falhas_totais += 1
                    
        log("\n[MULTICLIENTE] Finalizado processamento de todos os condomínios. Compilando relatório consolidado...", "SUCCESS")
        resultados_finais = compilar_resultados_fiscais()
        gerar_excel_resumo(resultados_finais)
        salvar_estado_global(False, "Varredura Concluída", total_clientes, total_clientes, sucessos_totais, falhas_totais)
        
        if cliente_filtro and destinatario:
            msg_final = (
                f"🤖 *Consulta de Nota Fiscal XML Concluída!*\n\n"
                f"• *Cliente:* {condominios[0]['nome']}\n"
                f"• *Notas Baixadas:* {sucessos_totais}\n"
                f"• *Falhas:* {falhas_totais}\n\n"
                f"Os arquivos XML das novas notas já foram enviados no seu chat."
            )
            enviar_whatsapp(msg_final, config, destinatario)
        
    else:
        # Modo normal (individual / contador)
        cnpj_cert, nome_cert, cert_first_char = obter_dados_certificado()
        log(f"Iniciando processo de Consulta de Nota Fiscal XML para o certificado {nome_cert} (CNPJ: {cnpj_cert})", "SYSTEM")
        salvar_estado_global(True, "Inicializando Navegador...", 1, 0, 0, 0)
        
        pfx_path = None
        import glob
        pfx_files = glob.glob("*.pfx")
        if pfx_files:
            pfx_path = os.path.abspath(pfx_files[0])
            
        client_certs = []
        if pfx_path:
            pfx_filename = os.path.basename(pfx_path)
            senha_padrao = config.get("senha_padrao_pfx", "123456")
            senha_encontrada = senha_padrao
            match_senha = re.search(r'senha\s*([a-zA-Z0-9@#$_-]+)', pfx_filename, re.IGNORECASE)
            if match_senha:
                senha_encontrada = match_senha.group(1)
                
            modern_pfx = modernizar_pfx_se_necessario(pfx_path, senha_encontrada)
            client_certs = [
                {"origin": "https://www.nfe.fazenda.gov.br", "pfxPath": modern_pfx, "passphrase": senha_encontrada},
                {"origin": "https://nfe.fazenda.gov.br", "pfxPath": modern_pfx, "passphrase": senha_encontrada},
                {"origin": "https://cav.receita.fazenda.gov.br", "pfxPath": modern_pfx, "passphrase": senha_encontrada},
                {"origin": "https://sso.acesso.gov.br", "pfxPath": modern_pfx, "passphrase": senha_encontrada}
            ]
            
        if not client_certs:
            log("Nenhum arquivo PFX local modernizado. Usando thread de foco do Windows...", "WARNING")
            import threading
            threading.Thread(
                target=executar_confirmacao_certificado_em_loop,
                args=(config, "[AUTO-LOGIN-NFE]"),
                daemon=True
            ).start()
            
        with sync_playwright() as p:
            user_data_dir = os.path.join(os.getcwd(), "temp", "chrome_profile_nota_fiscal_xml")
            limpar_cache_perfil_chrome(user_data_dir, None)
            os.makedirs(user_data_dir, exist_ok=True)
            
            lock_file = os.path.join(user_data_dir, "SingletonLock")
            if os.path.exists(lock_file):
                try: os.remove(lock_file)
                except: pass
                
            log("Abrindo Google Chrome nativo em modo visível...", "SYSTEM")
            try:
                context = p.chromium.launch_persistent_context(
                    user_data_dir=user_data_dir,
                    headless=False,
                    channel="chrome",
                    client_certificates=client_certs if client_certs else None,
                    args=[
                        "--disable-blink-features=AutomationControlled",
                        "--disable-infobars",
                        "--disable-session-crashed-bubble",
                        "--disable-features=BubbleSessionCrashedBubble"
                    ],
                    no_viewport=True
                )
                page = context.new_page()
                page.set_default_timeout(30000)
                page.on("dialog", lambda dialog: (log(f"Diálogo do navegador aceito automaticamente: '{dialog.message}'", "SYSTEM"), dialog.accept()))
                
                erros_hoje = {}
                s_c, f_c, t_c = executar_varredura_nfe_para_cliente(
                    page, config, cnpj_cert, nome_cert, erros_hoje, destinatario
                )
                
                context.close()
                
            except Exception as e_launch:
                log(f"Erro crítico ao abrir navegador: {e_launch}", "ERROR")
                salvar_estado_global(False, f"Erro ao iniciar Chrome: {e_launch}", 1, 0, 0, 0)
                return
                
        resultados_finais = compilar_resultados_fiscais(erros_hoje=erros_hoje, cnpj_cert_ativo=cnpj_cert)
        gerar_excel_resumo(resultados_finais)
        salvar_estado_global(False, "Varredura Concluída", len(resultados_finais) or 1, len(resultados_finais) or 1, s_c, f_c)

if __name__ == "__main__":
    main()

