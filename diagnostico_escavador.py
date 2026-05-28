import os
import re
import csv
import json
import openpyxl

def log(msg, level="INFO"):
    print(f"[{level}] {msg}")

def test_diagnostico():
    # 1. Carregar configurações
    config_path = "config.json"
    config = {}
    if os.path.exists(config_path):
        try:
            with open(config_path, "r", encoding="utf-8") as f:
                config = json.load(f)
        except Exception:
            pass
            
    path_servidor = config.get("caminho_certificados_servidor", r"\\Srvjej\contabilidade 2023\8 - CERTIFICADOS")
    path_local = config.get("caminho_certificados_local", r"C:\Certificados_Escavador")
    path_interno = os.path.abspath("certificados")
    senha_padrao = config.get("senha_padrao_pfx", "123456")
    
    log("=== DIAGNÓSTICO DO DIRETÓRIO DE CERTIFICADOS E CONDOMÍNIOS ===", "SYSTEM")
    
    # 2. Verificar acessibilidade das pastas
    caminho_ativo = None
    if os.path.exists(path_servidor):
        caminho_ativo = path_servidor
        log(f"Pasta do Servidor encontrada e ACESSÍVEL: '{path_servidor}'", "SUCCESS")
    elif os.path.exists(path_local):
        caminho_ativo = path_local
        log(f"Pasta Local de Contingência encontrada e ACESSÍVEL: '{path_local}'", "SUCCESS")
    elif os.path.exists(path_interno):
        caminho_ativo = path_interno
        log(f"Pasta Interna do Robô encontrada e ACESSÍVEL: '{path_interno}'", "SUCCESS")
    else:
        log(f"Pasta do Servidor NÃO ACESSÍVEL: '{path_servidor}'", "WARNING")
        log(f"Pasta Local de Contingência NÃO ACESSÍVEL: '{path_local}'", "WARNING")
        log(f"Pasta Interna do Robô ('certificados') NÃO ACESSÍVEL: '{path_interno}'", "ERROR")
        log("Crie a pasta 'certificados' na raiz do robô e coloque a planilha e os certificados lá.", "INFO")
        return
            
    # 3. Localizar planilha
    excel_path = os.path.join(caminho_ativo, "Controle_Certificados.xlsx")
    log(f"Buscando planilha Controle_Certificados.xlsx em: '{excel_path}'", "INFO")
    if not os.path.exists(excel_path):
        log(f"Planilha não localizada em '{excel_path}'", "ERROR")
        return
        
    log("Planilha localizada com sucesso! Lendo dados...", "SUCCESS")
    
    # 4. Ler planilha e mapear colunas
    try:
        wb = openpyxl.load_workbook(excel_path, data_only=True)
        ws = wb.active
        log(f"Aba ativa da planilha: '{ws.title}' | Total de linhas: {ws.max_row}", "INFO")
        
        headers = [str(ws.cell(row=1, column=c).value).strip().upper() for c in range(1, ws.max_column + 1)]
        log(f"Cabeçalhos detectados: {headers}", "INFO")
        
        col_cnpj = -1
        col_cliente = -1
        col_tipo = -1
        col_arquivo = -1
        
        for idx, h in enumerate(headers, start=1):
            if "CNPJ" in h or "DADOS" in h:
                col_cnpj = idx
            elif "CLIENTE" in h or "NOME" in h or "IDENTIFIC" in h:
                col_cliente = idx
            elif "TIPO" in h:
                col_tipo = idx
            elif "ARQUIVO" in h or "CERTIFICADO" in h:
                col_arquivo = idx
                
        log(f"Mapeamento de colunas realizado:", "SYSTEM")
        log(f"  - Coluna TIPO: {col_tipo} ('{headers[col_tipo-1]}' se encontrado)", "INFO")
        log(f"  - Coluna CLIENTE: {col_cliente} ('{headers[col_cliente-1]}' se encontrado)", "INFO")
        log(f"  - Coluna CNPJ: {col_cnpj} ('{headers[col_cnpj-1]}' se encontrado)", "INFO")
        log(f"  - Coluna ARQUIVO: {col_arquivo} ('{headers[col_arquivo-1]}' se encontrado)", "INFO")
        
        if col_cnpj == -1 or col_cliente == -1 or col_tipo == -1:
            log("Erro: Não foi possível identificar as colunas necessárias na planilha.", "ERROR")
            return
            
        # 5. Iterar e verificar cada condomínio
        condominios_encontrados = 0
        condominios_com_pfx = 0
        
        print("\n" + "="*80)
        print(f"{'IDENTIFICAÇÃO DO CONDOMÍNIO':<45} | {'CNPJ':<15} | {'PFX STATUS'}")
        print("="*80)
        
        for row_idx in range(2, ws.max_row + 1):
            tipo_val = str(ws.cell(row=row_idx, column=col_tipo).value or "").strip().upper()
            if "CONDOM" in tipo_val:
                cnpj_val = str(ws.cell(row=row_idx, column=col_cnpj).value or "").strip()
                nome_val = str(ws.cell(row=row_idx, column=col_cliente).value or "").strip()
                arquivo_val = str(ws.cell(row=row_idx, column=col_arquivo).value or "").strip() if col_arquivo != -1 else ""
                
                cnpj_limpo = "".join(filter(str.isdigit, cnpj_val))
                if not cnpj_limpo:
                    continue
                cnpj_limpo = cnpj_limpo.zfill(14)
                
                condominios_encontrados += 1
                
                # Buscar PFX
                caminho_pfx_encontrado = None
                pfx_filename = None
                
                pastas_busca = [caminho_ativo]
                if path_local and path_local not in pastas_busca:
                    pastas_busca.append(path_local)
                if path_interno and path_interno not in pastas_busca:
                    pastas_busca.append(path_interno)
                
                # 1. Nome especificado
                if arquivo_val and arquivo_val != "None":
                    for pasta in pastas_busca:
                        if not os.path.exists(pasta):
                            continue
                        p_file = os.path.join(pasta, arquivo_val)
                        if os.path.exists(p_file):
                            caminho_pfx_encontrado = p_file
                            pfx_filename = os.path.basename(p_file)
                            break
                        else:
                            for root, dirs, files in os.walk(pasta):
                                if arquivo_val in files:
                                    caminho_pfx_encontrado = os.path.join(root, arquivo_val)
                                    pfx_filename = arquivo_val
                                    break
                            if caminho_pfx_encontrado:
                                break
                                
                # 2. Busca recursiva por CNPJ
                if not caminho_pfx_encontrado:
                    for pasta in pastas_busca:
                        if not os.path.exists(pasta):
                            continue
                        for root, dirs, files in os.walk(pasta):
                            found = False
                            for f in files:
                                if cnpj_limpo in f and (f.lower().endswith(".pfx") or f.lower().endswith(".p12")):
                                    caminho_pfx_encontrado = os.path.join(root, f)
                                    pfx_filename = f
                                    found = True
                                    break
                            if found:
                                break
                        if caminho_pfx_encontrado:
                            break
                            
                if caminho_pfx_encontrado:
                    condominios_com_pfx += 1
                    senha_encontrada = senha_padrao
                    match_senha = re.search(r'senha\s*([a-zA-Z0-9@#$_-]+)', pfx_filename, re.IGNORECASE)
                    if match_senha:
                        senha_encontrada = match_senha.group(1)
                        status_str = f"Encontrado em '{os.path.relpath(caminho_pfx_encontrado, caminho_ativo)}' (Senha: {senha_encontrada})"
                    else:
                        status_str = f"Encontrado em '{os.path.relpath(caminho_pfx_encontrado, caminho_ativo)}' (Senha Padrão: {senha_padrao})"
                else:
                    status_str = "NÃO ENCONTRADO NO DISCO (PFX FALTANDO)"
                    
                nome_formatado = nome_val[:42] + "..." if len(nome_val) > 42 else nome_val
                print(f"{nome_formatado:<45} | {cnpj_limpo:<15} | {status_str}")
                
        print("="*80)
        log(f"Resumo do Diagnóstico: Encontrados {condominios_encontrados} condomínio(s) na planilha. {condominios_com_pfx} possui(em) arquivo PFX localizado.", "SUCCESS")
        
    except Exception as e:
        log(f"Erro ao executar diagnóstico da planilha: {e}", "ERROR")

if __name__ == "__main__":
    test_diagnostico()
