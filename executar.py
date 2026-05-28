import os
import sys
import csv
import time
import json
import ctypes
import datetime
import requests
from playwright.sync_api import sync_playwright, TimeoutError
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

class JapeException(Exception):
    pass

# Configuração de Logs
def log(msg, level="INFO"):
    timestamp = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    msg_str = str(msg)
    
    # Tenta obter a codificação do console, se não tiver usa utf-8
    encoding = sys.stdout.encoding or 'utf-8'
    try:
        msg_encoded = msg_str.encode(encoding, errors='replace').decode(encoding)
        print(f"[{timestamp}] [{level}] {msg_encoded}")
    except Exception:
        try:
            print(f"[{timestamp}] [{level}] {msg_str.encode('ascii', errors='replace').decode('ascii')}")
        except Exception:
            pass
            
    # Salvar em arquivo de log
    log_dir = "logs"
    try:
        os.makedirs(log_dir, exist_ok=True)
        today = datetime.date.today().strftime("%Y-%m-%d")
        with open(os.path.join(log_dir, f"execucao_{today}.log"), "a", encoding="utf-8") as f:
            f.write(f"[{timestamp}] [{level}] {msg_str}\n")
    except Exception:
        pass

# Função preventiva para encerrar processos órfãos da automação e liberar arquivos de bloqueio
def limpar_processos_automatizados_antigos():
    log("Iniciando limpeza preventiva de processos órfãos da automação...", "SYSTEM")
    import subprocess
    # Encerra processos chrome e edge que tenham a nossa pasta de perfil temporária nos argumentos de linha de comando
    cmd = (
        'powershell -NoProfile -Command "'
        'Get-CimInstance Win32_Process -Filter \\"Name = \'chrome.exe\' or Name = \'msedge.exe\' or Name = \'chromedriver.exe\' or Name = \'msedgedriver.exe\'\\" '
        '| Where-Object { $_.CommandLine -like \'*chrome_profile_chrome*\' or $_.CommandLine -like \'*chrome_profile_msedge*\' or $_.CommandLine -like \'*chrome_profile_chromium*\' } '
        '| ForEach-Object { Stop-Process -Id $_.ProcessId -Force }"'
    )
    try:
        res = subprocess.run(cmd, shell=True, capture_output=True, text=True)
        log("Processos órfãos anteriores analisados e finalizados.", "SUCCESS")
    except Exception as e:
        log(f"Aviso ao limpar processos antigos via PowerShell: {e}", "WARNING")

    # Deletar arquivos SingletonLock para destravar o Playwright
    for folder in ["chrome_profile_chrome", "chrome_profile_msedge", "chrome_profile"]:
        lock_path = os.path.join("temp", folder, "SingletonLock")
        if os.path.exists(lock_path):
            try:
                os.remove(lock_path)
                log(f"Arquivo de lock removido com sucesso: {lock_path}", "SUCCESS")
            except Exception as e:
                log(f"Aviso ao remover {lock_path} (pode já ter sido destravado): {e}", "WARNING")

# Enviar o caractere inicial (se fornecido) e depois a tecla ENTER para o Windows
def dialogo_certificado_aberto():
    EnumWindows = ctypes.windll.user32.EnumWindows
    EnumWindowsProc = ctypes.WINFUNCTYPE(ctypes.c_bool, ctypes.c_int, ctypes.c_int)
    GetWindowText = ctypes.windll.user32.GetWindowTextW
    GetWindowTextLength = ctypes.windll.user32.GetWindowTextLengthW
    IsWindowVisible = ctypes.windll.user32.IsWindowVisible
    
    aberto = [False]
    
    def foreach_window(hwnd, lParam):
        if IsWindowVisible(hwnd):
            length = GetWindowTextLength(hwnd)
            buff = ctypes.create_unicode_buffer(length + 1)
            GetWindowText(hwnd, buff, length + 1)
            title = buff.value
            title_lower = title.lower()
            
            # Blacklist para evitar planilhas, códigos e editores abertos do usuário
            blacklist = ["- excel", "- word", "- notepad", "visual studio", "code", ".py", ".xlsx", ".xls", "cmd.exe", "powershell", "escavador"]
            if any(b in title_lower for b in blacklist):
                return True
                
            termos_cert = ["selecione um certificado", "selecione o certificado", "confirmar certificado", 
                           "select a certificate", "confirm certificate", "segurança do windows", 
                           "windows security", "credenciais de segurança", "security credentials", 
                           "pin do certificado", "insira o pin", "controle de acesso"]
            if any(t in title_lower for t in termos_cert) or (any(x in title_lower for x in ["certificado", "segurança"]) and not any(b in title_lower for b in blacklist)):
                aberto[0] = True
                return False
        return True

    EnumWindows(EnumWindowsProc(foreach_window), 0)
    return aberto[0]

def focar_janela_certificado(keywords_browser):
    EnumWindows = ctypes.windll.user32.EnumWindows
    EnumWindowsProc = ctypes.WINFUNCTYPE(ctypes.c_bool, ctypes.c_int, ctypes.c_int)
    GetWindowText = ctypes.windll.user32.GetWindowTextW
    GetWindowTextLength = ctypes.windll.user32.GetWindowTextLengthW
    IsWindowVisible = ctypes.windll.user32.IsWindowVisible
    SetForegroundWindow = ctypes.windll.user32.SetForegroundWindow
    ShowWindow = ctypes.windll.user32.ShowWindow
    
    found_hwnds = []
    
    def foreach_window(hwnd, lParam):
        if IsWindowVisible(hwnd):
            length = GetWindowTextLength(hwnd)
            buff = ctypes.create_unicode_buffer(length + 1)
            GetWindowText(hwnd, buff, length + 1)
            title = buff.value
            title_lower = title.lower()
            
            # Blacklist para evitar planilhas, códigos e editores abertos do usuário
            blacklist = ["- excel", "- word", "- notepad", "visual studio", "code", ".py", ".xlsx", ".xls", "cmd.exe", "powershell", "escavador"]
            if any(b in title_lower for b in blacklist):
                return True
                
            termos_cert = ["selecione um certificado", "selecione o certificado", "confirmar certificado", 
                           "select a certificate", "confirm certificate", "segurança do windows", 
                           "windows security", "credenciais de segurança", "security credentials", 
                           "pin do certificado", "insira o pin", "controle de acesso"]
            
            # 1. Diálogos de certificado ou segurança
            if any(t in title_lower for t in termos_cert) or (any(x in title_lower for x in ["certificado", "segurança"]) and not any(b in title_lower for b in blacklist)):
                found_hwnds.append((hwnd, title, 2))
            # 2. Janela do navegador da nossa automação
            elif any(x in title_lower for x in keywords_browser):
                found_hwnds.append((hwnd, title, 1))
        return True

    EnumWindows(EnumWindowsProc(foreach_window), 0)
    
    if not found_hwnds:
        log("[AUTO-LOGIN] Nenhuma janela de certificado ou navegador encontrada para focar.", "WARNING")
        return False
        
    found_hwnds.sort(key=lambda x: x[2], reverse=True)
    target_hwnd, title, priority = found_hwnds[0]
    
    log(f"[AUTO-LOGIN] Janela alvo encontrada (Prioridade {priority}): '{title}'", "SYSTEM")
    try:
        current_active = ctypes.windll.user32.GetForegroundWindow()
        if current_active != target_hwnd:
            # Simula toque no ALT para destravar SetForegroundWindow
            ctypes.windll.user32.keybd_event(0x12, 0, 0, 0)
            ctypes.windll.user32.keybd_event(0x12, 0, 2, 0)
            ShowWindow(target_hwnd, 9) # SW_RESTORE
            SetForegroundWindow(target_hwnd)
            time.sleep(0.5)
        else:
            log("[AUTO-LOGIN] Janela alvo já está ativa. Ignorando refocus.", "SYSTEM")
        return True
    except Exception as e:
        log(f"[AUTO-LOGIN] Erro ao focar janela: {e}", "WARNING")
        return False

def press_enter(first_char=None):
    focar_janela_certificado(["e-cac", "cav.receita", "receita federal", "chrome", "edge"])
    if first_char:
        char_upper = first_char.upper()
        if len(char_upper) == 1 and (char_upper.isalnum() or char_upper == " "):
            vk_code = ord(char_upper)
            log(f"Enviando tecla '{char_upper}' (VK: {hex(vk_code)}) para focar no certificado...", "SYSTEM")
            ctypes.windll.user32.keybd_event(vk_code, 0, 0, 0) # Key Down
            time.sleep(0.05)
            ctypes.windll.user32.keybd_event(vk_code, 0, 2, 0) # Key Up
            time.sleep(0.5)
            
    log("Enviando comando ENTER para o Windows...", "SYSTEM")
    VK_RETURN = 0x0D
    ctypes.windll.user32.keybd_event(VK_RETURN, 0, 0, 0) # Key Down
    ctypes.windll.user32.keybd_event(VK_RETURN, 0, 2, 0) # Key Up
    log("ENTER enviado.", "SYSTEM")

def executar_confirmacao_certificado_em_loop(config, log_prefix):
    import time
    first_char = config.get("cert_first_char", "J")
    log(f"{log_prefix} Iniciando rotina de confirmacao de certificado SSL (Letra: '{first_char}')...", "SYSTEM")
    
    # Aguarda 3.5 segundos para a janela do Chrome iniciar o redirecionamento e abrir o diálogo
    time.sleep(3.5)
    
    # Realiza 3 tentativas espaçadas de focar e dar ENTER
    for attempt in range(1, 4):
        log(f"{log_prefix} Tentativa {attempt}/3 de confirmacao...", "SYSTEM")
        press_enter(first_char if attempt == 1 else None)
        time.sleep(1.5)
        
    log(f"{log_prefix} Rotina de confirmacao concluida.", "SUCCESS")

# Função auxiliar para fechar modais jQuery UI e Caixa Postal que cobrem a tela
def fechar_modais_e_overlays(page):
    try:
        # 1. Tentar fechar pelo botão (X) do dialog jQuery UI
        btn_close = page.locator(".ui-dialog-titlebar-close, button:has-text('Close'), button[title='Close']").first
        if btn_close.is_visible():
            log("[MODAL-CLEANUP] Botão de fechar (X) detectado. Clicando...", "INFO")
            btn_close.click(timeout=1500)
            time.sleep(0.5)
    except Exception:
        pass

    try:
        # 2. Enviar a tecla ESCAPE para fechar diálogos ativos
        page.keyboard.press("Escape")
    except Exception:
        pass

    try:
        # 3. Remover fisicamente qualquer overlay/dialog do DOM para garantir 100% de passagem e evitar cliques interceptados
        overlays = page.locator(".ui-widget-overlay, .ui-dialog, .ui-widget-shadow")
        if overlays.first.is_visible():
            log("[MODAL-CLEANUP] Removendo modais/overlays remanescentes do DOM...", "WARNING")
            page.evaluate("""
                document.querySelectorAll('.ui-widget-overlay, .ui-dialog, .ui-widget-shadow, .ui-dialog-buttonpane').forEach(el => {
                    el.style.display = 'none';
                    el.remove();
                });
            """)
            time.sleep(0.3)
    except Exception as e:
        log(f"[MODAL-CLEANUP] Falha ao limpar DOM via JS: {e}", "WARNING")

# Função auxiliar para checar e tratar a Caixa Postal bloqueante (lendo mensagens importantes para unblock)
def checar_e_tratar_caixa_postal(page, client_dir, config):
    try:
        # Detectar se o modal de aviso bloqueante está presente
        btn_caixa = None
        for selector in [
            "button:has-text('Ir para a Caixa Postal')",
            "input[value='Ir para a Caixa Postal']",
            "text=Ir para a Caixa Postal"
        ]:
            try:
                loc = page.locator(selector).first
                if loc.is_visible(timeout=1000):
                    btn_caixa = loc
                    break
            except Exception:
                continue
        
        if btn_caixa:
            log("[ALERTA CRÍTICO] Caixa Postal bloqueante detectada! O e-CAC exige a leitura de mensagens importantes para unblock.", "WARNING")
            
            # Clicar em "Ir para a Caixa Postal"
            btn_caixa.click()
            page.wait_for_load_state("load")
            page.wait_for_timeout(2000)
            
            # Buscar e abrir a mensagem com indicativo de alerta (!) ou não lida
            mensagem_aberta = False
            assunto_tabela = ""
            for sel in [
                "tr:has-text('!') a",
                "tr.nao-lida a",
                "xpath=//table//tr/td[4]//a",
                "xpath=//table//tr/td[4]",
                "xpath=//tr/td[4]//a",
                "xpath=//tr/td[4]",
                "xpath=//td[contains(@class, 'assunto')]//a",
                "xpath=//td[contains(@class, 'assunto')]",
                "xpath=//a[contains(@href, 'mensagem') or contains(@href, 'Mensagem') or contains(@href, 'Visualizar')]",
                "xpath=//tr[contains(., 'RECEITA FEDERAL')]//td[4]//a",
                "css=table tbody tr td a",
                "css=table tbody tr td",
                "td.assunto a",
                "a:has-text('ALERTA')",
                "a:has-text('risco')",
                "a[href*='Mensagem']",
                "//tr[contains(., '!')]//a"
            ]:
                try:
                    loc = page.locator(sel).first
                    if loc.is_visible(timeout=1500):
                        texto_msg = loc.inner_text().strip()
                        assunto_tabela = texto_msg
                        log(f"[CAIXA-POSTAL] Abrindo mensagem importante usando seletor '{sel}': '{texto_msg}'", "ACTION")
                        loc.click()
                        mensagem_aberta = True
                        break
                except Exception:
                    continue
            
            if mensagem_aberta:
                page.wait_for_timeout(2000)
                
                # Extrair assunto e conteúdo da mensagem
                assunto = ""
                conteudo = ""
                try:
                    assunto = page.locator("h1, h2, .titulo-mensagem, td.assunto").first.inner_text().strip()
                except Exception:
                    pass
                    
                if not assunto or len(assunto) < 3:
                    assunto = assunto_tabela.replace("!", "").strip() if assunto_tabela else "Alerta Importante e-CAC"
                    
                try:
                    conteudo = page.locator("body").inner_text().strip()
                except Exception:
                    conteudo = "Não foi possível extrair o corpo completo da mensagem."
                
                log(f"[CAIXA-POSTAL] Alerta de Caixa Postal extraído: '{assunto[:60]}...'", "SUCCESS")
                
                # Gravar arquivo de alerta TXT e JSON para estruturação fácil no Excel
                alerta_path = os.path.join(client_dir, "ALERTA_CAIXA_POSTAL.txt")
                with open(alerta_path, "w", encoding="utf-8") as f:
                    f.write("============================================================\n")
                    f.write(" MENSAGEM IMPORTANTE / ALERTA DE EXCLUSÃO CAPTURADO NO e-CAC\n")
                    f.write("============================================================\n")
                    f.write(f"Data Captura : {datetime.datetime.now().strftime('%d/%m/%Y %H:%M:%S')}\n")
                    f.write(f"Assunto      : {assunto}\n")
                    f.write(f"Conteúdo     :\n\n{conteudo}\n")
                
                alerta_json_path = os.path.join(client_dir, "ALERTA_CAIXA_POSTAL.json")
                with open(alerta_json_path, "w", encoding="utf-8") as f:
                    json.dump({
                        "assunto": assunto,
                        "conteudo": conteudo,
                        "data_captura": datetime.datetime.now().strftime('%d/%m/%Y %H:%M:%S')
                    }, f, indent=4, ensure_ascii=False)
                
                log(f"[CAIXA-POSTAL] Alerta salvo com sucesso em TXT e JSON: {client_dir}", "SUCCESS")
                
                # Retornar para a Home do e-CAC para resetar e continuar o fluxo sem o modal bloqueante!
                log("[CAIXA-POSTAL] Retornando para a página inicial do e-CAC para prosseguir...", "ACTION")
                page.goto(config["portal_url"])
                page.wait_for_selector("text=Alterar perfil de acesso", timeout=10000)
                
            else:
                log("[CAIXA-POSTAL] Caixa Postal aberta, mas não foi possível localizar a mensagem unread correspondente.", "WARNING")
                
    except Exception as e:
        log(f"Aviso ao tentar tratar Caixa Postal bloqueante: {e}. Prosseguindo com o fluxo...", "WARNING")

# Função auxiliar para navegar para o e-CAC de forma orgânica via busca do Google
def navegar_via_google_para_ecac(page):
    log("Iniciando acesso organico via busca do Google para evitar falhas de rede e bloqueios...", "INFO")
    import random
    
    try:
        # 1. Abre o Google
        log("Abrindo google.com...", "ACTION")
        page.goto("https://www.google.com", timeout=20000)
        
        # Pausa humana aleatoria (1.5s a 3.0s)
        tempo_reacao = random.uniform(1.5, 3.0)
        page.wait_for_timeout(tempo_reacao * 1000)
        
        # 2. Localizar o input de busca do Google
        selector_busca = 'textarea[name="q"], input[name="q"]'
        page.wait_for_selector(selector_busca, timeout=8000)
        
        # Simular clique humano para focar no campo
        log("Focando no campo de busca do Google...", "ACTION")
        page.click(selector_busca)
        page.wait_for_timeout(random.uniform(0.5, 1.0) * 1000)
        
        # Digitar caractere por caractere com atraso variavel
        log("Digitando 'ecac' caractere por caractere...", "ACTION")
        for char in "ecac":
            page.keyboard.type(char)
            # Atraso aleatorio por tecla (150ms a 300ms)
            page.wait_for_timeout(random.uniform(0.15, 0.3) * 1000)
            
        # Pausa antes de enviar (1.0s a 2.0s)
        page.wait_for_timeout(random.uniform(1.0, 2.0) * 1000)
        
        # Pressionar Enter
        log("Enviando busca do Google...", "ACTION")
        page.keyboard.press("Enter")
        
        # 3. Aguardar os resultados e inserir pausa de leitura
        log("Aguardando exibicao dos resultados...", "ACTION")
        selector_link = 'a[href*="cav.receita.fazenda.gov.br"]'
        page.wait_for_selector(selector_link, timeout=12000)
        
        # Simular leitura humana dos resultados (2s a 4s)
        tempo_leitura = random.uniform(2.0, 4.0)
        log(f"Simulando leitura humana dos resultados por {tempo_leitura:.1f} segundos...", "INFO")
        page.wait_for_timeout(tempo_leitura * 1000)
        
        # Focar no link e clicar de forma suave
        log("Clicando no link do e-CAC nos resultados da busca...", "ACTION")
        link_loc = page.locator(selector_link).first
        link_loc.focus()
        page.wait_for_timeout(random.uniform(0.5, 1.2) * 1000)
        link_loc.click()
        
        # 4. Aguardar o carregamento final
        page.wait_for_load_state("load")
        log("Portal e-CAC carregado com sucesso via Google!", "SUCCESS")
        page.wait_for_timeout(1000) # Pequena pausa de assentamento
        
    except Exception as e:
        log(f"Falha na navegacao via busca do Google: {e}. Fazendo fallback para acesso direto...", "WARNING")
        # Fallback silencioso direto para a URL do e-CAC
        page.goto("https://cav.receita.fazenda.gov.br/ecac/", timeout=25000)

# Função auxiliar para garantir que a sessão com o e-CAC está ativa e realizar login se necessário
def verificar_e_reestabelecer_sessao(page, config):
    log("Verificando se a sessão com o e-CAC está ativa...", "INFO")
    try:
        # Acessa a URL principal
        page.goto(config["portal_url"])
        fechar_modais_e_overlays(page)
        
        # Tenta esperar pelo elemento do painel autenticado
        page.wait_for_selector("text=Alterar perfil de acesso", timeout=8000)
        log("Sessão e-CAC validada e ativa.", "SUCCESS")
        return True
    except Exception:
        log("Sessão expirada ou deslogada. Tentando reestabelecer login automaticamente...", "WARNING")
        
        try:
            # Se não estiver no e-CAC, vai para a página inicial via Google
            if "cav.receita.fazenda.gov.br" not in page.url:
                navegar_via_google_para_ecac(page)
            
            # Se o botão de login do Gov.br estiver visível, clica nele
            btn_gov = page.locator('input[alt="Acesso Gov BR"]').first
            if btn_gov.is_visible(timeout=5000):
                log("Botão 'Acesso Gov BR' detectado. Clicando...", "ACTION")
                btn_gov.click()
                page.wait_for_timeout(2000)
                
            # Aguarda o botão do certificado
            page.wait_for_selector('button#login-certificate, #login-certificate', timeout=15000)
            
            # Dispara a thread para dar ENTER com monitoramento em loop
            import threading
            threading.Thread(
                target=executar_confirmacao_certificado_em_loop,
                args=(config, "[AUTO-LOGIN-RETRY]"),
                daemon=True
            ).start()
            
            log("Clicando no botão 'Seu certificado digital'...", "ACTION")
            try:
                page.click('button#login-certificate', timeout=5000)
            except Exception as e_click:
                log(f"Aviso no clique do certificado (TLS travou a thread temporariamente, esperado): {e_click}", "INFO")
            
            # Aguarda o painel
            page.wait_for_selector("text=Alterar perfil de acesso", timeout=45000)
            log("Login restabelecido com sucesso na sessão ativa!", "SUCCESS")
            
            # Salvar o estado da sessão atualizado em state.json
            try:
                page.context.storage_state(path="state.json")
                log("Sessão salva com sucesso em 'state.json'.", "SUCCESS")
            except Exception as e:
                log(f"Erro ao salvar estado da sessão: {e}", "WARNING")
                
            return True
        except Exception as e:
            log(f"Falha ao tentar reestabelecer login automaticamente: {e}. Solicitando intervenção manual se necessário...", "ERROR")
            if not config["headless"]:
                log("Aguardando login manual na tela (limite de 120 segundos)...", "IMPORTANT")
                
                url_atual = "N/A"
                if page:
                    try:
                        url_atual = page.url
                    except Exception:
                        pass
                
                mensagem_alerta = (
                    "⚠️ *ALERTA DO ROBÔ e-CAC (SESSÃO CAIU)*\n\n"
                    "A sessão do e-CAC caiu durante a varredura e o login automático falhou.\n"
                    f"• *Link da Página Atual*: {url_atual}\n"
                    "• *O que fazer*: Acesse o computador e conclua o login manualmente na tela do navegador (resolva o CAPTCHA se houver).\n\n"
                    "O robô aguardará por até 2 minutos."
                )
                enviar_whatsapp(mensagem_alerta, config)
                
                try:
                    page.wait_for_selector("text=Alterar perfil de acesso", timeout=120000)
                    log("Login manual detectado com sucesso!", "SUCCESS")
                    try:
                        page.context.storage_state(path="state.json")
                    except Exception:
                        pass
                    return True
                except Exception:
                    pass
            return False

# Carregar arquivo de configuração
def load_config():
    config_path = "config.json"
    private_config_path = "config_private.json"
    config = {
        "headless": False,
        "timeout_ms": 30000,
        "relatorios_dir": "relatorios",
        "clientes_file": "clientes.csv",
        "portal_url": "https://cav.receita.fazenda.gov.br/eCAC/Default.aspx#",
        "download_timeout_ms": 60000,
        "whatsapp_enabled": True,
        "whatsapp_number": "",
        "whatsapp_zapi_instance": "",
        "whatsapp_zapi_token": "",
        "whatsapp_zapi_client_token": "",
        "openai_api_key": ""
    }
    if os.path.exists(config_path):
        try:
            with open(config_path, "r", encoding="utf-8") as f:
                dados = json.load(f)
                config.update(dados)
        except Exception:
            pass
    if os.path.exists(private_config_path):
        try:
            with open(private_config_path, "r", encoding="utf-8") as f:
                dados_privados = json.load(f)
                config.update(dados_privados)
        except Exception:
            pass
    return config

# Carregar lista de clientes
def load_clients(filepath):
    clients = []
    if not os.path.exists(filepath):
        log(f"Arquivo de clientes não encontrado: {filepath}", "ERROR")
        return clients
        
    try:
        # Detectar e abrir com suporte a UTF-8-BOM (comum em exports do Excel no Windows)
        with open(filepath, "r", encoding="utf-8-sig") as f:
            reader = csv.DictReader(f)
            
            # Normalizar os nomes das colunas (remover espaços e colocar em minúsculas)
            raw_headers = reader.fieldnames if reader.fieldnames else []
            headers = [h.strip().lower() if h else "" for h in raw_headers]
            
            # Mapeamento inteligente e flexível de colunas
            col_cnpj = next((h for h in headers if h in ["cnpj", "c.n.p.j.", "cnpj_cliente", "documento"]), "cnpj")
            col_nome = next((h for h in headers if h in ["nome", "nome_cliente", "razao_social", "razao social", "cliente", "nome cliente"]), "nome")
            col_ativo = next((h for h in headers if h in ["ativo", "active", "status", "habilitado"]), "ativo")
            
            log(f"Colunas mapeadas do CSV: CNPJ -> '{col_cnpj}', Nome -> '{col_nome}', Ativo -> '{col_ativo}'", "INFO")
            
            for row in reader:
                raw_cnpj = ""
                raw_nome = ""
                raw_ativo = "True"
                
                # Buscar valores contornando possíveis diferenças de caixa e espaços no CSV
                for k, v in row.items():
                    if not k:
                        continue
                    k_lower = k.strip().lower()
                    if k_lower == col_cnpj:
                        raw_cnpj = v
                    elif k_lower == col_nome:
                        raw_nome = v
                    elif k_lower == col_ativo:
                        raw_ativo = v
                
                # Higienização dos dados extraídos
                cnpj = "".join(filter(str.isdigit, raw_cnpj or ""))
                nome = (raw_nome or "").strip()
                ativo = (raw_ativo or "True").strip().lower() in ["true", "1", "yes", "ativo", "sim", "s"]
                
                # Fallback de segurança para evitar colapsar caminhos se o nome estiver em branco
                if cnpj and not nome:
                    nome = f"Cliente_{cnpj}"
                    
                if cnpj:
                    clients.append({"cnpj": cnpj, "nome": nome, "ativo": ativo})
                    
        log(f"Carregados com sucesso {len(clients)} clientes a partir de '{filepath}'.", "SUCCESS")
    except Exception as e:
        log(f"Erro crítico ao ler o arquivo de clientes '{filepath}': {e}", "ERROR")
        
    return clients


# Salvar status do processamento do cliente
def save_client_status(relatorios_dir, cnpj, nome, status, details=""):
    today = datetime.date.today().strftime("%Y-%m-%d")
    # Limpar nome para criar pasta segura, mantendo letras acentuadas, cedilhas e espaços
    nome_limpo = "".join(c if c.isalnum() or c in " _-ÇçÁáÉéÍíÓóÚúÃãÕõÂâÊêÔôÀàÜü" else "_" for c in nome).strip()
    cnpj_limpo = "".join(filter(str.isdigit, cnpj))
    client_dir = os.path.join(relatorios_dir, f"{cnpj_limpo}_{nome_limpo}")
    os.makedirs(client_dir, exist_ok=True)
    
    status_path = os.path.join(client_dir, "status.json")
    contador_falhas = 0
    
    # Se já existir o status.json e for do dia de hoje, resgata o contador de falhas
    if os.path.exists(status_path):
        try:
            with open(status_path, "r", encoding="utf-8") as f:
                data = json.load(f)
                if data.get("data_consulta") == today:
                    contador_falhas = data.get("contador_falhas_hoje", 0)
        except Exception:
            pass
            
    # Ajustar contador conforme o novo status
    if status == "Erro":
        contador_falhas += 1
    elif status == "Sucesso":
        contador_falhas = 0
        
    status_data = {
        "cnpj": cnpj,
        "nome": nome,
        "data_consulta": today,
        "hora_consulta": datetime.datetime.now().strftime("%H:%M:%S"),
        "status": status,
        "detalhes": details,
        "contador_falhas_hoje": contador_falhas
    }
    
    with open(status_path, "w", encoding="utf-8") as f:
        json.dump(status_data, f, indent=4, ensure_ascii=False)
        
    return client_dir

def format_cnpj(cnpj):
    if len(cnpj) == 14:
        return f"{cnpj[:2]}.{cnpj[2:5]}.{cnpj[5:8]}/{cnpj[8:12]}-{cnpj[12:]}"
    elif len(cnpj) == 11:
        return f"{cnpj[:3]}.{cnpj[3:6]}.{cnpj[6:9]}-{cnpj[9:]}"
    return cnpj

def clean_filename(name):
    return "".join(c if c.isalnum() or c in " _-ÇçÁáÉéÍíÓóÚúÃãÕõÂâÊêÔôÀàÜü" else "_" for c in name).strip()

def remover_arquivos_fiscais_obsoletos(client_dir, arquivo_mantido_path=None):
    """
    Remove todos os arquivos e subpastas na pasta do cliente (client_dir),
    exceto o status.json e o arquivo_mantido_path (se fornecido).
    """
    try:
        import shutil
        if not os.path.exists(client_dir):
            return
        for item in os.listdir(client_dir):
            item_path = os.path.join(client_dir, item)
            if item.lower() == "status.json":
                continue
            if arquivo_mantido_path and os.path.normpath(item_path) == os.path.normpath(arquivo_mantido_path):
                continue
            try:
                if os.path.isdir(item_path):
                    shutil.rmtree(item_path)
                    log(f"Removida subpasta antiga: {item_path}", "INFO")
                else:
                    os.remove(item_path)
                    log(f"Removido arquivo obsoleto: {item_path}", "INFO")
            except Exception as e:
                log(f"Não foi possível remover item obsoleto {item_path}: {e}", "WARNING")
    except Exception as e:
        log(f"Erro ao remover arquivos obsoletos: {e}", "WARNING")

def gerar_consolidado_excel(clientes, relatorios_dir, output_path):
    log(f"Iniciando a geração do Painel Excel Consolidado: {output_path}", "INFO")
    try:
        import openpyxl
        from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
        
        today_str = datetime.date.today().strftime("%Y-%m-%d")
        month_str = datetime.date.today().strftime("%Y-%m")
        rows_data = []
        
        for c in clientes:
            cnpj_raw = c["cnpj"]
            nome_raw = c["nome"]
            ativo = c["ativo"]
            
            if not ativo:
                rows_data.append({
                    "cnpj": format_cnpj(cnpj_raw),
                    "nome": nome_raw,
                    "status": "Inativo",
                    "detalhes": "Ignorado",
                    "data_hora": "-",
                    "observacao": "Cliente inativo no cadastro"
                })
                continue
                
            nome_limpo = clean_filename(nome_raw)
            cnpj_limpo = "".join(filter(str.isdigit, cnpj_raw))
            status_path = os.path.join(relatorios_dir, f"{cnpj_limpo}_{nome_limpo}", "status.json")
            
            status = "Pendente"
            detalhes = "Não Consultado"
            data_hora = "-"
            observacao = "Aguardando processamento"
            
            if os.path.exists(status_path):
                try:
                    with open(status_path, "r", encoding="utf-8") as f:
                        data = json.load(f)
                        status = data.get("status", "Pendente")
                        detalhes = data.get("detalhes", "Não Consultado")
                        data_hora = f"{data.get('data_consulta', today_str)} {data.get('hora_consulta', '')}".strip()
                        if status == "Erro":
                            detalhes_lower = detalhes.lower()
                            if "procuração" in detalhes_lower or "procuracao" in detalhes_lower or "cancelada" in detalhes_lower or "inexistente" in detalhes_lower or "expirada" in detalhes_lower or "não cadastrada" in detalhes_lower or "revogada" in detalhes_lower:
                                status = "Erro (Procuração)"
                                msg_limpa = detalhes
                                if "Erro do e-CAC:" in msg_limpa:
                                    msg_limpa = msg_limpa.split("Erro do e-CAC:", 1)[1].strip()
                                detalhes = "Procuração Inválida / Inexistente"
                                observacao = msg_limpa
                            else:
                                observacao = detalhes
                                detalhes = "Falha no Processamento"
                        elif status == "Sucesso":
                            if "Certidão" in detalhes:
                                observacao = "CND emitida com sucesso (Sem Pendências)"
                            elif "Sem Pendências" in detalhes:
                                observacao = "Regular na Receita Federal (Informativo Gravado)"
                            elif "Relatório" in detalhes:
                                observacao = "Pendências tributárias ativas (Relatório PDF baixado)"
                            else:
                                observacao = "Concluído com sucesso"
                except Exception as e:
                    observacao = f"Erro ao ler status: {e}"
                    
            # Verifica se há alertas de Caixa Postal capturados para este cliente
            alerta_file = os.path.join(relatorios_dir, f"{cnpj_limpo}_{nome_limpo}", "ALERTA_CAIXA_POSTAL.txt")
            alerta_json = os.path.join(relatorios_dir, f"{cnpj_limpo}_{nome_limpo}", "ALERTA_CAIXA_POSTAL.json")
            alerta_texto = "-"
            
            if os.path.exists(alerta_json):
                try:
                    with open(alerta_json, "r", encoding="utf-8") as f_json:
                        alerta_data = json.load(f_json)
                        # Salva apenas o assunto no Excel consolidado para manter a legibilidade,
                        # mantendo o conteúdo da mensagem completo nos arquivos locais do cliente (.json/.txt)
                        alerta_texto = alerta_data.get("assunto", "Alerta Caixa Postal").strip()
                except Exception:
                    alerta_texto = "Erro ao carregar assunto da mensagem."
            elif os.path.exists(alerta_file):
                alerta_texto = "Alerta Caixa Postal (detalhes no arquivo .txt)"
                
            if os.path.exists(alerta_file):
                observacao = f"[ATENÇÃO] Alerta de Caixa Postal capturado! | {observacao}"
                
            rows_data.append({
                "cnpj": format_cnpj(cnpj_raw),
                "nome": nome_raw,
                "status": status,
                "detalhes": detalhes,
                "data_hora": data_hora,
                "alerta_caixa_postal": alerta_texto,
                "observacao": observacao
            })
            
        # Check if the Excel file exists on the Desktop and reuse it
        if os.path.exists(output_path):
            try:
                log(f"Reutilizando arquivo Excel existente na Area de Trabalho: {output_path}", "INFO")
                wb = openpyxl.load_workbook(output_path)
                # Se a planilha "Painel de Controle e-CAC" ja existir, removemos ela para criá-la do zero
                # garantindo que nao fiquem dados obsoletos do passado, mas preservando outras planilhas.
                if "Painel de Controle e-CAC" in wb.sheetnames:
                    idx = wb.sheetnames.index("Painel de Controle e-CAC")
                    wb.remove(wb["Painel de Controle e-CAC"])
                    ws = wb.create_sheet("Painel de Controle e-CAC", idx)
                else:
                    ws = wb.create_sheet("Painel de Controle e-CAC")
            except Exception as load_err:
                log(f"Aviso ao carregar Excel existente ({load_err}). Criando novo do zero.", "WARNING")
                wb = openpyxl.Workbook()
                ws = wb.active
                ws.title = "Painel de Controle e-CAC"
        else:
            log(f"Arquivo Excel nao encontrado. Criando novo consolidado: {output_path}", "INFO")
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Painel de Controle e-CAC"
            
        ws.views.sheetView[0].showGridLines = True
        
        # Design system styles (Navy & Premium accents)
        navy_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
        white_font_16 = Font(name="Calibri", size=16, bold=True, color="FFFFFF")
        gray_font_11 = Font(name="Calibri", size=11, bold=True, color="595959")
        gray_font_10 = Font(name="Calibri", size=10, italic=True, color="595959")
        
        header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
        header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
        
        white_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
        zebra_fill = PatternFill(start_color="FAFAFA", end_color="FAFAFA", fill_type="solid")
        
        green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        green_font = Font(name="Calibri", size=10, bold=True, color="006100")
        
        red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        red_font = Font(name="Calibri", size=10, bold=True, color="9C0006")
        
        yellow_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
        yellow_font = Font(name="Calibri", size=10, bold=True, color="9C6500")
        
        gray_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
        gray_font = Font(name="Calibri", size=10, color="595959")
        
        orange_fill = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")
        orange_font = Font(name="Calibri", size=10, bold=True, color="C65911")
        
        thin_border_side = Side(border_style="thin", color="D3D3D3")
        data_border = Border(left=thin_border_side, right=thin_border_side, top=thin_border_side, bottom=thin_border_side)
        
        # Helper function to apply styles to merged cell ranges
        def style_range(ws, cell_range, font=None, fill=None, alignment=None, border=None):
            for row in ws[cell_range]:
                for cell in row:
                    if font is not None: cell.font = font
                    if fill is not None: cell.fill = fill
                    if alignment is not None: cell.alignment = alignment
                    if border is not None: cell.border = border
 
        # Merge columns for titles and style them perfectly without cutouts (7 columns now!)
        ws.merge_cells("A1:G1")
        ws["A1"] = "J&J CONTABILIDADE — PAINEL DE CONTROLE DE PENDÊNCIAS FISCAIS (e-CAC)"
        style_range(ws, "A1:G1", font=white_font_16, fill=navy_fill, alignment=Alignment(horizontal="center", vertical="center"))
        ws.row_dimensions[1].height = 40
        
        ws.merge_cells("A2:G2")
        ws["A2"] = "Varredura de Débitos Federais e Situação Fiscal via Procuração Digital"
        style_range(ws, "A2:G2", font=gray_font_11, alignment=Alignment(horizontal="center", vertical="center"))
        ws.row_dimensions[2].height = 22
        
        ws.merge_cells("A3:G3")
        hoje_fmt = datetime.datetime.now().strftime("%d/%m/%Y às %H:%M")
        ws["A3"] = f"Relatório emitido em {hoje_fmt} | Escavador de Pendências Automatizado v1.0"
        style_range(ws, "A3:G3", font=gray_font_10, alignment=Alignment(horizontal="center", vertical="center"))
        ws.row_dimensions[3].height = 20
        
        ws.row_dimensions[4].height = 15
        
        headers = [
            "CNPJ",
            "Razão Social",
            "Status da Consulta",
            "Resultado / Documento",
            "Data/Hora Varredura",
            "Alerta Caixa Postal (e-CAC)",
            "Observações / Detalhes do Processamento"
        ]
        
        for col_idx, h in enumerate(headers, 1):
            cell = ws.cell(row=5, column=col_idx, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center" if col_idx not in [2, 6, 7] else "left", vertical="center", wrap_text=True)
            cell.border = data_border
            
        ws.row_dimensions[5].height = 28
        
        current_row = 6
        for idx, r in enumerate(rows_data):
            ws.cell(row=current_row, column=1, value=r["cnpj"]).alignment = Alignment(horizontal="center", vertical="center")
            ws.cell(row=current_row, column=2, value=r["nome"]).alignment = Alignment(horizontal="left", vertical="center")
            
            status_cell = ws.cell(row=current_row, column=3, value=r["status"])
            status_cell.alignment = Alignment(horizontal="center", vertical="center")
            
            detalhe_cell = ws.cell(row=current_row, column=4, value=r["detalhes"])
            detalhe_cell.alignment = Alignment(horizontal="center", vertical="center")
            
            ws.cell(row=current_row, column=5, value=r["data_hora"]).alignment = Alignment(horizontal="center", vertical="center")
            
            # Coluna 6: Texto completo do Alerta da Caixa Postal
            alerta_cell = ws.cell(row=current_row, column=6, value=r["alerta_caixa_postal"])
            alerta_cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
            
            # Coluna 7: Observações e Detalhes
            ws.cell(row=current_row, column=7, value=r["observacao"]).alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
            
            row_fill = zebra_fill if idx % 2 == 1 else white_fill
            
            for col_idx in range(1, 8):
                cell = ws.cell(row=current_row, column=col_idx)
                cell.border = data_border
                cell.fill = row_fill
                if col_idx not in [3, 4]:
                    cell.font = Font(name="Calibri", size=10)
                    
            st = r["status"]
            dt = r["detalhes"]
            
            if st == "Sucesso":
                if "Certidão" in dt or "Sem Pendências" in dt:
                    status_cell.fill = green_fill
                    status_cell.font = green_font
                    detalhe_cell.fill = green_fill
                    detalhe_cell.font = green_font
                elif "Relatório" in dt:
                    status_cell.fill = red_fill
                    status_cell.font = red_font
                    detalhe_cell.fill = red_fill
                    detalhe_cell.font = red_font
            elif st == "Erro (Procuração)":
                status_cell.fill = orange_fill
                status_cell.font = orange_font
                detalhe_cell.fill = orange_fill
                detalhe_cell.font = orange_font
            elif st == "Erro":
                status_cell.fill = yellow_fill
                status_cell.font = yellow_font
                detalhe_cell.fill = yellow_fill
                detalhe_cell.font = yellow_font
            else:
                status_cell.fill = gray_fill
                status_cell.font = gray_font
                detalhe_cell.fill = gray_fill
                detalhe_cell.font = gray_font
                
            ws.row_dimensions[current_row].height = 22
            current_row += 1
            
        ws.row_dimensions[current_row].height = 15
        current_row += 1
        
        totais_row = current_row
        ws.cell(row=totais_row, column=1, value="TOTAIS").alignment = Alignment(horizontal="right", vertical="center")
        ws.cell(row=totais_row, column=2, value=f"=COUNTA(B6:B{totais_row-2})").alignment = Alignment(horizontal="left", vertical="center")
        ws.cell(row=totais_row, column=3, value=f'=COUNTIF(C6:C{totais_row-2}, "Sucesso")').alignment = Alignment(horizontal="center", vertical="center")
        ws.cell(row=totais_row, column=4, value=f'=COUNTIF(D6:D{totais_row-2}, "*Certidão*") + COUNTIF(D6:D{totais_row-2}, "*Sem Pendências*")').alignment = Alignment(horizontal="center", vertical="center")
        ws.cell(row=totais_row, column=5, value=f'=COUNTIF(D6:D{totais_row-2}, "*Relatório*")').alignment = Alignment(horizontal="center", vertical="center")
        ws.cell(row=totais_row, column=6, value="").alignment = Alignment(horizontal="center", vertical="center")
        ws.cell(row=totais_row, column=7, value=f'=COUNTIF(C6:C{totais_row-2}, "Erro*")').alignment = Alignment(horizontal="left", vertical="center")
        
        # Professional borders for accounting totals (Double bottom line)
        double_bottom = Side(border_style="double", color="FFFFFF")
        thin_top = Side(border_style="thin", color="FFFFFF")
        thin_side = Side(border_style="thin", color="D3D3D3")
        totals_border = Border(left=thin_side, right=thin_side, top=thin_top, bottom=double_bottom)
        
        for col_idx in range(1, 8):
            cell = ws.cell(row=totais_row, column=col_idx)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = totals_border
            if col_idx == 1:
                cell.alignment = Alignment(horizontal="right", vertical="center")
            elif col_idx in [3, 4, 5]:
                cell.alignment = Alignment(horizontal="center", vertical="center")
            else:
                cell.alignment = Alignment(horizontal="left", vertical="center")
            
        ws.row_dimensions[totais_row].height = 26
        
        col_widths = {
            "A": 22, 
            "B": 48, 
            "C": 20, 
            "D": 38, 
            "E": 22, 
            "F": 65, # Coluna larga para o texto da Caixa Postal
            "G": 55  # Observações
        }
        for col_letter, width in col_widths.items():
            ws.column_dimensions[col_letter].width = width
            
        # Try saving the workbook. If there's a PermissionError (e.g. file open in Excel), wait and retry.
        saved = False
        for attempt in range(4):
            try:
                wb.save(output_path)
                saved = True
                log(f"Painel Excel Consolidado salvo com sucesso: {output_path}", "SUCCESS")
                break
            except PermissionError:
                if attempt < 3:
                    log(f"O arquivo Excel '{output_path}' esta aberto ou bloqueado. Por favor, feche o Excel! Tentando novamente em 5 segundos (tentativa {attempt + 1}/4)...", "WARNING")
                    time.sleep(5)
                else:
                    log(f"Nao foi possivel salvar o arquivo Excel '{output_path}' pois ele esta bloqueado. Por favor, feche o programa e tente novamente.", "ERROR")
            except Exception as save_err:
                log(f"Erro inesperado ao salvar o arquivo Excel consolidado: {save_err}", "ERROR")
                break
                
        if not saved:
            raise IOError(f"Nao foi possivel salvar a planilha Excel no caminho '{output_path}'. Verifique permissoes ou se o arquivo esta aberto.")
    except Exception as e:
        log(f"Falha critica ao gerar arquivo consolidado Excel: {e}", "ERROR")

def alterar_perfil(page, cnpj, procurador_cnpj):
    is_cpf = (len(cnpj) == 11)
    doc_tipo = "CPF" if is_cpf else "CNPJ"
    log(f"Alterando perfil de acesso para o {doc_tipo}: {cnpj}...", "ACTION")
    
    # 0. Limpar modais de aviso/mensagem (como aviso de caixa postal)
    fechar_modais_e_overlays(page)

    # 1. Clicar no botão "Alterar perfil de acesso"
    alterar_perfil_btn = None
    for selector in ["text=Alterar perfil de acesso", "#alterarPerfil", "a:has-text('Alterar')", ".btn-alterar-perfil"]:
        try:
            loc = page.locator(selector).first
            loc.wait_for(state="visible", timeout=3000)
            alterar_perfil_btn = loc
            break
        except Exception:
            continue
            
    if not alterar_perfil_btn:
        raise Exception("Não foi possível localizar o botão 'Alterar perfil de acesso' no cabeçalho.")
        
    try:
        alterar_perfil_btn.click(timeout=5000)
        page.wait_for_timeout(1000)
    except Exception as click_err:
        log(f"Erro ao clicar no botão de perfil: {click_err}. Tentando limpar overlays de recuperação...", "WARNING")
        fechar_modais_e_overlays(page)
        # Tenta localizar e clicar novamente
        alterar_perfil_btn = page.locator("text=Alterar perfil de acesso").first
        alterar_perfil_btn.click(timeout=5000)
        page.wait_for_timeout(1000)
    
    # Aguarda o modal de alteração de perfil estar 100% visível na tela
    log("Aguardando exibição do modal de alteração de perfil...", "ACTION")
    modal_carregado = False
    for sel in ["text=Procurador de pessoa jurídica", "text=Procurador de Pessoa Jurídica", "text=Procurador de pessoa física", "text=Procurador de Pessoa Física"]:
        try:
            loc = page.locator(sel).first
            loc.wait_for(state="visible", timeout=4000)
            modal_carregado = True
            log(f"Modal carregado com sucesso (elemento '{sel}' visível)", "SUCCESS")
            break
        except Exception:
            continue
            
    if not modal_carregado:
        log("Aviso: Tempo limite excedido ao aguardar visibilidade do modal. Tentando prosseguir diretamente.", "WARNING")
        
    # Se o CNPJ/CPF de destino for o próprio procurador, mudamos para o perfil Titular usando o botão dedicado
    if procurador_cnpj and cnpj == procurador_cnpj:
        log("Mudando de volta para o perfil Titular...", "ACTION")
        try:
            btn_titular = None
            for selector in [
                "input[value='Titular']",
                "button:has-text('Titular')",
                "text=Titular",
                "text=titular",
                "//input[@value='Titular']",
                "//button[contains(text(), 'Titular')]"
            ]:
                try:
                    loc = page.locator(selector).first
                    if loc.is_visible():
                        btn_titular = loc
                        break
                except Exception:
                    continue
                    
            if not btn_titular:
                raise Exception("Botão 'Titular' não localizado na tela.")
                
            btn_titular.click(timeout=5000)
            log("Botão 'Titular' clicado com sucesso para retornar ao perfil original.", "SUCCESS")
            page.wait_for_timeout(2000)
            return
        except Exception as e:
            raise Exception(f"Não foi possível reverter para o perfil Titular: {e}")
            
    # 2. Localizar o campo de entrada (CPF ou CNPJ) de Procurador e seu respectivo botão Alterar
    input_doc = None
    btn_alterar = None
    
    # Determinar labels e índice de fallback com base no tipo de documento
    labels_busca = ["Procurador de pessoa física", "Procurador de Pessoa Física"] if is_cpf else ["Procurador de pessoa jurídica", "Procurador de Pessoa Jurídica"]
    idx_fallback = 0 if is_cpf else 1
    
    # Método 1: Busca baseada em relação ao texto da label (Ultra-resiliente e preciso)
    for text_val in labels_busca:
        try:
            # Encontra o primeiro input de texto após a label correspondente
            inp_xpath = f"xpath=(//*[contains(text(), '{text_val}')]/following::input)[1]"
            # Encontra o primeiro botão de Alterar após a label correspondente
            btn_xpath = f"xpath=(//*[contains(text(), '{text_val}')]/following::input[@value='Alterar' or @type='submit' or contains(@value, 'Alterar')])[1]"
            
            inp = page.locator(inp_xpath).first
            btn = page.locator(btn_xpath).first
            
            inp.wait_for(state="visible", timeout=3000)
            btn.wait_for(state="visible", timeout=3000)
            
            input_doc = inp
            btn_alterar = btn
            log(f"Campos do perfil {doc_tipo} localizados com sucesso via XPath baseado em label: '{text_val}'", "INFO")
            break
        except Exception:
            continue
            
    # Método 2: Fallback por índice global (O primeiro input para CPF e o segundo input para CNPJ)
    if not input_doc or not btn_alterar:
        log(f"Busca por label falhou ou campos não ficaram visíveis. Usando fallback de índice global para {doc_tipo}...", "WARNING")
        try:
            inputs = page.locator("input[type='text'], input:not([type])")
            buttons = page.locator("input[value='Alterar'], input[type='submit'], button:has-text('Alterar')")
            
            if inputs.count() > idx_fallback and buttons.count() > idx_fallback:
                input_doc = inputs.nth(idx_fallback)
                btn_alterar = buttons.nth(idx_fallback)
                log(f"Campos do perfil {doc_tipo} localizados via fallback de índice global (índice {idx_fallback}).", "SUCCESS")
        except Exception as e:
            log(f"Falha ao executar fallback de índice global: {e}", "ERROR")
            
    if not input_doc or not btn_alterar:
        raise Exception(f"Não foi possível localizar o campo de texto do {doc_tipo} ou o botão 'Alterar' correspondente no modal.")
        
    # 3. Preencher o CPF/CNPJ no campo correto
    try:
        input_doc.clear()
        input_doc.fill(cnpj)
        log(f"Campo de {doc_tipo} de Procurador preenchido com: {cnpj}", "ACTION")
    except Exception as e:
        raise Exception(f"Falha ao preencher o campo de {doc_tipo} do procurador: {e}")
        
    # 4. Clicar no respectivo botão Alterar da linha
    try:
        btn_alterar.click(timeout=5000)
        log(f"Botão 'Alterar' correspondente ao perfil {doc_tipo} clicado com sucesso.", "ACTION")
    except Exception as e:
        raise Exception(f"Falha ao clicar no botão 'Alterar' do perfil {doc_tipo}: {e}")
        
    # 5. Validar se o perfil foi realmente alterado de forma dinâmica
    log("Aguardando atualização do cabeçalho com o novo perfil...", "ACTION")
    
    cnpj_sem_formatacao = cnpj.replace(".", "").replace("/", "").replace("-", "")
    sucesso = False
    
    # Fazemos uma verificação em loop a cada segundo por até 10 segundos
    for i in range(10):
        # Verifica se o CNPJ agora aparece no corpo/cabeçalho da página (indicando sucesso)
        header_text = page.locator("body").inner_text()
        header_text_limpo = header_text.replace(".", "").replace("/", "").replace("-", "")
        
        if cnpj_sem_formatacao in header_text_limpo:
            sucesso = True
            break
            
        # Também verifica se há alguma mensagem de erro de procuração ou do e-CAC explícita na tela durante a espera
        try:
            body_text = page.locator("body").inner_text()
            body_text_l = body_text.lower()
            
            # Procurar por linhas ou trechos de atenção no corpo do e-CAC
            for line in body_text.split("\n"):
                line_s = line.strip()
                line_l = line_s.lower()
                if not line_s:
                    continue
                
                is_atencao = ("atenção" in line_l or "atencao" in line_l or "atençâo" in line_l)
                is_procuracao = ("procuração" in line_l or "procuracao" in line_l or "procurador" in line_l)
                is_cancelada_inexistente = ("cancelada" in line_l or "inexistente" in line_l or "expirada" in line_l or "não cadastrada" in line_l or "revogada" in line_l)
                
                if (is_atencao and (is_procuracao or is_cancelada_inexistente)) or (is_procuracao and is_cancelada_inexistente):
                    raise Exception(f"Erro do e-CAC: {line_s}")
            
            # Fallback para o caso de o erro estar estruturado de outra forma
            if "procuração" in body_text_l or "procuracao" in body_text_l:
                if "cancelada" in body_text_l or "inexistente" in body_text_l or "expirada" in body_text_l:
                    raise Exception("Erro do e-CAC: Procuração cancelada ou inexistente detectada na página.")
                    
        except Exception as body_err:
            if "Erro do e-CAC" in str(body_err):
                raise body_err

        # Outros seletores clássicos de erro se estiverem visíveis
        for erro_selector in [".mensagem-erro", "#erro", ".alert-danger", ".msg-erro"]:
            try:
                loc = page.locator(erro_selector).first
                if loc.is_visible():
                    erro_msg = loc.inner_text().strip()
                    if erro_msg:
                        raise Exception(f"Erro do e-CAC: {erro_msg}")
            except Exception as e:
                if "Erro do e-CAC" in str(e):
                    raise e
                    
        page.wait_for_timeout(1000)
        
    if sucesso:
        log(f"Perfil alterado com sucesso para o cliente: {cnpj}", "SUCCESS")
        return True
    else:
        # Se falhou e esgotou os 10 segundos, extrai o texto visível para relatar a causa exata
        page_text = page.locator("body").inner_text()
        for palavra_chave in ["não cadastrada", "expirada", "inexistente", "erro", "restrição", "inválido", "restricao", "procuração", "automatizado", "automatizada"]:
            if palavra_chave in page_text.lower():
                raise Exception(f"A alteração de perfil falhou. Mensagem do e-CAC: '{page_text.strip()[:180]}...'")
                
        raise Exception("A alteração de perfil foi concluída, mas o CNPJ correspondente não foi identificado no cabeçalho.")

def baixar_certidao_regularidade_fiscal(page, context, client_dir, cnpj, config):
    log("Iniciando emissão da Certidão de Regularidade Fiscal no e-CAC...", "ACTION")
    
    # 1. Voltar para a aba principal do e-CAC e clicar em "Certidões e Situação Fiscal"
    aba_clicada = False
    for selector in [
        "text=Certidões e Situação Fiscal",
        "#btn198",
        "//span[contains(text(), 'Certidões e Situação Fiscal')]",
        "//a[contains(text(), 'Certidões e Situação Fiscal')]"
    ]:
        try:
            loc = page.locator(selector).first
            if loc.is_visible(timeout=3000):
                loc.click()
                aba_clicada = True
                break
        except Exception:
            continue
            
    page.wait_for_timeout(1500)
    
    # 2. Localizar e clicar no link de Emissão de Certidão
    link_clicado = False
    new_page = None
    
    for selector in [
        "text=Certidão de Débitos Relativos a Créditos Tributários Federais e à Dívida Ativa da União",
        "text=Emitir Certidão de Regularidade Fiscal",
        "text=Certidão de Regularidade Fiscal",
        "a[href*='certidao']",
        "a[href*='Certidao']",
        "//a[contains(text(), 'Créditos Tributários Federais')]"
    ]:
        try:
            loc = page.locator(selector).first
            if loc.is_visible(timeout=3000):
                # Esperar a abertura da nova aba ao clicar
                with context.expect_page() as new_page_info:
                    loc.click()
                new_page = new_page_info.value
                link_clicado = True
                log(f"Link de certidão localizado e clicado: '{selector}'", "INFO")
                break
        except Exception:
            continue
            
    if not link_clicado or not new_page:
        log("Aviso: Link direto no e-CAC para Emissão de Certidão não localizado. Tentando emitir pela página pública da Receita...", "WARNING")
        try:
            new_page = context.new_page()
            new_page.goto("https://solucoes.receita.fazenda.gov.br/Servicos/certidaointernet/PJ/Emitir", timeout=30000)
        except Exception as fallback_err:
            log(f"Falha ao carregar página de fallback pública: {fallback_err}", "ERROR")
            raise Exception("Não foi possível acessar a área de emissão de Certidão.")
        
    new_page.wait_for_load_state("load")
    
    # 3. Na página de emissão da certidão, preencher o CNPJ e emitir
    log("Aguardando carregamento da página de emissão de certidão...", "ACTION")
    try:
        # Se for a página pública de emissão da PJ, ela terá um campo para CNPJ
        input_cnpj = new_page.locator("input#cnpj, input[name*='cnpj'], input[name*='CNPJ']").first
        if input_cnpj.is_visible(timeout=3000):
            input_cnpj.clear()
            input_cnpj.fill(cnpj)
            log("CNPJ preenchido no formulário de emissão de certidão.", "ACTION")
            
            # Clicar no botão de Emitir do formulário
            btn_emitir = new_page.locator("input[value='Emitir'], button:has-text('Emitir'), input[type='submit']").first
            btn_emitir.click()
            new_page.wait_for_load_state("load")
            
        # Localizar botão de download / imprimir
        log("Localizando botão de download da certidão PDF...", "ACTION")
        botao_download = None
        for selector in [
            "text='Imprimir'",
            "text='Download'",
            "text='PDF'",
            "button:has-text('Imprimir')",
            "button:has-text('Gerar')",
            "button:has-text('Emitir')",
            "a:has-text('Imprimir')",
            "//input[@value='Imprimir']",
            "//input[@value='Gerar']"
        ]:
            try:
                loc = new_page.locator(selector).first
                if loc.is_visible(timeout=3000):
                    botao_download = loc
                    break
            except Exception:
                continue
                
        if not botao_download:
            raise Exception("Botão de download da certidão não localizado na página de emissão.")
            
        with new_page.expect_download(timeout=config["download_timeout_ms"]) as download_info:
            botao_download.click()
        download = download_info.value
        
        hoje_limpo = datetime.date.today().strftime("%Y%m%d")
        pdf_filename = f"CertidaoRegularidadeFiscal-{cnpj}-{hoje_limpo}.pdf"
        pdf_path = os.path.join(client_dir, pdf_filename)
        
        download.save_as(pdf_path)
        log(f"Certidão de regularidade fiscal baixada com sucesso: {pdf_path}", "SUCCESS")
        remover_arquivos_fiscais_obsoletos(client_dir, pdf_path)
        new_page.close()
        return f"Certidão Baixada: {pdf_filename}"
        
    except Exception as e:
        try:
            new_page.close()
        except Exception:
            pass
        log(f"Falha ao baixar a Certidão: {e}. Salvando informativo TXT como fallback.", "WARNING")
        # Fallback caso dê timeout ou erro de rede
        cnd_path = os.path.join(client_dir, f"Sem_Pendencias_Fiscais_Regular-{cnpj}.txt")
        with open(cnd_path, "w", encoding="utf-8") as f:
            f.write(f"Consulta realizada em {datetime.date.today().strftime('%d/%m/%Y')} às {datetime.datetime.now().strftime('%H:%M:%S')}\n")
            f.write("Status: Regular - Nenhuma pendência fiscal encontrada na Receita Federal (Emissão manual recomendada).")
        remover_arquivos_fiscais_obsoletos(client_dir, cnd_path)
        return "Sem Pendências (Informativo Gravado)"

def baixar_relatorio_situacao_fiscal(page, context, client_dir, cnpj, config, ja_possui_relatorio=False):
    log("Navegando para o serviço de Consulta de Situação Fiscal...", "ACTION")
    
    # 1. Clicar na aba "Certidões e Situação Fiscal" no painel principal do e-CAC
    aba_clicada = False
    for selector in [
        "text=Certidões e Situação Fiscal",
        "#btn198", # ID comum do botão de certidões no e-CAC
        "//span[contains(text(), 'Certidões e Situação Fiscal')]",
        "//a[contains(text(), 'Certidões e Situação Fiscal')]"
    ]:
        try:
            # Click direto utilizando o auto-waiting nativo do Playwright de até 4s
            page.click(selector, timeout=4000)
            aba_clicada = True
            break
        except Exception:
            continue
            
    if not aba_clicada:
        raise Exception("Não foi possível acessar o menu 'Certidões e Situação Fiscal'.")
        
    page.wait_for_timeout(1000)
    
    # 2. Localizar e clicar em "Consulta Pendências - Situação Fiscal" (abre em nova aba)
    link_clicado = False
    new_page = None
    
    for selector in [
        "a:has-text('Consulta Pendências - Situação Fiscal')",
        "text=Consulta Pendências - Situação Fiscal",
        "a[href*='servico/pendencias']",
        "//a[contains(text(), 'Consulta Pendências')]"
    ]:
        try:
            # Esperar a abertura da nova aba ao clicar (definimos timeout de 10s)
            with context.expect_page(timeout=10000) as new_page_info:
                # Usamos dispatch_event('click') para simular o clique a nível de evento DOM
                # Isso evita que o cursor físico dispare o hover do tooltip que cobre o link
                page.locator(selector).first.dispatch_event("click")
            new_page = new_page_info.value
            link_clicado = True
            break
        except Exception as click_err:
            log(f"Falha ao clicar com dispatch_event no seletor '{selector}': {click_err}. Tentando clique físico...", "WARNING")
            try:
                with context.expect_page(timeout=10000) as new_page_info:
                    page.locator(selector).first.click(timeout=4000)
                new_page = new_page_info.value
                link_clicado = True
                break
            except Exception:
                continue
            
    if not link_clicado or not new_page:
        raise Exception("Não foi possível clicar em 'Consulta Pendências - Situação Fiscal' ou a nova aba não abriu.")
        
    log("Nova aba de Consulta de Situação Fiscal detectada.", "ACTION")
    new_page.wait_for_load_state("load")
    
    # 2.5 Verificar se a nova página redirecionou para uma tela de login/autenticação intermediária (Gov.br)
    # Isso costuma ocorrer quando mudamos de domínio de cav.receita.fazenda.gov.br para servicos.receitafederal.gov.br
    try:
        btn_gov = None
        for sel in ["text=Entrar com GovBR", "text=Entrar com o GovBR", "text=Entrar com o gov.br", "button:has-text('Entrar com')", "a:has-text('Entrar com')"]:
            try:
                loc = new_page.locator(sel).first
                if loc.is_visible(timeout=1000):
                    btn_gov = loc
                    break
            except Exception:
                continue
        if btn_gov:
            log("Tela de autenticação intermediária do Gov.br detectada no novo domínio. Clicando em 'Entrar com GovBR'...", "WARNING")
            btn_gov.click()
            # Aguarda a nova página carregar após o redirecionamento de login
            new_page.wait_for_load_state("load")
            new_page.wait_for_timeout(3000)
    except Exception as e:
        log(f"Sem tela de login intermediária do Gov.br ou falha ao clicar: {e}. Prosseguindo...", "INFO")
        
    # 2.7 Trocar Representação diretamente no Novo Portal (se necessário)
    try:
        avatar_trigger = new_page.locator("#avatar-dropdown-trigger, .br-sign-in").first
        if avatar_trigger.is_visible(timeout=4000):
            header_text = avatar_trigger.inner_text()
            cnpj_limpo = "".join(filter(str.isdigit, cnpj))
            header_cnpj_limpo = "".join(filter(str.isdigit, header_text))
            
            # Se o CNPJ ativo no cabeçalho do novo portal for diferente do CNPJ procurado, fazemos a troca
            if cnpj_limpo not in header_cnpj_limpo:
                log(f"[NOVO-PORTAL] CNPJ ativo no novo portal é diferente do alvo ({cnpj}). Iniciando troca...", "INFO")
                
                # Clicar com time de 2 segundos no trigger do avatar
                avatar_trigger.click()
                new_page.wait_for_timeout(2000)
                
                # Escrever o CNPJ ou CPF do cliente no input
                log(f"[NOVO-PORTAL] Preenchendo CNPJ de representação: {cnpj}...", "ACTION")
                cnpj_input = new_page.locator("#input-representar-cpfcnpj, input[name='representar-cpfcnpj']").first
                cnpj_input.clear()
                cnpj_input.fill(cnpj)
                
                # Aguardar 1 segundo
                new_page.wait_for_timeout(1000)
                
                # Clicar no dropdown de perfil de representação
                log("[NOVO-PORTAL] Clicando no dropdown de perfil de representação...", "ACTION")
                select_profile = new_page.locator("ng-select, .ng-select-container, .ng-placeholder").first
                select_profile.click()
                new_page.wait_for_timeout(500)
                
                # Selecionar "PROCURADOR"
                opcao_procurador = new_page.locator(".ng-option:has-text('PROCURADOR'), .ng-option:has-text('Procurador'), text='PROCURADOR'").first
                opcao_procurador.click()
                new_page.wait_for_timeout(500)
                
                # Clicar no botão Representar
                log("[NOVO-PORTAL] Clicando no botão Representar...", "ACTION")
                btn_representar = new_page.locator("button:has-text('Representar'), .btn:has-text('Representar'), [role='button']:has-text('Representar')").first
                btn_representar.click()
                
                # Aguardar 4 segundos para a página recarregar a representação
                new_page.wait_for_timeout(4000)
                
                # Validar se a alteração foi refletida no cabeçalho
                header_text_after = avatar_trigger.inner_text()
                header_cnpj_after = "".join(filter(str.isdigit, header_text_after))
                if cnpj_limpo not in header_cnpj_after:
                    raise Exception(f"CNPJ ativo após a troca de representação ({header_cnpj_after}) não corresponde ao CNPJ esperado ({cnpj_limpo}).")
                
                # Clicar no centro da página (body) para conferir/remover dropdown e prosseguir
                new_page.click("body")
                new_page.wait_for_timeout(1000)
                log(f"[NOVO-PORTAL] Representação alterada com sucesso para o CNPJ: {cnpj}", "SUCCESS")
            else:
                log(f"[NOVO-PORTAL] Representação ativa no novo portal já corresponde ao CNPJ {cnpj}.", "SUCCESS")
    except Exception as rep_err:
        log(f"[NOVO-PORTAL] Erro crítico ao gerenciar representação diretamente no novo portal: {rep_err}", "ERROR")
        raise Exception(f"Não foi possível estabelecer a representação no novo portal para o CNPJ {cnpj}: {rep_err}")

    # 3. Aguardar o carregamento dos dados da situação fiscal
    # Esta página costuma fazer consultas lentas em APIs internas. Vamos aguardar pacientemente.
    log("Aguardando carregamento da situação fiscal do cliente na Receita Federal (isso pode levar alguns segundos)...", "ACTION")
    
    # Espera até 40 segundos no total com polling ativo para avançar IMEDIATAMENTE quando a página carregar
    selectors = [
        "text=Gerar Relatório",
        "text=Baixar Relatório",
        "text=Baixar Certidão",
        "text=não foram encontradas",
        "text=regularidade",
        "text=Sem pendência",
        "button:has-text('Relatório')"
    ]
    
    carregado = False
    for _ in range(40): # 40 segundos no máximo
        # Verificar se a tela de erro temporário JAPE (107.6) apareceu
        try:
            page_text = new_page.locator("body").inner_text()
            if "Não foi possível concluir a ação para o contribuinte" in page_text or "107.6 -" in page_text:
                log("[JAPE] Erro temporário do e-CAC (107.6) detectado para este contribuinte. Capturando screenshot e pulando...", "WARNING")
                
                # Capturar screenshot do erro na pasta do cliente
                screenshot_path = os.path.join(client_dir, "erro_jape.png")
                try:
                    new_page.screenshot(path=screenshot_path)
                    log(f"Screenshot do erro JAPE salvo em: {screenshot_path}", "INFO")
                    remover_arquivos_fiscais_obsoletos(client_dir, screenshot_path)
                except Exception as snap_err:
                    log(f"Não foi possível capturar screenshot do erro JAPE: {snap_err}", "WARNING")
                
                # Fechar aba do relatório
                try:
                    new_page.close()
                except Exception:
                    pass
                
                raise JapeException("Não foi possível concluir a ação para o contribuinte (Erro 107.6 / JAPE)")
        except JapeException:
            raise
        except Exception:
            pass

        for sel in selectors:
            try:
                if new_page.locator(sel).first.is_visible():
                    carregado = True
                    break
            except Exception:
                continue
        if carregado:
            break
        new_page.wait_for_timeout(1000)

    if carregado:
        log("Carregamento da situação fiscal concluído!", "SUCCESS")
    else:
        log("Aviso: Tempo limite de carregamento de 40 segundos atingido. Verificando elementos presentes na página...", "WARNING")
        
    # 4. Analisar se o cliente tem pendências ou se está regular
    page_text = new_page.locator("body").inner_text()
    
    # Se houver a mensagem amarela exigindo atualização da consulta, clicar em "Atualizar"
    if "necessário gerar uma nova consulta" in page_text or "situação fiscal, acionando a opção" in page_text:
        log("Alerta de consulta desatualizada detectado! Clicando em 'Atualizar'...", "WARNING")
        try:
            btn_atualizar = None
            for selector in [
                "button:has-text('Atualizar')",
                "text=Atualizar",
                "//button[contains(., 'Atualizar')]"
            ]:
                loc = new_page.locator(selector).first
                if loc.is_visible(timeout=3000):
                    btn_atualizar = loc
                    break
            
            if btn_atualizar:
                btn_atualizar.click()
                log("Botão 'Atualizar' clicado com sucesso. Aguardando processamento da Receita...", "INFO")
                new_page.wait_for_timeout(3000)
                # Esperar até que o alerta amarelo desapareça (máximo 30 segundos)
                for _ in range(30):
                    banner_visivel = False
                    try:
                        banner_visivel = new_page.locator("text=gerar uma nova consulta").first.is_visible()
                    except Exception:
                        pass
                    if not banner_visivel:
                        log("Nova consulta de situação fiscal concluída com sucesso!", "SUCCESS")
                        break
                    new_page.wait_for_timeout(1000)
                
                # Atualizar o page_text com a nova situação
                page_text = new_page.locator("body").inner_text()
            else:
                log("Aviso: Alerta de consulta desatualizada presente, mas botão 'Atualizar' não localizado.", "WARNING")
        except Exception as e_act:
            log(f"Erro ao tentar clicar no botão 'Atualizar': {e_act}", "WARNING")
            
    hoje_limpo = datetime.date.today().strftime("%Y%m%d")
    
    # Verificação se o status é "Sem pendência"
    if "não foram encontradas" in page_text.lower() or "sem pendência" in page_text.lower() or "regularidade" in page_text.lower():
        log("Mensagem de ausência de pendências detectada ('Sem pendência' / 'Regularidade')!", "SUCCESS")
        log("Buscando botão de download da CERTIDÃO (CND)...", "ACTION")
        
        botao_cnd = None
        for selector in [
            "app-botao-emitir-certidao button",
            "button.secondary.btn-acao:has-text('Baixar Certidão')",
            "button.btn-acao:has-text('Baixar Certidão')",
            "text=Baixar Certidão",
            "button:has-text('Baixar Certidão')",
            "a:has-text('Baixar Certidão')",
            "//button[contains(., 'Certidão')]",
            "//a[contains(., 'Certidão')]"
        ]:
            try:
                loc = new_page.locator(selector).first
                if loc.is_visible(timeout=3000):
                    botao_cnd = loc
                    break
            except Exception:
                continue
                
        if botao_cnd:
            log("Botão 'Baixar Certidão' localizado. Iniciando download do PDF da CND...", "ACTION")
            try:
                with new_page.expect_download(timeout=config["download_timeout_ms"]) as download_info:
                    botao_cnd.click()
                download = download_info.value
                
                pdf_filename = f"CertidaoRegularidadeFiscal-{cnpj}-{hoje_limpo}.pdf"
                pdf_path = os.path.join(client_dir, pdf_filename)
                download.save_as(pdf_path)
                log(f"Certidão de regularidade fiscal baixada com sucesso: {pdf_path}", "SUCCESS")
                remover_arquivos_fiscais_obsoletos(client_dir, pdf_path)
                new_page.close()
                return "Certidão Baixada"
            except Exception as e:
                log(f"Falha ao clicar no botão 'Baixar Certidão': {e}. Tentando fallback...", "WARNING")
                
        # Se não localizou o botão ou falhou o download na página de pendências, tenta emitir pela aba tradicional do e-CAC
        log("Tentando emitir CND via menu tradicional do e-CAC como fallback...", "INFO")
        new_page.close()
        resultado_fallback = baixar_certidao_regularidade_fiscal(page, context, client_dir, cnpj, config)
        return resultado_fallback
        
    else:
        # COM PENDÊNCIAS - Baixar o Relatório PDF ou ignorar se já existe no mês
        if ja_possui_relatorio:
            log("A situação fiscal continua com pendências e o relatório do mês corrente já existe na pasta. Ignorando novo download redundante.", "INFO")
            new_page.close()
            return "Relatório Existente (Pendências mantidas)"
            
        log("Pendências fiscais ativas detectadas. Buscando botão de download do RELATÓRIO...", "ACTION")
        
        botao_relatorio = None
        for selector in [
            "text=Baixar Relatório",
            "text='Gerar Relatório'",
            "text='Baixar Relatório'",
            "button:has-text('Baixar Relatório')",
            "button:has-text('Gerar Relatório')",
            "button:has-text('Relatório')",
            "button:has-text('PDF')",
            "a:has-text('Gerar Relatório')",
            "//button[contains(., 'Relatório')]",
            "//button[contains(., 'Gerar')]"
        ]:
            try:
                loc = new_page.locator(selector).first
                if loc.is_visible(timeout=3000):
                    botao_relatorio = loc
                    break
            except Exception:
                continue
                
        if not botao_relatorio:
            raise Exception("Botão 'Baixar Relatório' / 'Gerar Relatório' não localizado na página de pendências.")
            
        log("Botão de relatório localizado. Iniciando download do PDF do Relatório...", "ACTION")
        try:
            with new_page.expect_download(timeout=config["download_timeout_ms"]) as download_info:
                botao_relatorio.click()
            download = download_info.value
            
            pdf_filename = f"RelatorioSituacaoFiscal-{cnpj}-{hoje_limpo}.pdf"
            pdf_path = os.path.join(client_dir, pdf_filename)
            
            download.save_as(pdf_path)
            log(f"Relatório fiscal baixado e salvo com sucesso: {pdf_path}", "SUCCESS")
            remover_arquivos_fiscais_obsoletos(client_dir, pdf_path)
            new_page.close()
            return "Relatório Baixado"
        except Exception as e:
            new_page.close()
            raise Exception(f"Falha ao realizar download do PDF do Relatório: {e}")

def realizar_login_manual(config):
    # Forçar a limpeza logo no início do login manual por segurança
    limpar_processos_automatizados_antigos()

    log("=" * 60, "LOGIN")
    log(" INICIANDO MODO DE AUTENTICAÇÃO E-CAC (100% AUTOMATIZADO)", "LOGIN")
    log("=" * 60, "LOGIN")
    
    browser_choice = config.get("browser", "chrome").lower()
    log(f"Navegador configurado para login manual: {browser_choice.upper()}", "INFO")
    
    state_file = "state.json"
    
    with sync_playwright() as p:
        context = None
        
        # Define os argumentos de inicialização
        launch_args = {
            "headless": False,
            "args": [
                "--disable-blink-features=AutomationControlled",
                "--disable-infobars",
                "--disable-session-crashed-bubble",
                "--disable-features=BubbleSessionCrashedBubble"
            ],
            "no_viewport": True
        }
        
        if browser_choice == "msedge":
            user_data_dir = os.path.join(os.getcwd(), "temp", "chrome_profile_msedge")
            launch_args["channel"] = "msedge"
            browser_name_readable = "Microsoft Edge nativo"
        elif browser_choice == "chrome":
            user_data_dir = os.path.join(os.getcwd(), "temp", "chrome_profile_chrome")
            launch_args["channel"] = "chrome"
            browser_name_readable = "Google Chrome nativo"
        else:
            user_data_dir = os.path.join(os.getcwd(), "temp", "chrome_profile_chromium")
            browser_name_readable = "Chromium padrão"
            
        launch_args["user_data_dir"] = user_data_dir
        os.makedirs(user_data_dir, exist_ok=True)
        
        # Tenta duas vezes (com limpeza preventiva entre elas se falhar)
        for tentativa in range(1, 3):
            try:
                log(f"Iniciando {browser_name_readable} com perfil persistente e modo stealth (Tentativa {tentativa}/2)...", "INFO")
                context = p.chromium.launch_persistent_context(**launch_args)
                log(f"Navegador {browser_name_readable} iniciado com sucesso.", "SUCCESS")
                break
            except Exception as e:
                log(f"Erro ao iniciar o {browser_name_readable} na tentativa {tentativa}: {e}", "WARNING")
                if tentativa == 1:
                    log("Executando limpeza de processos órfãos e arquivos de lock para tentar novamente...", "WARNING")
                    limpar_processos_automatizados_antigos()
                else:
                    raise Exception(f"Falha persistente ao abrir o navegador {browser_name_readable} após duas tentativas. Verifique se o navegador está instalado ou aberto em outra tarefa: {e}")
            
        page = context.pages[0] if context.pages else context.new_page()
        
        # Acessa e-CAC diretamente
        log("Acessando a página de login do e-CAC diretamente...", "INFO")
        try:
            page.goto("https://cav.receita.fazenda.gov.br/ecac/", timeout=30000)
            page.wait_for_load_state("load")
        except Exception as goto_err:
            log(f"Erro ao acessar diretamente: {goto_err}. Tentando via busca orgânica do Google...", "WARNING")
            navegar_via_google_para_ecac(page)
        
        # 1. Aguarda e clica automaticamente no botão de login do Gov.br
        try:
            log("Buscando o botão de login 'Acesso Gov BR'...", "INFO")
            page.wait_for_selector('input[alt="Acesso Gov BR"]', timeout=10000)
            log("Clicando no botão de login do Gov.br...", "ACTION")
            page.click('input[alt="Acesso Gov BR"]')
            
            # 2. Aguarda redirecionamento para o Gov.br SSO e o botão "Seu certificado digital"
            log("Aguardando redirecionamento para o portal Gov.br...", "INFO")
            page.wait_for_selector('button#login-certificate, #login-certificate', timeout=15000)
            log("Botão 'Seu certificado digital' (#login-certificate) detectado!", "SUCCESS")
            
            # 3. Disparar thread em background para simular o ENTER no diálogo do Windows com monitoramento em loop
            import threading
            threading.Thread(
                target=executar_confirmacao_certificado_em_loop,
                args=(config, "[AUTO-LOGIN]"),
                daemon=True
            ).start()
            
            log("Clicando no botão 'Seu certificado digital' e aguardando confirmação do Windows...", "ACTION")
            try:
                page.click('button#login-certificate', timeout=5000)
            except Exception as e_click:
                log(f"Aviso no clique do certificado (TLS travou a thread temporariamente, esperado): {e_click}", "INFO")
            
        except Exception as auto_err:
            log(f"Aviso no fluxo automático: {auto_err}", "WARNING")
            log("Fluxo automático parou. Por favor, conclua o login manualmente na tela se necessário.", "IMPORTANT")
            
        try:
            # Tenta aguardar o login automático rápido por 12 segundos primeiro
            page.wait_for_selector("text=Alterar perfil de acesso", timeout=12000)
            log("Conexão e login detectados com sucesso no e-CAC de forma rápida!", "SUCCESS")
        except Exception:
            # Se demorou mais de 12 segundos, provavelmente precisa de intervenção manual (CAPTCHA, etc.)
            log("O login automático não concluiu in 12 segundos. Provavelmente é necessário resolver CAPTCHA ou selecionar certificado. Enviando alerta via WhatsApp...", "WARNING")
            
            url_atual = "N/A"
            if page:
                try:
                    url_atual = page.url
                except Exception:
                    pass
            
            mensagem_alerta = (
                "⚠️ *ALERTA DO ROBÔ e-CAC (INÍCIO DE CICLO)*\n\n"
                "A automação precisa de intervenção para concluir o login no e-CAC!\n"
                f"• *Link da Página Atual*: {url_atual}\n"
                "• *Causa*: Provavelmente surgiu um CAPTCHA de imagem ou confirmação de certificado na tela.\n"
                "• *O que fazer*: Acesse a tela do navegador aberto, resolva o CAPTCHA e conclua o login Gov.br.\n\n"
                "O robô aguardará por até 5 minutos a conclusão do login."
            )
            enviar_whatsapp(mensagem_alerta, config)
            
            # Agora aguarda o tempo longo restante (288 segundos) para a conclusão manual
            try:
                page.wait_for_selector("text=Alterar perfil de acesso", timeout=288000)
                log("Conexão e login detectados com sucesso no e-CAC após intervenção manual!", "SUCCESS")
            except Exception as e:
                log(f"Tempo limite de 5 minutos excedido ou erro no login: {e}", "ERROR")
                context.close()
                return False
                
        # Salvar o estado da sessão em state.json
        try:
            context.storage_state(path=state_file)
            log(f"Sessão gravada com sucesso em '{state_file}'!", "SUCCESS")
        except Exception as e_save:
            log(f"Aviso ao gravar sessão: {e_save}", "WARNING")
        context.close()
        return True

def enviar_whatsapp(mensagem, config):
    if not config.get("whatsapp_enabled"):
        log("Notificação via WhatsApp desabilitada nas configurações.", "INFO")
        return False
        
    number = config.get("whatsapp_number")
    if not number:
        log("Erro: Número de telefone do WhatsApp não configurado.", "ERROR")
        return False
        
    def worker():
        url = "http://127.0.0.1:3000/api/send-message"
        payload = {
            "to": number,
            "message": mensagem
        }
        log("Enviando notificação WhatsApp via gateway local em segundo plano...", "INFO")
        try:
            response = requests.post(url, json=payload, timeout=15)
            if response.status_code in [200, 201]:
                log("Notificação via WhatsApp enviada com sucesso!", "SUCCESS")
            else:
                log(f"Falha ao enviar WhatsApp. Status: {response.status_code}, Resposta: {response.text}", "ERROR")
        except Exception as e:
            log(f"Erro ao enviar requisição HTTP do WhatsApp: {e}", "ERROR")

    import threading
    threading.Thread(target=worker, daemon=True).start()
    return True

def atualizar_excel_dinamico(clientes, config):
    try:
        desktop_dir = os.path.join(os.path.expanduser("~"), "Desktop")
        desktop_excel_fixed = os.path.join(desktop_dir, "Painel_Consolidado_Pendencias_eCAC.xlsx")
        # Atualiza diretamente o arquivo fixo e único na Area de Trabalho, sem gerar cópias datadas.
        gerar_consolidado_excel(clientes, config.get("relatorios_dir", "relatorios"), desktop_excel_fixed)
        log(f"Painel Excel Consolidado principal atualizado no Desktop: {desktop_excel_fixed}", "SUCCESS")
    except Exception as d_err:
        log(f"Erro ao atualizar painel Excel no Desktop: {d_err}", "ERROR")

def tentar_logout_ecac(page):
    if not page:
        return
    log("Tentando realizar logout seguro do e-CAC...", "ACTION")
    try:
        # Tenta localizar e clicar nos seletores comuns de Sair no e-CAC
        for selector in [
            "span.botaoAzul:has-text('Sair com Segurança')",
            "span:has-text('Sair com Segurança')",
            ".botaoAzul",
            "text='Sair com segurança'", "text='Sair'", "a:has-text('Sair')",
            "#btnSair", ".btn-sair", "//a[contains(text(), 'Sair')]"
        ]:
            loc = page.locator(selector).first
            if loc.is_visible(timeout=2000):
                loc.click(timeout=3000)
                log("Botão 'Sair' clicado com sucesso.", "SUCCESS")
                page.wait_for_timeout(1000)
                break
    except Exception as e:
        log(f"Aviso: Não foi possível clicar no botão de Sair (fechando diretamente): {e}", "WARNING")

def main():
    limpar_processos_automatizados_antigos()
    config = load_config()
    clientes = load_clients(config["clientes_file"])
    
    if not clientes:
        log("Nenhum cliente ativo para processamento. Finalizando robô.", "WARNING")
        return
        
    state_file = "state.json"
    
    while True:
        # Identificar procurador e letra inicial pelo arquivo .pfx (usamos para focar o certificado no Windows e evitar re-troca de perfil)
        import glob
        import re
        pfx_files = glob.glob("*.pfx")
        procurador_cnpj = ""
        cert_first_char = "J"
        if pfx_files:
            filename = os.path.basename(pfx_files[0])
            # Extrair CNPJ usando regex para buscar qualquer sequência de 14 dígitos consecutivos
            match = re.search(r"\d{14}", filename)
            if match:
                procurador_cnpj = match.group(0)
            
            # Extrair a primeira letra da razão social (removendo números iniciais e sublinhado)
            clean_name = re.sub(r"^\d+_", "", filename)
            if clean_name:
                cert_first_char = clean_name[0].upper()
                
        config["cert_first_char"] = cert_first_char
        log(f"CNPJ do Procurador detectado: {procurador_cnpj or 'Não identificado'}", "INFO")
        log(f"Caractere inicial do certificado detectado: '{cert_first_char}'", "INFO")
        
        # Se o usuário passou --login na linha de comando, ou se o arquivo state.json não existe, força o login manual
        forçar_login = "--login" in sys.argv
        if forçar_login or not os.path.exists(state_file):
            if not os.path.exists(state_file):
                log("Sessão ativa não encontrada ('state.json'). Precisamos logar uma vez para salvar o acesso.", "INFO")
            else:
                log("Forçando renovação da sessão ativa conforme solicitado...", "INFO")
                
            sucesso = realizar_login_manual(config)
            if not sucesso:
                log("Falha ao salvar a sessão autenticada. A execução da varredura foi cancelada.", "ERROR")
                return
        
        log(f"Iniciando rotina de processamento para {len([c for c in clientes if c['ativo']])} clientes ativos...", "INFO")
        
        success_count = 0
        failure_count = 0
        skipped_count = 0
        
        need_restart = False
        
        with sync_playwright() as p:
            context = None
            page = None
            need_browser_restart = False
            
            def iniciar_navegador():
                nonlocal context, page, need_browser_restart
                
                if context:
                    try:
                        log("Fechando navegador e-CAC anterior para iniciar uma sessão limpa...", "INFO")
                        context.close()
                    except Exception:
                        pass
                    context = None
                    page = None
                
                if need_browser_restart:
                    log("Aguardando 1 minuto antes de reabrir o navegador e-CAC...", "WARNING")
                    time.sleep(60)
                    need_browser_restart = False
                
                # Executa a limpeza preventiva de processos no início da inicialização
                limpar_processos_automatizados_antigos()
                
                browser_choice = config.get("browser", "chrome").lower()
                log(f"Iniciando navegador com perfil persistente ({browser_choice.upper()}) para varredura automática...", "SYSTEM")
                
                # Define os argumentos de inicialização
                launch_args = {
                    "headless": config["headless"],
                    "args": [
                        "--disable-blink-features=AutomationControlled",
                        "--disable-infobars",
                        "--disable-session-crashed-bubble",
                        "--disable-features=BubbleSessionCrashedBubble"
                    ],
                    "no_viewport": True
                }
                
                if browser_choice == "msedge":
                    user_data_dir = os.path.join(os.getcwd(), "temp", "chrome_profile_msedge")
                    launch_args["channel"] = "msedge"
                    browser_name_readable = "Microsoft Edge nativo"
                elif browser_choice == "chrome":
                    user_data_dir = os.path.join(os.getcwd(), "temp", "chrome_profile_chrome")
                    launch_args["channel"] = "chrome"
                    browser_name_readable = "Google Chrome nativo"
                else:
                    user_data_dir = os.path.join(os.getcwd(), "temp", "chrome_profile_chromium")
                    browser_name_readable = "Chromium padrão"
                    
                launch_args["user_data_dir"] = user_data_dir
                os.makedirs(user_data_dir, exist_ok=True)
                
                # Tenta duas vezes (com limpeza preventiva entre elas se falhar)
                for tentativa in range(1, 3):
                    try:
                        log(f"Abrindo {browser_name_readable} (Tentativa {tentativa}/2)...", "INFO")
                        context = p.chromium.launch_persistent_context(**launch_args)
                        log(f"Navegador {browser_name_readable} iniciado com sucesso.", "SUCCESS")
                        break
                    except Exception as e:
                        log(f"Erro ao iniciar o {browser_name_readable} na tentativa {tentativa}: {e}", "WARNING")
                        if tentativa == 1:
                            log("Executando limpeza de processos órfãos e arquivos de lock para tentar novamente...", "WARNING")
                            limpar_processos_automatizados_antigos()
                        else:
                            raise Exception(f"Falha persistente ao abrir o navegador {browser_name_readable} após duas tentativas. Verifique se o navegador está instalado ou aberto em outra tarefa: {e}")
                
                # Injetar os cookies salvos em state.json no contexto antes de acessar a página
                if os.path.exists(state_file):
                    try:
                        with open(state_file, "r", encoding="utf-8") as f:
                            state_data = json.load(f)
                            cookies = state_data.get("cookies", [])
                            if cookies:
                                context.add_cookies(cookies)
                                log("Cookies da sessão carregados com sucesso do state.json", "SUCCESS")
                    except Exception as e_cookies:
                        log(f"Aviso: Não foi possível injetar cookies do state.json: {e_cookies}", "WARNING")
                
                page = context.pages[0] if context.pages else context.new_page()
                
                # Acessar diretamente a página inicial do e-CAC (já autenticado!)
                log("Acessando portal e-CAC (reutilizando sessão ativa)...", "ACTION")
                try:
                    page.goto("https://cav.receita.fazenda.gov.br/ecac/", timeout=30000)
                    # Validar se a sessão ainda está ativa
                    page.wait_for_selector("text=Alterar perfil de acesso", timeout=15000)
                    log("Sessão autenticada validada e ativa com sucesso!", "SUCCESS")
                except Exception as e:
                    log(f"Sessão expirada ou erro ao acessar: {e}. Tentando reestabelecer a sessão automaticamente...", "WARNING")
                    if verificar_e_reestabelecer_sessao(page, config):
                        log("Sessão autenticada reestabelecida com sucesso na inicialização!", "SUCCESS")
                    else:
                        log("Falha crítica: Não foi possível reestabelecer a sessão do e-CAC de forma automática/manual.", "ERROR")
                        raise e
            
            # Tentar iniciar o navegador na primeira vez
            try:
                iniciar_navegador()
            except Exception:
                log("A sessão salva expirou ou foi invalidada pela Receita Federal. Precisamos renovar o login.", "WARNING")
                if os.path.exists(state_file):
                    try:
                        os.remove(state_file)
                    except Exception:
                        pass
                need_restart = True
            
            if not need_restart:
                # 2. Iterar sobre a lista de clientes
                for cliente in clientes:
                    cnpj = cliente["cnpj"]
                    nome = cliente["nome"]
                    ativo = cliente["ativo"]
                    
                    if not ativo:
                        log(f"Cliente {nome} ({cnpj}) está inativo no CSV. Ignorando.", "INFO")
                        skipped_count += 1
                        continue
                        
                    # Verificar se o cliente já foi processado com sucesso hoje ou se já possui regularidade
                    today_str = datetime.date.today().strftime("%Y-%m-%d")
                    nome_limpo = clean_filename(nome)
                    cnpj_limpo = "".join(filter(str.isdigit, cnpj))
                    client_dir = os.path.join(config["relatorios_dir"], f"{cnpj_limpo}_{nome_limpo}")
                    status_path = os.path.join(client_dir, "status.json")
                    
                    # Salva cópia do status antigo caso precise restaurar
                    old_status = None
                    if os.path.exists(status_path):
                        try:
                            with open(status_path, "r", encoding="utf-8") as f:
                                old_status = json.load(f)
                        except Exception:
                            pass
                            
                    forcar_todos = "--forcar-todos" in sys.argv
                    
                    # 1. Verificar se já existem arquivos de regularidade (Certidão ou Informativo Regular) na pasta do cliente
                    cnd_pdfs = glob.glob(os.path.join(client_dir, "CertidaoRegularidadeFiscal-*.pdf"))
                    cnd_txts = glob.glob(os.path.join(client_dir, "Sem_Pendencias_Fiscais_Regular-*.txt"))
                    relatorio_pdfs = glob.glob(os.path.join(client_dir, "RelatorioSituacaoFiscal-*.pdf"))
                    
                    possui_registro_existente = len(cnd_pdfs) > 0 or len(cnd_txts) > 0 or len(relatorio_pdfs) > 0
                    
                    if not forcar_todos and (cnd_pdfs or cnd_txts):
                        log(f"Cliente {nome} ({cnpj}) já possui Certidão/Regularidade ativa na pasta do cliente. Pulando consulta no e-CAC.", "SUCCESS")
                        detalhes_status = "Certidão Baixada" if cnd_pdfs else "Sem Pendências (Informativo Gravado)"
                        save_client_status(config["relatorios_dir"], cnpj, nome, "Sucesso", detalhes_status)
                        success_count += 1
                        # Atualiza a planilha de forma dinâmica
                        atualizar_excel_dinamico(clientes, config)
                        continue
                        
                    # 2. Verificar se o cliente já foi consultado com SUCESSO hoje de fato
                    ja_processado_hoje = False
                    if not forcar_todos and os.path.exists(status_path):
                        try:
                            with open(status_path, "r", encoding="utf-8") as f:
                                status_data = json.load(f)
                                if status_data.get("status") == "Sucesso" and status_data.get("data_consulta") == today_str:
                                    ja_processado_hoje = True
                        except Exception:
                            pass
                            
                    if ja_processado_hoje:
                        log(f"Cliente {nome} ({cnpj}) já foi consultado com SUCESSO hoje. Pulando...", "SUCCESS")
                        success_count += 1
                        # Atualiza a planilha de forma dinâmica
                        atualizar_excel_dinamico(clientes, config)
                        continue
                        
                    # 2.5. Verificar se o cliente já falhou hoje
                    ja_falhou_hoje = False
                    if not forcar_todos and os.path.exists(status_path):
                        try:
                            with open(status_path, "r", encoding="utf-8") as f:
                                status_data = json.load(f)
                                if status_data.get("data_consulta") == today_str:
                                    if status_data.get("status") == "Erro" and status_data.get("contador_falhas_hoje", 0) >= 1:
                                        ja_falhou_hoje = True
                        except Exception:
                            pass
                            
                    if ja_falhou_hoje:
                        log(f"Cliente {nome} ({cnpj}) já falhou hoje. Pulando processamento para evitar travamento...", "WARNING")
                        failure_count += 1
                        # Atualiza a planilha de forma dinâmica
                        atualizar_excel_dinamico(clientes, config)
                        continue
                        
                    # 3. Verificar se já existe relatório de pendência na pasta do cliente
                    relatorio_pdfs = glob.glob(os.path.join(client_dir, "RelatorioSituacaoFiscal-*.pdf"))
                    tem_relatorio_existente = len(relatorio_pdfs) > 0
                    
                    log(f"Processando Cliente: {nome} ({cnpj})...", "INFO")
                    if tem_relatorio_existente:
                        log(f"Identificado relatório de pendências pré-existente na pasta do cliente. Caso continue irregular, o download será pulado.", "INFO")
                    
                    # Criar diretório para salvar o relatório do cliente (já inicializa como pendente)
                    client_dir = save_client_status(config["relatorios_dir"], cnpj, nome, "Pendente", "Iniciado")
                    enviar_whatsapp(f"🤖 *Processando*: {nome} ({cnpj})", config)
                    
                    # Limpar todos os arquivos e pastas se for varredura completa. Se não, mantemos para caso de erro.
                    if forcar_todos:
                        remover_arquivos_fiscais_obsoletos(client_dir)
                        
                    max_tentativas = 1
                    sucesso_cliente = False
                    erro_final = None
                    url_erro = "N/A"
                    
                    for tentativa in range(1, max_tentativas + 1):
                        try:
                            # Se for uma tentativa maior que 1, ou se o navegador foi fechado/marcado para reiniciar
                            if tentativa > 1 or context is None or page is None or need_browser_restart:
                                need_browser_restart = True # Força o atraso de 1 minuto em iniciar_navegador
                                iniciar_navegador()
                                
                            # 0.5. Limpar modais iniciais ou remanescentes e tratar Caixa Postal bloqueante
                            fechar_modais_e_overlays(page)
                            checar_e_tratar_caixa_postal(page, client_dir, config)
                            
                            # 1. Verificar qual é o perfil atualmente ativo na tela do e-CAC
                            header_text = page.locator("body").inner_text()
                            header_text_limpo = header_text.replace(".", "").replace("/", "").replace("-", "").replace(" ", "").replace("\n", "").replace("\r", "")
                            
                            cnpj_sem_formatacao = cnpj.replace(".", "").replace("/", "").replace("-", "")
                            
                            # Caso A: O cliente que queremos consultar é o próprio procurador (JEJ)
                            if procurador_cnpj and cnpj == procurador_cnpj:
                                if "Procuradorde" in header_text_limpo:
                                    log(f"Perfil atual não é Titular. Solicitando alteração de perfil para voltar ao Titular ({nome})...", "INFO")
                                    alterar_perfil(page, cnpj, procurador_cnpj)
                                    checar_e_tratar_caixa_postal(page, client_dir, config)
                                else:
                                    log(f"Já estamos no perfil Titular do procurador ({nome}). Consultando diretamente.", "INFO")
                                    
                            # Caso B: O cliente que queremos consultar é uma empresa representada (Tome & Lopes)
                            else:
                                if cnpj_sem_formatacao in header_text_limpo:
                                    log(f"O perfil ativo no e-CAC já corresponde a {nome} ({cnpj}). Pulando alteração de perfil.", "SUCCESS")
                                else:
                                    log(f"Perfil atual diferente de {nome}. Alterando perfil para o CNPJ: {cnpj}...", "INFO")
                                    alterar_perfil(page, cnpj, procurador_cnpj)
                                    checar_e_tratar_caixa_postal(page, client_dir, config)
                                
                            # Acessar Situação Fiscal e baixar relatório
                            resultado = baixar_relatorio_situacao_fiscal(page, context, client_dir, cnpj, config, ja_possui_relatorio=tem_relatorio_existente)
                            
                            # Salvar metadados de sucesso
                            save_client_status(config["relatorios_dir"], cnpj, nome, "Sucesso", resultado)
                            enviar_whatsapp(f"✅ *Sucesso*: {nome} - {resultado}", config)
                            success_count += 1
                            sucesso_cliente = True
                            break
                        except JapeException as e:
                            log(f"[ERRO JAPE] Erro temporário e-CAC (107.6) para {nome} ({cnpj}): {e}", "WARNING")
                            erro_final = e
                            if page:
                                try:
                                    url_erro = page.url
                                except Exception:
                                    pass
                            
                            # Fechar abas extras se acumularam
                            try:
                                if context and len(context.pages) > 1:
                                    for p_extra in context.pages[1:]:
                                        p_extra.close()
                            except Exception:
                                pass
                                
                            # Redirecionar página principal de volta para a Home do e-CAC
                            try:
                                log("Redirecionando a página principal de volta para a Home do e-CAC...", "INFO")
                                page.goto(config["portal_url"], timeout=20000)
                                page.wait_for_load_state("load")
                                # Esperar até que o botão "Alterar perfil de acesso" esteja visível
                                page.locator("text=Alterar perfil de acesso").first.wait_for(state="visible", timeout=10000)
                                log("Página principal retornada com sucesso para a Home do e-CAC (botão Alterar perfil visível).", "SUCCESS")
                            except Exception as nav_err:
                                log(f"Aviso: Falha ao redirecionar para a Home após erro JAPE: {nav_err}. A automação tentará prosseguir.", "WARNING")

                            sucesso_cliente = False
                            break
                            
                        except Exception as e:
                            log(f"[ERRO - TENTATIVA {tentativa}/{max_tentativas}] Erro ao processar cliente {nome} ({cnpj}): {e}", "WARNING")
                            erro_final = e
                            if page:
                                try:
                                    url_erro = page.url
                                except Exception:
                                    pass
                            
                            # Enriquecer o erro com textos visíveis na tela (mensagens de erro do e-CAC/Gov.br)
                            msg_tela = ""
                            if page:
                                try:
                                    body_text = page.locator("body").inner_text()
                                    linhas_erro = []
                                    for line in body_text.split("\n"):
                                        line_s = line.strip()
                                        line_l = line_s.lower()
                                        if not line_s or len(line_s) < 5:
                                            continue
                                        # Captura avisos, mensagens de erro do gov.br ou do e-cac
                                        if any(x in line_l for x in ["atenção", "atencao", "erro", "restrição", "restricao", "inexistente", "cancelada", "expirada", "automatizado", "indisponível", "falha", "não cadastrada", "revogada"]):
                                            if len(line_s) < 250: # Evita blocos excessivamente longos
                                                linhas_erro.append(line_s)
                                    
                                    if linhas_erro:
                                        msg_tela = " | ".join(linhas_erro[:3])
                                        erro_final = Exception(f"{e} (Erro na tela: {msg_tela})")
                                except Exception:
                                    pass
                            
                            # Capturar screenshot do erro para diagnóstico visual na pasta do cliente
                            try:
                                screenshot_path = os.path.join(client_dir, f"erro_tentativa_{tentativa}.png")
                                page.screenshot(path=screenshot_path)
                                log(f"Screenshot do erro capturado e salvo em: {screenshot_path}", "INFO")
                                remover_arquivos_fiscais_obsoletos(client_dir, screenshot_path)
                            except Exception as snap_err:
                                log(f"Não foi possível capturar screenshot do erro: {snap_err}", "WARNING")
                                
                            # Fechar abas extras se acumularam
                            try:
                                if context and len(context.pages) > 1:
                                    for p_extra in context.pages[1:]:
                                        p_extra.close()
                            except Exception:
                                pass
                                
                            # Checagem de expurgo de cookies bloqueados / sessões inválidas
                            erro_detalhado_l = (str(erro_final) + " " + msg_tela).lower()
                            if any(palavra in erro_detalhado_l for palavra in ["automatizado", "bloqueio", "bloqueado", "expirada", "expirado", "invalido", "inválido", "sessão", "sessao", "desconectado"]):
                                log("[RECOVERY] Bloqueio ou expiração de sessão detectado! Expurando state.json...", "WARNING")
                                if os.path.exists(state_file):
                                    try:
                                        os.remove(state_file)
                                        log("Arquivo de sessão state.json apagado para renovação.", "SUCCESS")
                                    except Exception as rm_err:
                                        log(f"Não foi possível remover state.json: {rm_err}", "WARNING")
                                
                            # Se der erro em algum processo, fechar o e-CAC (sair) para entrar novamente do início na próxima tentativa/cliente
                            log("Realizando logout e fechando navegador do e-CAC devido a erro de execução...", "INFO")
                            need_browser_restart = True
                            if context:
                                try:
                                    tentar_logout_ecac(page)
                                except Exception:
                                    pass
                                try:
                                    context.close()
                                except Exception:
                                    pass
                                context = None
                                page = None
                                
                            # Verificar se é um erro permanente de procuração
                            erro_str = str(e).lower()
                            if "procuração" in erro_str or "procuracao" in erro_str or "inexistente" in erro_str or "cancelada" in erro_str or "expirada" in erro_str:
                                log(f"[ERRO PERMANENTE] Impedimento de procuração definitivo para {nome} ({cnpj}): {e}. Pulando tentativas adicionais.", "ERROR")
                                break
                                
                    if not sucesso_cliente:
                        # Se não for varredura forçada completa e já tivermos um registro/PDF salvo na pasta,
                        # nós pulamos o erro, restauramos o status de Sucesso e preservamos os arquivos!
                        if not forcar_todos and possui_registro_existente:
                            log(f"[RECOVERY] Falha ao processar {nome} ({cnpj}), mas mantendo o registro existente de CND/Relatório e ignorando o erro de consulta.", "SUCCESS")
                            
                            status_restaurado = "Sucesso"
                            detalhes_restaurados = "Preservado (Erro na nova consulta)"
                            if old_status and old_status.get("status") == "Sucesso":
                                detalhes_restaurados = old_status.get("detalhes", detalhes_restaurados)
                            
                            try:
                                save_client_status(config["relatorios_dir"], cnpj, nome, status_restaurado, detalhes_restaurados)
                            except Exception as save_err:
                                log(f"Não foi possível restaurar status de sucesso: {save_err}", "WARNING")
                            success_count += 1
                        else:
                            log(f"[FALHA FINAL] Não foi possível processar o cliente {nome} ({cnpj}) após {max_tentativas} tentativas.", "ERROR")
                            # Salvar log de erro definitivo
                            try:
                                save_client_status(config["relatorios_dir"], cnpj, nome, "Erro", str(erro_final))
                                enviar_whatsapp(f"❌ *Falha*: {nome} - {str(erro_final)}\n• *Página do Erro*: {url_erro}", config)
                            except Exception as save_err:
                                log(f"Não foi possível salvar status de erro para o cliente: {save_err}", "WARNING")
                            failure_count += 1
                    
                    # Atualiza a planilha Excel de forma dinâmica após processar o cliente
                    try:
                        atualizar_excel_dinamico(clientes, config)
                        log(f"Painel Excel consolidado atualizado dinamicamente para o cliente {nome}.", "SUCCESS")
                    except Exception as xls_err:
                        log(f"Aviso ao atualizar painel Excel dinamicamente: {xls_err}", "WARNING")
                        
                    # Atraso regulamentar de 5 segundos entre execuções de clientes (se não houve falha que reiniciou o navegador)
                    if not need_browser_restart:
                        log("Aguardando 5 segundos de intervalo regulamentar antes de prosseguir para o próximo cliente...", "INFO")
                        time.sleep(5)
                            
            if context:
                try:
                    tentar_logout_ecac(page)
                except Exception:
                    pass
                try:
                    context.close()
                except Exception:
                    pass
                
        if need_restart:
            # Evita loops infinitos de login manual se o argumento --login estiver na linha de comando
            if "--login" in sys.argv:
                sys.argv.remove("--login")
            log("Sessão limpa e resetada. Reiniciando fluxo de login manual...", "INFO")
            continue
        else:
            break
            
    # Garantir uma atualização final consolidada
    atualizar_excel_dinamico(clientes, config)
    
    # Relatório final
    log("=" * 60, "SUMMARY")
    log(" ROTINA DE EXECUÇÃO CONCLUÍDA", "SUMMARY")
    log(f" - Clientes com Sucesso: {success_count}", "SUMMARY")
    log(f" - Clientes com Falha: {failure_count}", "SUMMARY")
    log(f" - Clientes Ignorados: {skipped_count}", "SUMMARY")
    log("=" * 60, "SUMMARY")

    # Enviar notificação via WhatsApp (Z-API) se estiver ativo
    if config.get("whatsapp_enabled"):
        try:
            hoje_fmt = datetime.datetime.now().strftime("%d/%m/%Y às %H:%M")
            mensagem = (
                f"📢 *Escavador de Pendências e-CAC*\n"
                f"📅 Varredura concluída em {hoje_fmt}\n\n"
                f"✅ *Sucesso:* {success_count} empresas\n"
                f"❌ *Falha:* {failure_count} empresas\n"
                f"➖ *Ignorado:* {skipped_count} empresas\n\n"
            )
            
            # Detalhar pendências encontradas (relatórios baixados de hoje)
            detalhes_pendencias = []
            detalhes_erros = []
            today_str = datetime.date.today().strftime("%Y-%m-%d")
            
            for cliente in clientes:
                cnpj = cliente["cnpj"]
                nome = cliente["nome"]
                ativo = cliente["ativo"]
                if not ativo:
                    continue
                
                nome_limpo = clean_filename(nome)
                cnpj_limpo = "".join(filter(str.isdigit, cnpj))
                status_path = os.path.join(config["relatorios_dir"], f"{cnpj_limpo}_{nome_limpo}", "status.json")
                if os.path.exists(status_path):
                    with open(status_path, "r", encoding="utf-8") as f:
                        data = json.load(f)
                        if data.get("data_consulta") == today_str:
                            status = data.get("status")
                            detalhes = data.get("detalhes")
                            if status == "Sucesso" and "Relatório" in detalhes:
                                detalhes_pendencias.append(nome)
                            elif status == "Erro":
                                detalhes_erros.append(f"{nome} ({detalhes})")
            
            if detalhes_pendencias:
                mensagem += "⚠️ *Empresas com Pendências (Relatório PDF Baixado):*\n"
                for p in detalhes_pendencias:
                    mensagem += f" - {p}\n"
                mensagem += "\n"
                
            if detalhes_erros:
                mensagem += "🚨 *Empresas com Falha no Acesso:*\n"
                for e in detalhes_erros:
                    mensagem += f" - {e}\n"
                mensagem += "\n"
                
            mensagem += "📂 *O Painel de Controle Consolidado atualizado está disponível no seu Desktop!*"
            enviar_whatsapp(mensagem, config)
        except Exception as wa_err:
            log(f"Erro ao compilar e enviar notificação do WhatsApp: {wa_err}", "ERROR")

if __name__ == "__main__":
    main()
