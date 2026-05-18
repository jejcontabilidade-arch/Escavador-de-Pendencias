import os
import sys

def main():
    excel_path = r"C:\Users\jejco\Desktop\05_BLOCO_E_SQS_309_CNPJ_03495231000101.xlsx"
    
    if not os.path.exists(excel_path):
        print(f"Error: Example Excel not found at {excel_path}")
        return
        
    try:
        import openpyxl
    except ImportError:
        print("Installing openpyxl...")
        import subprocess
        subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl"])
        import openpyxl
        
    print(f"Opening Excel: {excel_path}")
    wb = openpyxl.load_workbook(excel_path, data_only=False)
    
    print(f"Sheet names: {wb.sheetnames}")
    for name in wb.sheetnames:
        sheet = wb[name]
        print(f"\n--- Sheet: {name} ---")
        print(f"Dimensions: {sheet.dimensions}")
        
        # Read the first 10 rows and 10 columns
        max_r = min(sheet.max_row, 30)
        max_c = min(sheet.max_column, 15)
        
        for r in range(1, max_r + 1):
            row_vals = []
            row_styles = []
            for c in range(1, max_c + 1):
                cell = sheet.cell(row=r, column=c)
                val = cell.value
                val_str = f"'{val}'" if val is not None else "None"
                
                # Extract style info if cell is not empty
                style_info = ""
                if cell.fill and cell.fill.fill_type:
                    fg = cell.fill.start_color.rgb if cell.fill.start_color else "None"
                    bg = cell.fill.end_color.rgb if cell.fill.end_color else "None"
                    if fg != "00000000":
                        style_info += f" Color:{fg}"
                if cell.font:
                    if cell.font.bold:
                        style_info += " Bold"
                    if cell.font.size:
                        style_info += f" Size:{cell.font.size}"
                    if cell.font.color and cell.font.color.rgb:
                        style_info += f" FontColor:{cell.font.color.rgb}"
                        
                row_vals.append(val_str)
                if style_info:
                    row_styles.append(f"C{c}:{style_info.strip()}")
                    
            print(f"Row {r:02d}: {', '.join(row_vals[:8])}")
            if row_styles:
                print(f"  Styles -> {', '.join(row_styles)}")

if __name__ == "__main__":
    main()
