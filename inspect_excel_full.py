import os
import sys
import openpyxl

def main():
    excel_path = r"C:\Users\jejco\Desktop\05_BLOCO_E_SQS_309_CNPJ_03495231000101.xlsx"
    out_path = r"C:\Users\jejco\Desktop\Escavador de Pendencias\excel_details.txt"
    
    if not os.path.exists(excel_path):
        with open(out_path, "w") as f:
            f.write(f"Error: File not found at {excel_path}\n")
        return
        
    wb = openpyxl.load_workbook(excel_path)
    
    with open(out_path, "w", encoding="utf-8") as f:
        f.write(f"Workbook Sheets: {wb.sheetnames}\n\n")
        
        for name in wb.sheetnames:
            sheet = wb[name]
            f.write(f"=== SHEET: {name} ===\n")
            f.write(f"Dimensions: {sheet.dimensions}\n")
            f.write(f"Max Row: {sheet.max_row}, Max Col: {sheet.max_column}\n\n")
            
            # Print the first 25 rows
            for r in range(1, min(sheet.max_row + 1, 40)):
                row_vals = []
                styles = []
                for c in range(1, min(sheet.max_column + 1, 15)):
                    cell = sheet.cell(row=r, column=c)
                    val = cell.value
                    row_vals.append(str(val) if val is not None else "")
                    
                    style = []
                    if cell.fill and cell.fill.fill_type:
                        fg = cell.fill.start_color.rgb
                        if fg and fg != "00000000":
                            style.append(f"BgColor:{fg}")
                    if cell.font:
                        font_style = []
                        if cell.font.bold:
                            font_style.append("Bold")
                        if cell.font.color and cell.font.color.rgb:
                            font_style.append(f"Color:{cell.font.color.rgb}")
                        if cell.font.size:
                            font_style.append(f"Size:{cell.font.size}")
                        if cell.font.name:
                            font_style.append(f"Name:{cell.font.name}")
                        if font_style:
                            style.append("Font(" + " ".join(font_style) + ")")
                    if cell.alignment:
                        align_style = []
                        if cell.alignment.horizontal:
                            align_style.append(f"H:{cell.alignment.horizontal}")
                        if cell.alignment.vertical:
                            align_style.append(f"V:{cell.alignment.vertical}")
                        if align_style:
                            style.append("Align(" + " ".join(align_style) + ")")
                    if style:
                        styles.append(f"Col{c}: {', '.join(style)}")
                        
                f.write(f"Row {r:02d}: {row_vals}\n")
                if styles:
                    f.write(f"   Styles: {'; '.join(styles)}\n")
            f.write("\n\n")

if __name__ == "__main__":
    main()
